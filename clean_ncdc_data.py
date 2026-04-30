"""
NCDC (TASKA) Data Cleaning Script
========================================
Cleans Status Kesihatan Kanak-Kanak Kebangsaan (NCDC) data and generates:
  1. NCDC_Cleaned.csv           — Cleansed data with WHO z-scores & indicators
  2. NCDC_QualityReport.xlsx    — Data quality report

Source: Multi-sheet Excel file with years 2023, 2024, 2025 measurements per child

Cleaning rules applied (drop invalid rows):
  Rule 1  : Drop Jantina not in (Lelaki, Perempuan) → mapped to Male/Female
  Rule 2  : Drop biologically implausible weight (≤0.5 or >35 kg) / height (≤30 or >130 cm)
  Rule 3  : Drop age at assessment > 5 years (≥60 months) or < 0
  Rule 4  : Drop records where measurement date < DOB (assessment before birth)
  Rule 5  : Drop implausible BMI > 40
  Rule 6  : Drop records where both weight AND height are null
  Rule 7  : Drop records where ANY z-score is null (ensures fully clean output)

WHO Indicators computed:
  WAZ  — Weight-for-Age Z-score  → Kurang Berat Badan (WAZ < -2SD)
  HAZ  — Height-for-Age Z-score  → Bantut (HAZ < -2SD)
  BAZ  — BMI-for-Age Z-score     → Susut (BAZ < -2SD)
  Ind_Normal — TRUE when child is healthy (all above indicators FALSE)

Standardization:
  - Gender: Male / Female (consistent with KPM/MyVASS)
  - Dates: YYYY-MM-DD format
  - Status labels: Malay (Kurang Berat Badan Teruk, Bantut, Susut, etc.)
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path for WHO z-score module ─────────────────────────────
_BACKEND = Path(__file__).parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore import compute_zscore
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height bounds (biologically plausible for 0–5 years)
BERAT_MIN, BERAT_MAX = 0.5, 35.0      # kg
TINGGI_MIN, TINGGI_MAX = 30.0, 130.0  # cm
BMI_MAX = 40.0

# Age bounds
AGE_MAX_MONTHS = 60  # under 5 years

# WHO BIV thresholds
BIV = {
    "WAZ": (-6.0, +5.0),
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}
_BIV_LO_WAZ, _BIV_HI_WAZ = BIV["WAZ"]
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]

# Gender mapping (standardize to English)
GENDER_MAP = {
    "LELAKI": "Male",
    "PEREMPUAN": "Female",
    "L": "Male",
    "P": "Female",
    "MALE": "Male",
    "FEMALE": "Female",
}

# ---------------------------------------------------------------------------
# Label functions
# ---------------------------------------------------------------------------

def _waz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Kurang Berat Badan Teruk"
    if z < -2:  return "Kurang Berat Badan"
    if z <= 2:  return "Berat Badan Normal"
    return "Mungkin Masalah Pertumbuhan"


def _haz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Bantut Teruk"
    if z < -2:  return "Bantut"
    if z <= 3:  return "Panjang/Tinggi Normal"
    return "Mungkin Masalah Endokrin"


def _baz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Susut Teruk"
    if z < -2:  return "Susut"
    if z <= 1:  return "Berat Badan Normal"
    if z <= 2:  return "Berisiko Berlebihan Berat Badan"
    if z <= 3:  return "Berlebihan Berat Badan"
    return "Obes"


# ---------------------------------------------------------------------------
# Load and reshape NCDC data
# ---------------------------------------------------------------------------

def load_ncdc_data(file_path: Path) -> pd.DataFrame:
    """
    Load NCDC Excel file (multi-sheet, wide format) and reshape to long format.
    Each row = one measurement (child × year).
    """
    xl = pd.ExcelFile(file_path)
    all_frames = []

    for sheet_name in xl.sheet_names:
        print(f"    Loading sheet: {sheet_name} ...")
        df_sheet = xl.parse(sheet_name)
        if df_sheet.empty:
            continue

        # Standardize column names
        df_sheet.columns = df_sheet.columns.str.strip()

        # Extract base info columns
        base_cols = [
            "Agensi", "Negeri", "Daerah", "Nama TASKA", "Nama Anak",
            "No. Mykid", "Pendapatan Keluarga", "Kumpulan Umur",
            "Tarikh Lahir", "Jantina"
        ]
        base_cols_present = [c for c in base_cols if c in df_sheet.columns]

        # Melt years 2023, 2024, 2025 into long format
        for year in [2023, 2024, 2025]:
            year_cols = {
                "Tarikh_Pengukuran": f"{year} Tarikh Pengukuran",
                "Berat_kg": f"{year} Berat (kg)",
                "Tinggi_cm": f"{year} Tinggi (cm)",
                "BMI_Original": f"{year} BMI",
                "Status_Berat_Original": f"{year} Status Berat",
                "Status_Tinggi_Original": f"{year} Status Tinggi",
                "Status_BMI_Original": f"{year} Status BMI",
            }

            # Check if year data exists
            if year_cols["Berat_kg"] not in df_sheet.columns:
                continue

            # Build row for this year
            df_year = df_sheet[base_cols_present].copy()
            df_year["Year"] = year

            for new_col, src_col in year_cols.items():
                if src_col in df_sheet.columns:
                    df_year[new_col] = df_sheet[src_col]
                else:
                    df_year[new_col] = np.nan

            all_frames.append(df_year)

    if not all_frames:
        raise ValueError("No data found in NCDC file")

    df_long = pd.concat(all_frames, ignore_index=True)
    print(f"    Total rows after reshape: {len(df_long):,}")
    return df_long


# ---------------------------------------------------------------------------
# Cleaning function
# ---------------------------------------------------------------------------

def clean_ncdc(raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Clean NCDC data and compute WHO z-scores."""
    stats: dict = {"raw_count": len(raw)}
    df = raw.copy()

    # ── Rename columns ────────────────────────────────────────────────────
    df = df.rename(columns={
        "Nama TASKA": "Nama_Taska",
        "Nama Anak": "Nama_Anak",
        "No. Mykid": "MyKid",
        "Pendapatan Keluarga": "Pendapatan_Keluarga",
        "Kumpulan Umur": "Kumpulan_Umur_Original",
        "Tarikh Lahir": "Tarikh_Lahir",
    })

    # ── Rule 1: Standardize gender, drop invalid ──────────────────────────
    df["Jantina_Raw"] = df["Jantina"].astype(str).str.strip().str.upper()
    df["Gender"] = df["Jantina_Raw"].map(GENDER_MAP)

    before = len(df)
    invalid_gender = df["Gender"].isna()
    stats["dropped_invalid_gender"] = int(invalid_gender.sum())
    df = df[~invalid_gender].copy()

    # Also store Jantina in Malay for compatibility
    df["Jantina"] = df["Gender"].map({"Male": "Lelaki", "Female": "Perempuan"})

    # ── Parse dates ───────────────────────────────────────────────────────
    df["Tarikh_Lahir"] = pd.to_datetime(df["Tarikh_Lahir"], errors="coerce")
    df["Tarikh_Pengukuran"] = pd.to_datetime(df["Tarikh_Pengukuran"], errors="coerce")

    # Format as date only (remove time component)
    df["Tarikh_Lahir"] = df["Tarikh_Lahir"].dt.normalize()
    df["Tarikh_Pengukuran"] = df["Tarikh_Pengukuran"].dt.normalize()

    stats["null_tarikh_lahir"] = int(df["Tarikh_Lahir"].isna().sum())
    stats["null_tarikh_pengukuran"] = int(df["Tarikh_Pengukuran"].isna().sum())

    # ── Rule 4: Drop assessment before DOB ────────────────────────────────
    before = len(df)
    bad_date_mask = (
        df["Tarikh_Lahir"].notna() &
        df["Tarikh_Pengukuran"].notna() &
        (df["Tarikh_Pengukuran"] < df["Tarikh_Lahir"])
    )
    stats["dropped_assessment_before_dob"] = int(bad_date_mask.sum())
    df = df[~bad_date_mask].copy()

    # ── Compute age at measurement (months) ───────────────────────────────
    has_both = df["Tarikh_Lahir"].notna() & df["Tarikh_Pengukuran"].notna()
    df["Age_Months"] = np.where(
        has_both,
        (df["Tarikh_Pengukuran"] - df["Tarikh_Lahir"]).dt.days / 30.4375,
        np.nan,
    )
    df["Age_Months"] = df["Age_Months"].round(2)

    # ── Rule 3: Drop age > 5 years or negative ────────────────────────────
    before = len(df)
    age_invalid = df["Age_Months"].notna() & (
        (df["Age_Months"] >= AGE_MAX_MONTHS) | (df["Age_Months"] < 0)
    )
    stats["dropped_age_invalid"] = int(age_invalid.sum())
    df = df[~age_invalid].copy()

    # ── Numeric measurements ──────────────────────────────────────────────
    df["Berat_kg"] = pd.to_numeric(df["Berat_kg"], errors="coerce")
    df["Tinggi_cm"] = pd.to_numeric(df["Tinggi_cm"], errors="coerce")

    # ── Rule 2: Drop measurement outliers ─────────────────────────────────
    berat_bad = (df["Berat_kg"] <= BERAT_MIN) | (df["Berat_kg"] > BERAT_MAX)
    tinggi_bad = (df["Tinggi_cm"] <= TINGGI_MIN) | (df["Tinggi_cm"] > TINGGI_MAX)

    before = len(df)
    outlier_mask = (berat_bad & df["Berat_kg"].notna()) | (tinggi_bad & df["Tinggi_cm"].notna())
    stats["dropped_measurement_outlier"] = int(outlier_mask.sum())
    df = df[~outlier_mask].copy()

    df["Berat_kg"] = df["Berat_kg"].round(2)
    df["Tinggi_cm"] = df["Tinggi_cm"].round(1)

    # ── Rule 6: Drop rows where both measurements are null ────────────────
    before = len(df)
    no_meas_mask = df["Berat_kg"].isna() & df["Tinggi_cm"].isna()
    stats["dropped_no_measurement"] = int(no_meas_mask.sum())
    df = df[~no_meas_mask].copy()

    # ── Compute BMI ───────────────────────────────────────────────────────
    valid_both = df["Berat_kg"].notna() & df["Tinggi_cm"].notna() & (df["Tinggi_cm"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["Berat_kg"] / ((df["Tinggi_cm"] / 100) ** 2)).round(2),
        np.nan,
    )

    # ── Rule 5: Drop implausible BMI > 40 ─────────────────────────────────
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()

    # ── Kumpulan Umur ─────────────────────────────────────────────────────
    def _age_group(m):
        if pd.isna(m): return None
        if m < 24:     return "Bawah 2 Tahun"
        if m < 60:     return "Bawah 5 Tahun"
        return None
    df["Kumpulan_Umur"] = df["Age_Months"].apply(_age_group)

    # ── WHO Z-scores ──────────────────────────────────────────────────────
    if _ZSCORE_AVAILABLE:
        print("    Computing WHO z-scores ...")
        sex_s = df["Gender"].astype(str)
        age_s = df["Age_Months"]
        berat_s = df["Berat_kg"]
        tinggi_s = df["Tinggi_cm"]
        bmi_s = df["BMI"]

        n = len(df)
        waz_vals, haz_vals, baz_vals = [], [], []

        for i in range(n):
            age = age_s.iat[i]
            sex = sex_s.iat[i]
            w = berat_s.iat[i]
            h = tinggi_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age) or pd.isna(sex):
                waz_vals.append(None)
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            waz = compute_zscore(float(w), sex, float(age), "WAZ") if pd.notna(w) else None
            haz = compute_zscore(float(h), sex, float(age), "HAZ") if pd.notna(h) else None
            baz = compute_zscore(float(bmi), sex, float(age), "BAZ") if pd.notna(bmi) else None

            # Apply BIV
            if waz is not None and not (_BIV_LO_WAZ <= waz <= _BIV_HI_WAZ): waz = None
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            waz_vals.append(round(waz, 3) if waz is not None else None)
            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 50000 == 0:
                print(f"      Z-scores: {i+1:,}/{n:,} ...")

        df["WAZ"] = waz_vals
        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["WAZ_Status"] = df["WAZ"].apply(_waz_label)
        df["HAZ_Status"] = df["HAZ"].apply(_haz_label)
        df["BAZ_Status"] = df["BAZ"].apply(_baz_label)

        # Indicator flags
        df["Ind_Kurang_Berat_Badan"] = df["WAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Bantut"] = df["HAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Susut"] = df["BAZ"].apply(lambda z: z is not None and z < -2)

        # ── Rule 7: Drop rows where ANY z-score is null ───────────────────
        before = len(df)
        null_zscore_mask = df["WAZ"].isna() | df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore_mask.sum())
        df = df[~null_zscore_mask].copy()
        print(f"    Rule 7: Dropped {stats['dropped_null_zscore']:,} rows with null z-scores")

        # Ind_Normal
        df["Ind_Normal"] = ~(df["Ind_Kurang_Berat_Badan"] | df["Ind_Bantut"] | df["Ind_Susut"])

    else:
        for col in ["WAZ", "HAZ", "BAZ", "WAZ_Status", "HAZ_Status", "BAZ_Status",
                    "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["ind_kurang_berat"] = int(df["Ind_Kurang_Berat_Badan"].sum())
        stats["ind_bantut"] = int(df["Ind_Bantut"].sum())
        stats["ind_susut"] = int(df["Ind_Susut"].sum())
        stats["ind_normal"] = int(df["Ind_Normal"].sum())

    return df, stats


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr


def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(stats: dict, df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "NCDC (TASKA) — DATA QUALITY REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    _hdr(ws1, 4, ["Metric", "Count", "% of Raw", "Notes"])
    total = stats["raw_count"]

    summary_rows = [
        ("Raw records (before cleaning)", stats["raw_count"], "100%", "After reshape to long format"),
        ("Dropped — Invalid Gender", stats.get("dropped_invalid_gender", 0),
         f"{stats.get('dropped_invalid_gender',0)/total*100:.2f}%", "Rule 1"),
        ("Dropped — Assessment before DOB", stats.get("dropped_assessment_before_dob", 0),
         f"{stats.get('dropped_assessment_before_dob',0)/total*100:.2f}%", "Rule 4"),
        ("Dropped — Age invalid", stats.get("dropped_age_invalid", 0),
         f"{stats.get('dropped_age_invalid',0)/total*100:.2f}%", "Rule 3"),
        ("Dropped — Measurement outliers", stats.get("dropped_measurement_outlier", 0),
         f"{stats.get('dropped_measurement_outlier',0)/total*100:.2f}%", "Rule 2"),
        ("Dropped — BMI > 40", stats.get("dropped_bmi_outlier", 0),
         f"{stats.get('dropped_bmi_outlier',0)/total*100:.2f}%", "Rule 5"),
        ("Dropped — No measurement", stats.get("dropped_no_measurement", 0),
         f"{stats.get('dropped_no_measurement',0)/total*100:.2f}%", "Rule 6"),
        ("Dropped — Null z-score(s)", stats.get("dropped_null_zscore", 0),
         f"{stats.get('dropped_null_zscore',0)/total*100:.2f}%", "Rule 7"),
        ("Final cleaned records", stats["final_count"],
         f"{stats['final_count']/total*100:.2f}%", "After all rules"),
    ]
    for i, r in enumerate(summary_rows, 5):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(r, 1):
            _cell(ws1, i, c, v, bg=bg)

    row = len(summary_rows) + 6

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["WHO Indicator", "Positive Count", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        ind_rows = [
            ("Kurang Berat Badan (WAZ < -2SD)", stats.get("ind_kurang_berat", 0),
             f"{stats.get('ind_kurang_berat',0)/max(stats['final_count'],1)*100:.1f}%", "WAZ < -2"),
            ("Bantut (HAZ < -2SD)", stats.get("ind_bantut", 0),
             f"{stats.get('ind_bantut',0)/max(stats['final_count'],1)*100:.1f}%", "HAZ < -2"),
            ("Susut (BAZ < -2SD)", stats.get("ind_susut", 0),
             f"{stats.get('ind_susut',0)/max(stats['final_count'],1)*100:.1f}%", "BAZ < -2"),
            ("Normal (healthy child)", stats.get("ind_normal", 0),
             f"{stats.get('ind_normal',0)/max(stats['final_count'],1)*100:.1f}%", "All indicators FALSE"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

    _auto_width(ws1)

    # ── Sheet 2: By Year ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Year")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "DISTRIBUTION BY YEAR"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _hdr(ws2, 3, ["Year", "Total", "Kurang Berat Badan", "Bantut", "Susut", "Normal"])

    if "Year" in df.columns:
        grp = df.groupby("Year")
        for i, (year, sub) in enumerate(sorted(grp), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                int(year), len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws2, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws2)

    # ── Sheet 3: By Negeri ────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Negeri")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:F1")
    ws3["A1"].value = "DISTRIBUTION BY NEGERI"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _hdr(ws3, 3, ["Negeri", "Total", "Kurang Berat Badan", "Bantut", "Susut", "Normal"])

    if "Negeri" in df.columns:
        grp = df.groupby("Negeri")
        for i, (neg, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                neg, len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws3, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws3)

    return wb


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def save_cleaned_csv(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()

    # Format dates as YYYY-MM-DD
    for col in ["Tarikh_Lahir", "Tarikh_Pengukuran"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Booleans as strings
    for col in ["Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)

    # Select output columns in order
    out_cols = [
        "Year", "Negeri", "Daerah", "Nama_Taska", "Agensi",
        "MyKid", "Gender", "Jantina",
        "Tarikh_Lahir", "Tarikh_Pengukuran", "Age_Months", "Kumpulan_Umur",
        "Berat_kg", "Tinggi_cm", "BMI",
        "WAZ", "HAZ", "BAZ",
        "WAZ_Status", "HAZ_Status", "BAZ_Status",
        "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal",
    ]
    out_cols = [c for c in out_cols if c in df_out.columns]
    df_out = df_out[out_cols]

    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# File picker helpers
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=initial_dir,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  NCDC (TASKA) Data Cleaning Tool")
    print("=" * 60)

    # Step 1: Pick input file
    print("\nStep 1: Select the NCDC Excel file")
    ncdc_path = _pick_file("Select NCDC Excel file", str(BASE_DIR))
    if not ncdc_path:
        print("  ✗ No file selected. Exiting.")
        return
    print(f"  ✓ File: {ncdc_path}")

    # Step 2: Pick output folder
    print("\nStep 2: Select the output folder")
    out_folder = _pick_folder("Select output folder", str(ncdc_path.parent))
    if not out_folder:
        out_folder = ncdc_path.parent / "NCDC_Cleaned"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output: {out_folder}")

    # Load and reshape
    print(f"\nLoading and reshaping NCDC data ...")
    try:
        raw = load_ncdc_data(ncdc_path)
    except Exception as e:
        print(f"  ✗ Error loading file: {e}")
        return
    print(f"  Raw rows (long format): {len(raw):,}")

    # Clean
    print("\nApplying cleaning rules ...")
    df, stats = clean_ncdc(raw)
    print(f"  Rule 1 (Invalid Gender)       dropped: {stats.get('dropped_invalid_gender', 0):,}")
    print(f"  Rule 4 (Assessment < DOB)     dropped: {stats.get('dropped_assessment_before_dob', 0):,}")
    print(f"  Rule 3 (Age invalid)          dropped: {stats.get('dropped_age_invalid', 0):,}")
    print(f"  Rule 2 (Measurement outliers) dropped: {stats.get('dropped_measurement_outlier', 0):,}")
    print(f"  Rule 5 (BMI > 40)             dropped: {stats.get('dropped_bmi_outlier', 0):,}")
    print(f"  Rule 6 (No measurement)       dropped: {stats.get('dropped_no_measurement', 0):,}")
    print(f"  Rule 7 (Null z-scores)        dropped: {stats.get('dropped_null_zscore', 0):,}")
    print(f"  ─────────────────────────────────────")
    print(f"  Total dropped : {stats['total_dropped']:,}  ({stats['total_dropped']/stats['raw_count']*100:.1f}%)")
    print(f"  Final records : {stats['final_count']:,}")

    if _ZSCORE_AVAILABLE:
        print(f"\n  WHO Indicators (in cleaned data):")
        print(f"    Kurang Berat Badan (WAZ < -2) : {stats.get('ind_kurang_berat', 0):,}")
        print(f"    Bantut             (HAZ < -2) : {stats.get('ind_bantut', 0):,}")
        print(f"    Susut              (BAZ < -2) : {stats.get('ind_susut', 0):,}")
        print(f"    Normal      (all FALSE)       : {stats.get('ind_normal', 0):,}")

    # Save quality report
    print("\nBuilding quality report ...")
    out_report = out_folder / "NCDC_QualityReport.xlsx"
    qr = build_quality_report(stats, df)
    qr.save(out_report)
    print(f"  Saved: {out_report}")

    # Save cleaned CSV
    print("\nSaving cleaned CSV ...")
    out_csv = out_folder / "NCDC_Cleaned.csv"
    save_cleaned_csv(df, out_csv)

    print()
    print("=" * 60)
    print("  Done! Files saved to:")
    print(f"    {out_report}")
    print(f"    {out_csv}")
    print("=" * 60)

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(
        "Done",
        f"NCDC cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n"
        f"• NCDC_Cleaned.csv\n"
        f"• NCDC_QualityReport.xlsx",
    )
    root.destroy()


if __name__ == "__main__":
    main()
