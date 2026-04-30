"""
Pemakanan Anthropometry Data Cleaning Script
=============================================
Cleans anthropometry data from MyVAS and generates:
  1. Pemakanan_Cleaned.csv - Cleansed data with WHO z-scores & indicators
  2. Pemakanan_QualityReport.xlsx - Data quality report

Source: Multiple Excel files (Jan-Apr 2026, Jan-Jun 2025, Jul-Sep 2025, Oct-Dec 2025)

DATA CLEANING RULES:
  Rule 1  : Drop all rows with ANY null value in critical columns
  Rule 2  : Validate IC format (12 digits, no suffix). Keep passports.
  Rule 3  : Drop biologically implausible weight (≤2 or >50 kg) / height (≤40 or >150 cm)
  Rule 4  : Drop age at measurement > 5 years (≥1826 days) or < 0
  Rule 5  : Drop records where measurement date < DOB
  Rule 6  : Drop implausible BMI > 40
  Rule 7  : Drop duplicate IC (keep most recent measurement date)
  Rule 8  : Gender validation - if invalid, try to extract from IC last digit
  Rule 9  : Drop records where ANY z-score is null

4 Core Indicators (WHO standards):
  1. Kurang Berat Badan  — WAZ < -2SD
  2. Bantut              — HAZ < -2SD
  3. Susut               — BAZ < -2SD
  4. Berlebihan BB/Obes  — BAZ > +1SD

2 Age Groups:
  - Bawah 2 Tahun: 0-729 days (0-23 months)
  - Bawah 5 Tahun: 0-1825 days (0-59 months)

Geographic Levels:
  - Malaysia (National)
  - Negeri (16 states)
  - Bahagian (for Sabah: 6, Sarawak: 12)
  - Daerah (Districts)
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path for WHO z-score module ─────────────────────────────
_BACKEND = Path(__file__).parent.parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore_daily import compute_zscore, classify_waz, classify_haz, classify_baz
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height bounds (biologically plausible for 0–5 years)
BERAT_MIN, BERAT_MAX = 2.0, 50.0      # kg
TINGGI_MIN, TINGGI_MAX = 40.0, 150.0  # cm
BMI_MAX = 40.0

# Age bounds (days)
AGE_MAX_DAYS = 1826  # under 5 years (60 months * 30.4375)

# WHO BIV thresholds
BIV = {
    "WAZ": (-6.0, +5.0),
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}
_BIV_LO_WAZ, _BIV_HI_WAZ = BIV["WAZ"]
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]

# Gender mapping
GENDER_MAP = {
    "MALE": "Male",
    "FEMALE": "Female",
    "LELAKI": "Male",
    "PEREMPUAN": "Female",
    "L": "Male",
    "P": "Female",
    "M": "Male",
    "F": "Female",
}

# Bahagian mapping for Sabah (6 divisions)
SABAH_BAHAGIAN = {
    "KOTA KINABALU": "Pantai Barat", "PENAMPANG": "Pantai Barat", "PUTATAN": "Pantai Barat",
    "PAPAR": "Pantai Barat", "TUARAN": "Pantai Barat", "KOTA BELUD": "Pantai Barat",
    "RANAU": "Pedalaman", "KUNDASANG": "Pedalaman", "TAMBUNAN": "Pedalaman",
    "KENINGAU": "Pedalaman", "TENOM": "Pedalaman", "NABAWAN": "Pedalaman",
    "SOOK": "Pedalaman", "KEMABONG": "Pedalaman",
    "KUDAT": "Kudat", "KOTA MARUDU": "Kudat", "PITAS": "Kudat", "MATUNGGONG": "Kudat",
    "SANDAKAN": "Sandakan", "KINABATANGAN": "Sandakan", "BELURAN": "Sandakan",
    "TONGOD": "Sandakan", "TELUPID": "Sandakan",
    "TAWAU": "Tawau", "LAHAD DATU": "Tawau", "SEMPORNA": "Tawau", "KUNAK": "Tawau",
    "BEAUFORT": "Beaufort", "KUALA PENYU": "Beaufort", "SIPITANG": "Beaufort", "MEMBAKUT": "Beaufort",
}

# Bahagian mapping for Sarawak (12 divisions)
SARAWAK_BAHAGIAN = {
    "KUCHING": "Kuching", "BAU": "Kuching", "LUNDU": "Kuching", "PADAWAN": "Kuching",
    "KOTA SAMARAHAN": "Samarahan", "SAMARAHAN": "Samarahan", "ASAJAYA": "Samarahan", 
    "SIMUNJAN": "Samarahan",
    "SERIAN": "Serian", "TEBEDU": "Serian",
    "SRI AMAN": "Sri Aman", "LUBOK ANTU": "Sri Aman", "ENGKILILI": "Sri Aman",
    "BETONG": "Betong", "PUSA": "Betong", "SARATOK": "Betong", "KABONG": "Betong",
    "SARIKEI": "Sarikei", "JULAU": "Sarikei", "MERADONG": "Sarikei", "PAKAN": "Sarikei",
    "SIBU": "Sibu", "SELANGAU": "Sibu", "KANOWIT": "Sibu",
    "MUKAH": "Mukah", "DALAT": "Mukah", "DARO": "Mukah", "MATU": "Mukah", "TANJUNG MANIS": "Mukah",
    "KAPIT": "Kapit", "SONG": "Kapit", "BELAGA": "Kapit", "BUKIT MABONG": "Kapit",
    "BINTULU": "Bintulu", "SEBAUH": "Bintulu", "TATAU": "Bintulu",
    "MIRI": "Miri", "MARUDI": "Miri", "SUBIS": "Miri", "TELANG USAN": "Miri", "BELURU": "Miri",
    "LIMBANG": "Limbang", "LAWAS": "Limbang",
}

# ---------------------------------------------------------------------------
# IC Validation
# ---------------------------------------------------------------------------

def validate_ic(ic_str):
    """
    Validate Malaysian IC format (12 digits, YYMMDDPPSSSSN)
    Returns: (is_valid, gender_from_ic)
    - is_valid: True if valid IC format
    - gender_from_ic: "Male" if last digit is odd, "Female" if even
    """
    ic_str = str(ic_str).strip().replace('-', '').replace(' ', '')
    
    # Check if 12 digits
    if not ic_str.isdigit() or len(ic_str) != 12:
        return (False, None)
    
    # Extract last digit for gender
    last_digit = int(ic_str[-1])
    gender = "Male" if last_digit % 2 == 1 else "Female"
    
    return (True, gender)


def is_passport(ic_str):
    """Check if the string looks like a passport (contains letters)"""
    ic_str = str(ic_str).strip()
    return any(c.isalpha() for c in ic_str)


def _get_bahagian(negeri: str, daerah: str) -> str | None:
    """Derive Bahagian from Negeri and Daerah for Sabah/Sarawak."""
    negeri_upper = str(negeri).upper().strip()
    daerah_upper = str(daerah).upper().strip()
    
    if negeri_upper == "SABAH":
        return SABAH_BAHAGIAN.get(daerah_upper, "Unknown")
    elif negeri_upper == "SARAWAK":
        return SARAWAK_BAHAGIAN.get(daerah_upper, "Unknown")
    return None


def _get_age_group(age_days: float) -> str:
    """Classify age into Bawah 2 Tahun or Bawah 5 Tahun."""
    if age_days is None or math.isnan(age_days):
        return "Unknown"
    if age_days < 730:  # 2 years
        return "Bawah 2 Tahun"
    elif age_days < 1826:  # 5 years
        return "2-5 Tahun"
    return "Lebih 5 Tahun"


# ---------------------------------------------------------------------------
# Load all files
# ---------------------------------------------------------------------------

def get_excel_files(folder_path: Path):
    """Return list of Excel files in the folder."""
    excel_files = list(folder_path.glob("*.xlsx"))
    print(f"  Found {len(excel_files)} Excel files")
    return excel_files


# ---------------------------------------------------------------------------
# Cleaning function
# ---------------------------------------------------------------------------

def clean_pemakanan_data(raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Clean Pemakanan data and compute WHO z-scores."""
    stats: dict = {"raw_count": len(raw)}
    df = raw.copy()

    # Standardize column names
    df.columns = df.columns.str.strip()

    # ── Rule 1: Drop rows with ANY null value in critical columns ─────────
    critical_cols = ['NAME', 'IC_NO_PASSPORT', 'DATE_OF_BIRTH', 'GENDER', 
                     'DOSE_DATE', 'LENGTH_HEIGHT_CM', 'WEIGHT_KG', 'STATE', 'DISTRICT']
    
    before = len(df)
    df = df.dropna(subset=critical_cols, how='any')
    stats["dropped_null_values"] = before - len(df)
    print(f"    Rule 1: Dropped {stats['dropped_null_values']:,} rows with null values")

    # ── Rule 8: Gender validation ─────────────────────────────────────────
    df["GENDER"] = df["GENDER"].astype(str).str.upper().str.strip()
    df["Gender"] = df["GENDER"].map(GENDER_MAP)
    
    # For invalid genders, try to extract from IC
    invalid_gender = df["Gender"].isna()
    if invalid_gender.sum() > 0:
        print(f"    Rule 8: Found {invalid_gender.sum():,} invalid genders, checking IC...")
        for idx in df[invalid_gender].index:
            ic = df.at[idx, 'IC_NO_PASSPORT']
            if not is_passport(ic):
                is_valid, gender_from_ic = validate_ic(ic)
                if is_valid and gender_from_ic:
                    df.at[idx, 'Gender'] = gender_from_ic
                    print(f"      Fixed gender for IC {ic}: {gender_from_ic}")
    
    # Drop rows with still invalid gender
    before = len(df)
    df = df[df["Gender"].notna()].copy()
    stats["dropped_invalid_gender"] = before - len(df)
    print(f"    Rule 8: Dropped {stats['dropped_invalid_gender']:,} invalid gender")

    # Jantina in Malay
    df["Jantina"] = df["Gender"].map({"Male": "Lelaki", "Female": "Perempuan"})

    # ── Rule 2: IC/Passport validation ────────────────────────────────────
    df["IC_NO_PASSPORT"] = df["IC_NO_PASSPORT"].astype(str).str.strip()
    df["Is_Passport"] = df["IC_NO_PASSPORT"].apply(is_passport)
    
    # Validate IC format (for non-passports)
    ic_validation = []
    for idx, row in df.iterrows():
        ic = row['IC_NO_PASSPORT']
        if row['Is_Passport']:
            ic_validation.append((True, None))  # Passports always valid
        else:
            is_valid, gender = validate_ic(ic)
            ic_validation.append((is_valid, gender))
    
    df['IC_Valid'] = [v[0] for v in ic_validation]
    df['Gender_From_IC'] = [v[1] for v in ic_validation]
    
    before = len(df)
    df = df[df['IC_Valid'] == True].copy()
    stats["dropped_invalid_ic"] = before - len(df)
    print(f"    Rule 2: Dropped {stats['dropped_invalid_ic']:,} invalid IC")

    # ── Standardize text fields ───────────────────────────────────────────
    for col in ["STATE", "DISTRICT", "FACILITY_NAME"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().str.strip()

    df["NAME"] = df["NAME"].astype(str).str.strip()

    # ── Parse dates ───────────────────────────────────────────────────────
    df["DATE_OF_BIRTH"] = pd.to_datetime(df["DATE_OF_BIRTH"], errors="coerce")
    df["DOSE_DATE"] = pd.to_datetime(df["DOSE_DATE"], errors="coerce")

    # Remove time component
    df["DATE_OF_BIRTH"] = df["DATE_OF_BIRTH"].dt.normalize()
    df["DOSE_DATE"] = df["DOSE_DATE"].dt.normalize()

    # ── Drop rows with null DOB or DOSE_DATE ──────────────────────────────
    before = len(df)
    df = df[df["DATE_OF_BIRTH"].notna() & df["DOSE_DATE"].notna()].copy()
    stats["dropped_null_dates"] = before - len(df)
    print(f"    Dropped {stats['dropped_null_dates']:,} null dates")

    # ── Rule 5: Drop where measurement date < DOB ─────────────────────────
    before = len(df)
    bad_date = df["DOSE_DATE"] < df["DATE_OF_BIRTH"]
    stats["dropped_date_before_dob"] = int(bad_date.sum())
    df = df[~bad_date].copy()
    print(f"    Rule 5: Dropped {stats['dropped_date_before_dob']:,} measurement before DOB")

    # ── Compute age at measurement (in DAYS for WHO z-score) ─────────────
    df["Age_Days"] = (df["DOSE_DATE"] - df["DATE_OF_BIRTH"]).dt.days
    df["Age_Months"] = (df["Age_Days"] / 30.4375).round(2)

    # ── Rule 4: Drop age >= 5 years or < 0 ────────────────────────────────
    before = len(df)
    age_invalid = (df["Age_Days"] < 0) | (df["Age_Days"] >= AGE_MAX_DAYS)
    stats["dropped_age_invalid"] = int(age_invalid.sum())
    df = df[~age_invalid].copy()
    print(f"    Rule 4: Dropped {stats['dropped_age_invalid']:,} age >= 5 years or < 0")

    # Add age group
    df["Kumpulan_Umur"] = df["Age_Days"].apply(_get_age_group)

    # ── Numeric measurements ──────────────────────────────────────────────
    df["WEIGHT_KG"] = pd.to_numeric(df["WEIGHT_KG"], errors="coerce")
    df["LENGTH_HEIGHT_CM"] = pd.to_numeric(df["LENGTH_HEIGHT_CM"], errors="coerce")
    df["BMI_KG_M2"] = pd.to_numeric(df["BMI_KG_M2"], errors="coerce")

    # ── Rule 3: Drop measurement outliers ─────────────────────────────────
    berat_bad = (df["WEIGHT_KG"] < BERAT_MIN) | (df["WEIGHT_KG"] > BERAT_MAX)
    tinggi_bad = (df["LENGTH_HEIGHT_CM"] < TINGGI_MIN) | (df["LENGTH_HEIGHT_CM"] > TINGGI_MAX)

    outlier_mask = (berat_bad & df["WEIGHT_KG"].notna()) | (tinggi_bad & df["LENGTH_HEIGHT_CM"].notna())
    stats["dropped_measurement_outlier"] = int(outlier_mask.sum())
    df = df[~outlier_mask].copy()
    print(f"    Rule 3: Dropped {stats['dropped_measurement_outlier']:,} measurement outliers")

    df["WEIGHT_KG"] = df["WEIGHT_KG"].round(1)
    df["LENGTH_HEIGHT_CM"] = df["LENGTH_HEIGHT_CM"].round(1)

    # ── Recalculate BMI ───────────────────────────────────────────────────
    valid_both = df["WEIGHT_KG"].notna() & df["LENGTH_HEIGHT_CM"].notna() & (df["LENGTH_HEIGHT_CM"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["WEIGHT_KG"] / ((df["LENGTH_HEIGHT_CM"] / 100) ** 2)).round(2),
        df["BMI_KG_M2"],
    )

    # ── Rule 6: Drop implausible BMI > 40 ─────────────────────────────────
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()
    print(f"    Rule 6: Dropped {stats['dropped_bmi_outlier']:,} BMI > 40")

    # ── Rule 7: Remove duplicate IC, keep most recent ─────────────────────
    before = len(df)
    df = df.sort_values("DOSE_DATE", ascending=False)
    df = df.drop_duplicates(subset=["IC_NO_PASSPORT"], keep="first")
    stats["dropped_duplicate_ic"] = before - len(df)
    print(f"    Rule 7: Dropped {stats['dropped_duplicate_ic']:,} duplicate IC (kept most recent)")

    # ── Add Bahagian for Sabah/Sarawak ────────────────────────────────────
    df["Bahagian"] = df.apply(
        lambda r: _get_bahagian(r.get("STATE", ""), r.get("DISTRICT", "")), axis=1
    )

    # ── WHO Z-scores (using DAILY LMS tables) ─────────────────────────────
    if _ZSCORE_AVAILABLE:
        print("    Computing WHO z-scores (daily LMS) ...")
        sex_s = df["Gender"].astype(str)
        age_s = df["Age_Days"]
        berat_s = df["WEIGHT_KG"]
        tinggi_s = df["LENGTH_HEIGHT_CM"]
        bmi_s = df["BMI"]

        n = len(df)
        waz_vals, haz_vals, baz_vals = [], [], []

        for i in range(n):
            age_d = age_s.iat[i]
            sex = sex_s.iat[i]
            w = berat_s.iat[i]
            h = tinggi_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age_d) or pd.isna(sex):
                waz_vals.append(None)
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            waz = compute_zscore("WAZ", float(w), int(age_d), sex) if pd.notna(w) else None
            haz = compute_zscore("HAZ", float(h), int(age_d), sex) if pd.notna(h) else None
            baz = compute_zscore("BAZ", float(bmi), int(age_d), sex) if pd.notna(bmi) else None

            # Apply BIV bounds
            if waz is not None and not (_BIV_LO_WAZ <= waz <= _BIV_HI_WAZ): waz = None
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            waz_vals.append(round(waz, 3) if waz is not None else None)
            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 5000 == 0:
                print(f"      Z-scores: {i+1:,}/{n:,} ...")

        df["WAZ"] = waz_vals
        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["WAZ_Status"] = df["WAZ"].apply(classify_waz)
        df["HAZ_Status"] = df["HAZ"].apply(classify_haz)
        df["BAZ_Status"] = df["BAZ"].apply(classify_baz)

        # 4 Core Indicator flags
        df["Ind_Kurang_Berat_Badan"] = df["WAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Bantut"] = df["HAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Susut"] = df["BAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Berlebihan_BB"] = df["BAZ"].apply(lambda z: z is not None and z > 1)
        df["Ind_Obes"] = df["BAZ"].apply(lambda z: z is not None and z > 2)

        df["Ind_Normal"] = ~(
            df["Ind_Kurang_Berat_Badan"] | df["Ind_Bantut"] | 
            df["Ind_Susut"] | df["Ind_Berlebihan_BB"]
        )

        # ── Rule 9: Drop rows where ANY z-score is null ───────────────────
        before = len(df)
        null_zscore = df["WAZ"].isna() | df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore.sum())
        df = df[~null_zscore].copy()
        print(f"    Rule 9: Dropped {stats['dropped_null_zscore']:,} null z-scores")

    else:
        for col in ["WAZ", "HAZ", "BAZ", "WAZ_Status", "HAZ_Status", "BAZ_Status",
                    "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", 
                    "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["ind_kurang_berat"] = int(df["Ind_Kurang_Berat_Badan"].sum())
        stats["ind_bantut"] = int(df["Ind_Bantut"].sum())
        stats["ind_susut"] = int(df["Ind_Susut"].sum())
        stats["ind_berlebihan_bb"] = int(df["Ind_Berlebihan_BB"].sum())
        stats["ind_obes"] = int(df["Ind_Obes"].sum())
        stats["ind_normal"] = int(df["Ind_Normal"].sum())

        stats["bawah_2_tahun"] = int((df["Kumpulan_Umur"] == "Bawah 2 Tahun").sum())
        stats["bawah_5_tahun"] = len(df)

    if "Gender" in df.columns:
        stats["gender_breakdown"] = df["Gender"].value_counts().to_dict()

    return df, stats


# ---------------------------------------------------------------------------
# Quality report (similar to NCDC)
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr


def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(stats: dict, df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "PEMAKANAN ANTHROPOMETRY — DATA QUALITY REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:D3")
    ws1["A3"].value = "Data Source: Anthropometry (MyVAS) | Jan 2025 - Apr 2026"
    ws1["A3"].font = Font(italic=True, size=9, color="7F7F7F")
    ws1["A3"].alignment = Alignment(horizontal="left", wrap_text=True)

    _hdr(ws1, 5, ["Metric", "Count", "Notes"])

    summary_rows = [
        ("Raw records (all files)", stats["raw_count"], "Source Excel"),
        ("Dropped — Null values", stats.get("dropped_null_values", 0), "Rule 1"),
        ("Dropped — Invalid Gender", stats.get("dropped_invalid_gender", 0), "Rule 8"),
        ("Dropped — Invalid IC", stats.get("dropped_invalid_ic", 0), "Rule 2"),
        ("Dropped — Null dates", stats.get("dropped_null_dates", 0), "Required"),
        ("Dropped — Measurement before DOB", stats.get("dropped_date_before_dob", 0), "Rule 5"),
        ("Dropped — Age invalid (≥5 yrs)", stats.get("dropped_age_invalid", 0), "Rule 4"),
        ("Dropped — Measurement outliers", stats.get("dropped_measurement_outlier", 0), "Rule 3"),
        ("Dropped — BMI > 40", stats.get("dropped_bmi_outlier", 0), "Rule 6"),
        ("Dropped — Duplicate IC", stats.get("dropped_duplicate_ic", 0), "Rule 7"),
        ("Dropped — Null z-score(s)", stats.get("dropped_null_zscore", 0), "Rule 9"),
        ("Final cleaned records", stats["final_count"], "After all rules"),
    ]
    for i, (label, val, note) in enumerate(summary_rows, 6):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        _cell(ws1, i, 1, label, bg=bg)
        _cell(ws1, i, 2, val, bg=bg, align="right")
        _cell(ws1, i, 3, note, bg=bg)

    row = len(summary_rows) + 7

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["4 Core Indicators", "Total", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        final = stats["final_count"]
        ind_rows = [
            ("1. Kurang Berat Badan (WAZ < -2SD)", stats.get("ind_kurang_berat", 0),
             f"{stats.get('ind_kurang_berat', 0)/max(final,1)*100:.1f}%", "WAZ < -2"),
            ("2. Bantut (HAZ < -2SD)", stats.get("ind_bantut", 0),
             f"{stats.get('ind_bantut', 0)/max(final,1)*100:.1f}%", "HAZ < -2"),
            ("3. Susut (BAZ < -2SD)", stats.get("ind_susut", 0),
             f"{stats.get('ind_susut', 0)/max(final,1)*100:.1f}%", "BAZ < -2"),
            ("4. Berlebihan BB (BAZ > +1SD)", stats.get("ind_berlebihan_bb", 0),
             f"{stats.get('ind_berlebihan_bb', 0)/max(final,1)*100:.1f}%", "BAZ > +1 (includes Obes)"),
            ("   └─ Obes (BAZ > +2SD)", stats.get("ind_obes", 0),
             f"{stats.get('ind_obes', 0)/max(final,1)*100:.1f}%", "BAZ > +2"),
            ("Normal (healthy child)", stats.get("ind_normal", 0),
             f"{stats.get('ind_normal', 0)/max(final,1)*100:.1f}%", "All indicators FALSE"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

        row = row + len(ind_rows) + 2
        _hdr(ws1, row, ["Age Group", "Count", "% of Total"], bg="548235")
        row += 1
        age_rows = [
            ("Bawah 2 Tahun (0-729 days)", stats.get("bawah_2_tahun", 0),
             f"{stats.get('bawah_2_tahun', 0)/max(final,1)*100:.1f}%"),
            ("Bawah 5 Tahun (0-1825 days)", final, "100.0%"),
        ]
        for i, r in enumerate(age_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

    _auto_width(ws1)

    # ── Sheet 2: By Negeri (similar structure) ────────────────────────────
    # ... (add similar sheets as NCDC report)

    return wb


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def save_cleaned_excel(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()

    for col in ["DATE_OF_BIRTH", "DOSE_DATE"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    for col in ["Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", 
                "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)

    out_cols = [
        "FACILITY_NAME", "STATE", "Bahagian", "DISTRICT", "LATITUDE", "LONGITUDE",
        "NAME", "IC_NO_PASSPORT", "Is_Passport", "Gender", "Jantina", "ETHNICITY",
        "DATE_OF_BIRTH", "DOSE_DATE", "Age_Days", "Age_Months", "Kumpulan_Umur",
        "WEIGHT_KG", "LENGTH_HEIGHT_CM", "BMI",
        "WAZ", "HAZ", "BAZ",
        "WAZ_Status", "HAZ_Status", "BAZ_Status",
        "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut",
        "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal",
        "SOURCE"
    ]
    out_cols = [c for c in out_cols if c in df_out.columns]
    df_out = df_out[out_cols]

    sort_cols = ["Gender", "STATE"]
    if sort_cols:
        df_out = df_out.sort_values(sort_cols)

    df_out.to_excel(out_path, index=False, engine='openpyxl')
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# Dataset registry
# ---------------------------------------------------------------------------

DATASETS = {
    "1": "pemakanan_anthropometry_jan_apr_2026",
    "2": "pemakanan_anthropometry jan jun 2025",
    "3": "pemakanan_anthropometry_july_sept_2025",
    "4": "pemakanan_anthropometry_oct_dec 2025",
}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _prompt_dataset() -> tuple[str, Path]:
    """
    Interactively ask the user which dataset to process and where to find it.
    Returns (dataset_label, input_file_path).
    """
    print("\n  Available datasets:")
    print("  " + "─" * 50)
    for key, name in DATASETS.items():
        print(f"    [{key}]  {name}")
    print("  " + "─" * 50)

    while True:
        choice = input("\n  Enter dataset number (1-4): ").strip()
        if choice in DATASETS:
            dataset_name = DATASETS[choice]
            break
        print("  ⚠  Invalid choice. Please enter a number between 1 and 4.")

    print(f"\n  Selected  : {dataset_name}")

    # ── Locate the input file ─────────────────────────────────────────────
    print("\n  Enter the full path to the Excel file for this dataset.")
    print(f"  Example   : C:\\Data\\{dataset_name}.xlsx")

    while True:
        raw_path = input("\n  Input file path: ").strip().strip('"').strip("'")
        input_path = Path(raw_path)
        if input_path.is_file() and input_path.suffix.lower() == ".xlsx":
            break
        if not input_path.exists():
            print(f"  ⚠  File not found: {input_path}")
        else:
            print(f"  ⚠  File must be an .xlsx file.")

    return dataset_name, input_path


def _prompt_output_folder() -> Path:
    """Ask the user where to save the output files."""
    print("\n  Enter the output folder path where results will be saved.")
    print(f"  Example   : C:\\Data\\Cleaned")
    print("  (Press Enter to use the same folder as the input file.)")

    raw = input("\n  Output folder path: ").strip().strip('"').strip("'")

    if raw == "":
        return None   # resolved later relative to input file
    out = Path(raw)
    out.mkdir(parents=True, exist_ok=True)
    return out


def main():
    print("=" * 70)
    print("  Pemakanan Anthropometry Data Cleaning Tool")
    print("  Status Pemakanan Kanak-Kanak Bawah Lima Tahun")
    print("=" * 70)

    # ── Step 1: Choose dataset & input file ───────────────────────────────
    dataset_name, input_path = _prompt_dataset()

    # ── Step 2: Choose output folder ─────────────────────────────────────
    out_folder = _prompt_output_folder()
    if out_folder is None:
        out_folder = input_path.parent / "Cleaned"
    out_folder.mkdir(parents=True, exist_ok=True)

    # ── Step 3: Load ──────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"  Processing  : {input_path.name}")
    print(f"  Dataset     : {dataset_name}")
    print(f"  Output to   : {out_folder}")
    print(f"{'='*70}")

    raw = pd.read_excel(input_path)
    print(f"\n  Rows: {len(raw):,}  |  Columns: {len(raw.columns)}")

    # Tag source so it appears in the cleaned file
    raw["SOURCE"] = dataset_name

    # ── Step 4: Clean ─────────────────────────────────────────────────────
    print("\n  Applying cleaning rules ...")
    df, stats = clean_pemakanan_data(raw)

    # ── Step 5: Summary ───────────────────────────────────────────────────
    print(f"\n  {'─'*60}")
    print(f"  Total dropped : {stats['total_dropped']:,}  "
          f"({stats['total_dropped'] / stats['raw_count'] * 100:.1f}%)")
    print(f"  Final records : {stats['final_count']:,}")

    if _ZSCORE_AVAILABLE:
        print(f"\n  4 Core Indicators:")
        print(f"    1. Kurang Berat Badan (WAZ < -2) : {stats.get('ind_kurang_berat', 0):,}")
        print(f"    2. Bantut (HAZ < -2)             : {stats.get('ind_bantut', 0):,}")
        print(f"    3. Susut (BAZ < -2)              : {stats.get('ind_susut', 0):,}")
        print(f"    4. Berlebihan BB (BAZ > +1)      : {stats.get('ind_berlebihan_bb', 0):,}")
        print(f"       └─ Obes (BAZ > +2)            : {stats.get('ind_obes', 0):,}")
        print(f"    Normal                           : {stats.get('ind_normal', 0):,}")

    # ── Step 6: Save outputs ──────────────────────────────────────────────
    # Use the original dataset name (no extension) as the base stem
    stem = dataset_name

    cleaned_path = out_folder / f"Cleaned_{stem}.xlsx"
    report_path  = out_folder / f"DataQualityReport_{stem}.xlsx"

    print(f"\n  Saving outputs ...")
    save_cleaned_excel(df, cleaned_path)

    wb = build_quality_report(stats, df)
    wb.save(report_path)
    print(f"  Saved: {report_path}")

    # ── Done ──────────────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print("  ✓  Done! Two files saved to:")
    print(f"     • Cleaned_{stem}.xlsx")
    print(f"     • DataQualityReport_{stem}.xlsx")
    print(f"  Folder: {out_folder}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()