"""
KKM Berat & Tinggi Data Cleaning Script
========================================
Cleans 2024 and 2025 KKM weight/height datasets and generates:
  1. KKM_BeratTinggi_Cleaned_2024_2025.xlsx  — Cleansed data
  2. KKM_BeratTinggi_QualityReport_2024_2025.xlsx — Quality report

Cleaning rules sourced from: KKM_ColumnCleaning_Reference.md
"""
import numpy as np
import pandas as pd
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths (defaults — overridden interactively at runtime)
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
VALID_NEGERI = {
    "JOHOR", "KEDAH", "KELANTAN", "MELAKA", "NEGERI SEMBILAN",
    "PAHANG", "PERAK", "PERLIS", "PULAU PINANG", "SABAH",
    "SARAWAK", "SELANGOR", "TERENGGANU",
    "WP KUALA LUMPUR", "WP LABUAN", "WP PUTRAJAYA",
    "WILAYAH PERSEKUTUAN KUALA LUMPUR",
    "WILAYAH PERSEKUTUAN LABUAN",
    "WILAYAH PERSEKUTUAN PUTRAJAYA",
}

BERAT_MIN, BERAT_MAX   = 12.0, 50.0   # kg  — BR-06
TINGGI_MIN, TINGGI_MAX = 100.0, 160.0  # cm  — BR-07
AGE_MIN, AGE_MAX       = 6.0, 8.0      # years — BR-04

# WHO BMI-for-age thresholds (7-year-old cohort, 2007 reference)
BMI_UNDERWEIGHT = 13.5
BMI_OVERWEIGHT  = 16.5
BMI_OBESE       = 18.5

# Height-for-age proxies for 7-year-olds (approx ±2 SD from WHO median ~120 cm)
STUNTED_THRESHOLD = 112.0
TALL_THRESHOLD    = 132.0


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------

def bmi_category(bmi: float) -> str:
    if pd.isna(bmi):
        return "Unknown"
    if bmi < BMI_UNDERWEIGHT:
        return "Underweight"
    if bmi < BMI_OVERWEIGHT:
        return "Normal"
    if bmi < BMI_OBESE:
        return "Overweight"
    return "Obese"


def height_category(height: float) -> str:
    if pd.isna(height):
        return "Unknown"
    if height < STUNTED_THRESHOLD:
        return "Stunted"
    if height > TALL_THRESHOLD:
        return "Tall"
    return "Normal"


def clean_dataset(raw: pd.DataFrame, year: int, drop_invalid: bool = False) -> tuple[pd.DataFrame, dict]:
    """
    Apply all cleaning rules to a raw KKM dataframe.
    Returns (cleaned_df, stats_dict).

    drop_invalid=True  → rows failing any rule are REMOVED entirely (used for 2025)
    drop_invalid=False → rows are flagged but kept (used for 2024)

    Actual column names in source files (normalised):
      TAHUN PERSEKOLAHAN, NEGERI, DAERAH, NAMA SEKOLAH,
      LOKASI (BANDAR/LUAR BANDAR), JENIS SEKOLAH, THN_TING,
      JANTINA, ID_MURID, TARIKH LAHIR,
      TARIKH PENGUKURAN BERAT/ TINGGI, BERAT (kg), TINGGI (cm)
    """
    df = raw.copy()

    # ── 0. Normalise all column names (strip whitespace + newlines) ────────
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    print(f"  Columns detected: {list(df.columns)}")

    stats: dict = {"year": year, "raw_count": len(df)}

    # ── 1. Rename columns to clean snake_case IDs ─────────────────────────
    df = df.rename(columns={
        "TAHUN PERSEKOLAHAN":                "Tahun_Persekolahan",
        "NEGERI":                            "Negeri",
        "DAERAH":                            "Daerah",
        "NAMA SEKOLAH":                      "Nama_Sekolah",
        "LOKASI (BANDAR/LUAR BANDAR)":       "Lokasi",
        "JENIS SEKOLAH":                     "Jenis_Sekolah",
        "THN_TING":                          "Thn_Ting",
        "JANTINA":                           "Jantina",
        "ID_MURID":                          "ID_Murid",
        "TARIKH LAHIR":                      "Tarikh_Lahir_Raw",
        "TARIKH PENGUKURAN BERAT/ TINGGI":   "Tarikh_Pengukuran_Raw",
        "BERAT (kg)":                        "Berat_kg_Raw",
        "TINGGI (cm)":                       "Tinggi_cm_Raw",
    })

    # ── 2. Text normalisation ─────────────────────────────────────────────
    for col in ["Negeri", "Daerah", "Jantina", "Jenis_Sekolah", "Thn_Ting"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().str.strip()

    for col in ["Nama_Sekolah", "ID_Murid"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    df["Lokasi"] = df["Lokasi"].astype(str).str.upper().str.strip()

    # ── 3. Validate Negeri ────────────────────────────────────────────────
    df["Negeri_Valid"] = df["Negeri"].isin(VALID_NEGERI)
    stats["invalid_negeri"] = int((~df["Negeri_Valid"]).sum())

    # ── 4. Lokasi mapping (BANDAR→Urban, LUAR BANDAR→Rural) ──────────────
    lokasi_map = {"BANDAR": "Urban", "LUAR BANDAR": "Rural"}
    df["Lokasi_EN"] = df["Lokasi"].map(lokasi_map).fillna("Unknown")

    # ── 5. Year level mapping ─────────────────────────────────────────────
    thn_map = {
        "TAHUN SATU":       "Year 1",
        "TAHUN DUA":        "Year 2",
        "TAHUN TIGA":       "Year 3",
        "KELAS KHAS RENDAH":"Special Ed",
    }
    df["Year_Level"] = df["Thn_Ting"].map(thn_map).fillna(df["Thn_Ting"])

    # ── 6. Gender mapping (BR-03) ─────────────────────────────────────────
    gender_map = {"LELAKI": "Male", "PEREMPUAN": "Female"}
    df["Gender"] = df["Jantina"].map(gender_map).fillna("Unknown")
    stats["ragu_gender"]    = int((df["Jantina"] == "RAGU").sum())
    stats["unknown_gender"] = int((df["Gender"] == "Unknown").sum())

    # ── 7. Parse Tarikh_Lahir ─────────────────────────────────────────────
    df["Tarikh_Lahir"] = pd.to_datetime(df["Tarikh_Lahir_Raw"], dayfirst=True, errors="coerce")
    stats["null_tarikh_lahir"] = int(df["Tarikh_Lahir"].isna().sum())

    # ── 8. Parse & validate Tarikh_Pengukuran (BR-05) ─────────────────────
    df["Tarikh_Pengukuran"] = pd.to_datetime(df["Tarikh_Pengukuran_Raw"], dayfirst=True, errors="coerce")

    epoch_mask  = df["Tarikh_Pengukuran"].dt.normalize() == pd.Timestamp("1970-01-01")
    pre20_mask  = df["Tarikh_Pengukuran"] < pd.Timestamp("2020-01-01")
    future_mask = df["Tarikh_Pengukuran"] > pd.Timestamp.now()
    bad_date    = epoch_mask | pre20_mask | future_mask

    stats["unix_epoch_dates"]      = int(epoch_mask.sum())
    stats["pre2020_dates"]         = int((pre20_mask & ~epoch_mask).sum())
    stats["future_dates"]          = int(future_mask.sum())
    stats["null_measurement_date"] = int(df["Tarikh_Pengukuran"].isna().sum())

    if drop_invalid:
        rows_before = len(df)
        df = df[~bad_date].copy()
        stats["rows_dropped_bad_date"] = rows_before - len(df)
    else:
        df.loc[bad_date, "Tarikh_Pengukuran"] = pd.NaT
        stats["rows_dropped_bad_date"] = 0
    stats["invalid_measurement_date"] = int(df["Tarikh_Pengukuran"].isna().sum())

    # ── SPECIAL: For 2024 data, set ALL dates to 31/12/2024 ───────────────
    # This allows age calculation for all students as of 31/12/2024
    if year == 2024:
        df["Tarikh_Pengukuran"] = pd.Timestamp("2024-12-31")
        stats["standardized_to_31dec2024"] = len(df)
        print(f"    ✓ Set ALL measurement dates to 31/12/2024 for {len(df):,} records")

    # ── 9. Age at measurement (in days and years) ────────────────────────
    has_both = df["Tarikh_Lahir"].notna() & df["Tarikh_Pengukuran"].notna()
    df["Age_Days"] = np.where(
        has_both,
        (df["Tarikh_Pengukuran"] - df["Tarikh_Lahir"]).dt.days,
        np.nan,
    )
    df["Age_At_Measurement"] = np.where(
        has_both,
        df["Age_Days"] / 365.25,
        np.nan,
    )
    df["Age_At_Measurement"] = df["Age_At_Measurement"].round(4)
    age_valid = df["Age_At_Measurement"].notna()
    age_out   = age_valid & ((df["Age_At_Measurement"] < AGE_MIN) | (df["Age_At_Measurement"] > AGE_MAX))
    stats["invalid_age"] = int(age_out.sum())

    # ── 10. Clean Berat_kg (BR-06) ────────────────────────────────────────
    df["Berat_kg"] = pd.to_numeric(df["Berat_kg_Raw"], errors="coerce")
    stats["null_berat_original"] = int(df["Berat_kg"].isna().sum())

    berat_oor = (df["Berat_kg"] < BERAT_MIN) | (df["Berat_kg"] > BERAT_MAX)
    stats["berat_out_of_range"] = int(berat_oor.sum())
    if drop_invalid:
        rows_before = len(df)
        df = df[~berat_oor].copy()
        stats["rows_dropped_berat"] = rows_before - len(df)
    else:
        df.loc[berat_oor, "Berat_kg"] = np.nan
        stats["rows_dropped_berat"] = 0
    df["Berat_kg"] = df["Berat_kg"].round(1)
    stats["null_berat_after_clean"] = int(df["Berat_kg"].isna().sum())

    # ── 11. Clean Tinggi_cm (BR-07) ───────────────────────────────────────
    df["Tinggi_cm"] = pd.to_numeric(df["Tinggi_cm_Raw"], errors="coerce")
    stats["null_tinggi_original"] = int(df["Tinggi_cm"].isna().sum())

    tinggi_oor = (df["Tinggi_cm"] < TINGGI_MIN) | (df["Tinggi_cm"] > TINGGI_MAX)
    stats["tinggi_out_of_range"] = int(tinggi_oor.sum())
    if drop_invalid:
        rows_before = len(df)
        df = df[~tinggi_oor].copy()
        stats["rows_dropped_tinggi"] = rows_before - len(df)
    else:
        df.loc[tinggi_oor, "Tinggi_cm"] = np.nan
        stats["rows_dropped_tinggi"] = 0
    df["Tinggi_cm"] = df["Tinggi_cm"].round(1)
    stats["null_tinggi_after_clean"] = int(df["Tinggi_cm"].isna().sum())

    # ── 12. BMI & derived categories (BR-13) ──────────────────────────────
    valid_meas = df["Berat_kg"].notna() & df["Tinggi_cm"].notna() & (df["Tinggi_cm"] > 0)
    df["BMI"] = np.where(
        valid_meas,
        (df["Berat_kg"] / ((df["Tinggi_cm"] / 100) ** 2)).round(2),
        np.nan,
    )
    df["BMI_Category"]    = df["BMI"].apply(bmi_category)
    df["Height_Category"] = df["Tinggi_cm"].apply(height_category)

    # ── 13. Completeness flags ─────────────────────────────────────────────
    df["Has_Complete_Measurements"] = df["Berat_kg"].notna() & df["Tinggi_cm"].notna()
    df["Is_Valid_Age"]              = age_valid & (df["Age_At_Measurement"] >= AGE_MIN) & (df["Age_At_Measurement"] <= AGE_MAX)
    df["Is_Valid_Measurement_Date"] = df["Tarikh_Pengukuran"].notna()

    # ── 14. Duplicate ID_Murid (BR-02) ────────────────────────────────────
    id_clean = df["ID_Murid"].replace({"nan": np.nan, "": np.nan})
    df["Is_Duplicate_ID"] = id_clean.duplicated(keep=False) & id_clean.notna()
    stats["duplicate_ic"] = int(df["Is_Duplicate_ID"].sum())

    # ── 15. Composite quality flag ─────────────────────────────────────────
    conditions = [
        (~df["Has_Complete_Measurements"], "Missing Measurements"),
        (~df["Is_Valid_Age"],              "Invalid Age"),
        (~df["Is_Valid_Measurement_Date"], "Invalid Date"),
        (df["Gender"] == "Unknown",        "Unknown Gender"),
        (df["Is_Duplicate_ID"],            "Duplicate ID"),
    ]

    issue_parts = []
    for mask, label in conditions:
        issue_parts.append(mask.astype(int).map({0: "", 1: label}))
    combined = issue_parts[0].copy()
    for part in issue_parts[1:]:
        need_sep = (combined != "") & (part != "")
        combined = combined + need_sep.map({True: "; ", False: ""}) + part
    combined = combined.replace("", "Valid")
    df["Data_Quality_Flag"] = combined

    stats["valid_records"]   = int((df["Data_Quality_Flag"] == "Valid").sum())
    stats["flagged_records"] = int((df["Data_Quality_Flag"] != "Valid").sum())
    stats["rows_after_clean"] = len(df)
    stats["total_rows_dropped"] = stats["raw_count"] - len(df)

    # ── 16. Select final output columns ───────────────────────────────────
    out_cols = [
        "Tahun_Persekolahan",
        "Negeri", "Daerah", "Nama_Sekolah",
        "Lokasi", "Lokasi_EN",
        "Jenis_Sekolah", "Thn_Ting", "Year_Level",
        "Gender", "Jantina",
        "ID_Murid",
        "Tarikh_Lahir", "Age_Days", "Age_At_Measurement",
        "Tarikh_Pengukuran",
        "Berat_kg", "Tinggi_cm",
        "BMI", "BMI_Category", "Height_Category",
        "Has_Complete_Measurements",
        "Is_Valid_Age", "Is_Valid_Measurement_Date", "Is_Duplicate_ID",
        "Data_Quality_Flag",
    ]
    cleaned = df[out_cols]
    return cleaned, stats


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _cell_style(
    ws, row: int, col: int, value,
    bold: bool = False,
    bg: str | None = None,
    font_color: str = "000000",
    align: str = "left",
    wrap: bool = False,
    number_format: str | None = None,
):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if number_format:
        cell.number_format = number_format
    thin = Side(style="thin", color="D0D0D0")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _header_row(ws, row: int, headers: list[str], bg: str = "1F4E79"):
    for c, h in enumerate(headers, 1):
        _cell_style(ws, row, c, h, bold=True, bg=bg, font_color="FFFFFF", align="center")


def _auto_col_width(ws, min_w: int = 10, max_w: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Quality report builder
# ---------------------------------------------------------------------------

def build_quality_report(all_stats: list[dict], df_all: pd.DataFrame):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False

    # Title block
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = "KKM BERAT & TINGGI — DATA QUALITY REPORT"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws1["A2"].font  = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # Build year-keyed lookup — works for 1 or 2 years
    stats_by_year: dict[int, dict] = {s["year"]: s for s in all_stats}
    years_present = sorted(stats_by_year.keys())

    # Dynamic headers: one column per year + Combined (only if >1 year) + Notes
    year_headers = [str(y) for y in years_present]
    if len(years_present) > 1:
        year_headers.append("Combined")
    headers = ["Metric"] + year_headers + ["Notes"]
    _header_row(ws1, 4, headers)

    def _y(key: str) -> list:
        """Return per-year values for a stats key."""
        return [stats_by_year[y].get(key, "—") for y in years_present]

    def _total(key: str):
        vals = [stats_by_year[y].get(key, 0) for y in years_present]
        return sum(v for v in vals if isinstance(v, (int, float)))

    def _pct(year: int) -> str:
        s = stats_by_year[year]
        raw = s.get("raw_count", 0)
        return f"{s.get('valid_records', 0) / raw * 100:.1f}%" if raw else "—"

    def _make_row(label: str, key: str, note: str) -> tuple:
        vals = _y(key)
        row = [label] + vals
        if len(years_present) > 1:
            row.append(_total(key))
        row.append(note)
        return tuple(row)

    total_raw    = _total("raw_count")
    total_valid  = _total("valid_records")
    total_flagged = _total("flagged_records")

    pct_vals = [_pct(y) for y in years_present]
    pct_row  = ["Data Quality Rate (%)"] + pct_vals
    if len(years_present) > 1:
        pct_row.append(f"{total_valid/total_raw*100:.1f}%" if total_raw else "—")
    pct_row.append("Valid / Total")

    rows_data = [
        _make_row("Total Records (Raw)",          "raw_count",                "Before any cleaning"),
        _make_row("Rows After Cleaning",           "rows_after_clean",         "After all rules applied"),
        _make_row("Rows Dropped (total)",          "total_rows_dropped",       "2025: dropped; 2024: flagged"),
        _make_row("Valid Records",                 "valid_records",            "Pass all business rules"),
        _make_row("Flagged Records",               "flagged_records",          "At least one issue found"),
        tuple(pct_row),
        _make_row("Missing Weight (after clean)",  "null_berat_after_clean",   "NULL after range validation"),
        _make_row("Missing Height (after clean)",  "null_tinggi_after_clean",  "NULL after range validation"),
        _make_row("Weight Out-of-Range",           "berat_out_of_range",       f"Outside {BERAT_MIN}–{BERAT_MAX} kg"),
        _make_row("Height Out-of-Range",           "tinggi_out_of_range",      f"Outside {TINGGI_MIN}–{TINGGI_MAX} cm"),
        _make_row("Invalid Measurement Date",      "invalid_measurement_date", "Unix epoch / pre-2020 / future"),
        _make_row("Unix Epoch Dates (1970-01-01)", "unix_epoch_dates",         "System error — nullified or dropped"),
        _make_row("Invalid Age (out of 6–8 yrs)",  "invalid_age",              "BR-04"),
        _make_row("Unknown / Invalid Gender",       "unknown_gender",           "RAGU or unrecognised value"),
        _make_row("Duplicate ID",                  "duplicate_ic",             "BR-02: same ID_MURID in dataset"),
    ]

    ALT_BG = "EBF3FA"
    for i, rd in enumerate(rows_data, 5):
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(rd, 1):
            _cell_style(ws1, i, c, v, bg=bg)

    _auto_col_width(ws1)

    # ── Sheet 2: Cleaning Rules Applied ───────────────────────────────────
    ws2 = wb.create_sheet("Cleaning Rules")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "CLEANING RULES APPLIED"
    ws2["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    DROP_INVALID_YEARS = {2025}

    def _yvals(key: str) -> list:
        return [stats_by_year[y].get(key, "—") for y in years_present]

    def _action_label(null_txt: str, drop_txt: str) -> str:
        actions = set(_action(y, null_txt, drop_txt) for y in years_present)
        return " / ".join(sorted(actions))

    def _action(year: int, null_txt: str, drop_txt: str) -> str:
        return drop_txt if year in DROP_INVALID_YEARS else null_txt

    rule_year_headers = [f"Affected ({y})" for y in years_present]
    rule_headers = ["Rule ID", "Column", "Category", "Rule Description", "Action Taken"] + rule_year_headers
    _header_row(ws2, 3, rule_headers)

    rules = [
        ("BR-01", "Negeri / Daerah / Jantina / Nama_Sekolah",
         "Text Normalisation", "UPPER() + TRIM() on all text fields",
         "Standardised") + tuple("All rows" for _ in years_present),
        ("BR-02", "ID_MURID",
         "Duplicate", "Flag duplicate Student IDs within each year",
         "Flagged (Is_Duplicate_ID = True)") + tuple(_yvals("duplicate_ic")),
        ("BR-03", "Jantina",
         "Gender", "Map LELAKI→Male, PEREMPUAN→Female; RAGU→Unknown",
         "Mapped; invalid flagged") + tuple(_yvals("unknown_gender")),
        ("BR-04", "Tarikh_Lahir + Tarikh_Pengukuran",
         "Age Validation", f"Age at measurement must be {AGE_MIN}–{AGE_MAX} years",
         "Flagged (Is_Valid_Age = False)") + tuple(_yvals("invalid_age")),
        ("BR-05a", "Tarikh_Pengukuran",
         "Date Validation", "Unix epoch 1970-01-01 (system error)",
         _action_label("Set to NULL", "Row dropped")) + tuple(_yvals("unix_epoch_dates")),
        ("BR-05b", "Tarikh_Pengukuran",
         "Date Validation", "Dates before 2020-01-01 (implausible)",
         _action_label("Set to NULL", "Row dropped")) + tuple(_yvals("pre2020_dates")),
        ("BR-05c", "Tarikh_Pengukuran",
         "Date Validation", "Future dates",
         _action_label("Set to NULL", "Row dropped")) + tuple(_yvals("future_dates")),
        ("BR-06", "Berat (kg)",
         "Measurement Range", f"Weight outside {BERAT_MIN}–{BERAT_MAX} kg",
         _action_label("Set to NULL", "Row dropped")) + tuple(_yvals("berat_out_of_range")),
        ("BR-07", "Tinggi (cm)",
         "Measurement Range", f"Height outside {TINGGI_MIN}–{TINGGI_MAX} cm",
         _action_label("Set to NULL", "Row dropped")) + tuple(_yvals("tinggi_out_of_range")),
        ("BR-13", "BMI",
         "Derived Metric", "BMI = weight_kg / (height_m²); NULL if inputs missing",
         "Computed") + tuple("—" for _ in years_present),
        ("BR-14", "BMI_Category",
         "Derived Metric",
         f"WHO 7-yr: <{BMI_UNDERWEIGHT}=Underweight, {BMI_UNDERWEIGHT}–{BMI_OVERWEIGHT}=Normal, {BMI_OVERWEIGHT}–{BMI_OBESE}=Overweight, ≥{BMI_OBESE}=Obese",
         "Computed") + tuple("—" for _ in years_present),
        ("BR-15", "Height_Category",
         "Derived Metric",
         f"Stunted <{STUNTED_THRESHOLD} cm, Normal {STUNTED_THRESHOLD}–{TALL_THRESHOLD} cm, Tall >{TALL_THRESHOLD} cm",
         "Computed") + tuple("—" for _ in years_present),
    ]

    for i, r in enumerate(rules, 4):
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(r, 1):
            _cell_style(ws2, i, c, v, bg=bg, wrap=(c == 4))

    _auto_col_width(ws2)

    # ── Sheet 3: BMI Distribution ─────────────────────────────────────────
    ws3 = wb.create_sheet("BMI Distribution")
    ws3.sheet_view.showGridLines = False

    year_label = " + ".join(str(y) for y in years_present)
    n_bmi_cols = 1 + len(years_present) * 2 + (1 if len(years_present) > 1 else 0)
    ws3.merge_cells(f"A1:{get_column_letter(n_bmi_cols)}1")
    ws3["A1"].value = f"BMI DISTRIBUTION — {year_label}"
    ws3["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    bmi_yr_hdrs = []
    for y in years_present:
        bmi_yr_hdrs += [f"{y} Count", f"{y} %"]
    if len(years_present) > 1:
        bmi_yr_hdrs.append("Combined")
    _header_row(ws3, 3, ["BMI Category"] + bmi_yr_hdrs)

    bmi_cats = ["Underweight", "Normal", "Overweight", "Obese", "Unknown"]
    for i, cat in enumerate(bmi_cats, 4):
        row_vals = [cat]
        counts = []
        for y in years_present:
            sub_y = df_all[df_all["Tahun_Persekolahan"] == y]
            cnt = int((sub_y["BMI_Category"] == cat).sum())
            tot = len(sub_y)
            row_vals += [cnt, f"{cnt/tot*100:.1f}%" if tot else "—"]
            counts.append(cnt)
        if len(years_present) > 1:
            row_vals.append(sum(counts))
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(row_vals, 1):
            _cell_style(ws3, i, c, v, bg=bg, align="center" if c > 1 else "left")

    _auto_col_width(ws3)

    # ── Sheet 4: By State ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("By State")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:H1")
    ws4["A1"].value = f"BMI DISTRIBUTION BY STATE — {year_label}"
    ws4["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    _header_row(ws4, 3, ["Negeri", "Total", "Underweight", "Normal", "Overweight", "Obese", "Unknown", "With BMI %"])

    state_grp = df_all.groupby("Negeri")["BMI_Category"].value_counts().unstack(fill_value=0)
    for cat in ["Underweight", "Normal", "Overweight", "Obese", "Unknown"]:
        if cat not in state_grp.columns:
            state_grp[cat] = 0
    state_grp["Total"] = state_grp.sum(axis=1)
    state_grp["Valid_Pct"] = (
        (state_grp.get("Underweight", 0) + state_grp.get("Normal", 0) +
         state_grp.get("Overweight", 0) + state_grp.get("Obese", 0)) /
        state_grp["Total"] * 100
    ).round(1)
    state_grp = state_grp.sort_values("Total", ascending=False)

    for i, (negeri, row) in enumerate(state_grp.iterrows(), 4):
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        vals = [negeri, int(row["Total"]),
                int(row.get("Underweight", 0)), int(row.get("Normal", 0)),
                int(row.get("Overweight", 0)), int(row.get("Obese", 0)),
                int(row.get("Unknown", 0)), f"{row['Valid_Pct']:.1f}%"]
        for c, v in enumerate(vals, 1):
            _cell_style(ws4, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_col_width(ws4)

    # ── Sheet 5: By Gender ────────────────────────────────────────────────
    ws5 = wb.create_sheet("By Gender")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:G1")
    ws5["A1"].value = f"BMI DISTRIBUTION BY GENDER — {year_label}"
    ws5["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws5["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    _header_row(ws5, 3, ["Gender", "Total", "Underweight", "Normal", "Overweight", "Obese", "Unknown"])

    for i, gender in enumerate(["Male", "Female", "Unknown"], 4):
        sub = df_all[df_all["Gender"] == gender]
        vc  = sub["BMI_Category"].value_counts()
        bg  = ALT_BG if i % 2 == 0 else "FFFFFF"
        vals = [gender, len(sub),
                int(vc.get("Underweight", 0)), int(vc.get("Normal", 0)),
                int(vc.get("Overweight", 0)), int(vc.get("Obese", 0)),
                int(vc.get("Unknown", 0))]
        for c, v in enumerate(vals, 1):
            _cell_style(ws5, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_col_width(ws5)

    # ── Sheet 6: Height (Stunting) Analysis ───────────────────────────────
    ws6 = wb.create_sheet("Height Analysis")
    ws6.sheet_view.showGridLines = False

    n_ht_cols = 1 + len(years_present) * 2 + (1 if len(years_present) > 1 else 0)
    ws6.merge_cells(f"A1:{get_column_letter(n_ht_cols)}1")
    ws6["A1"].value = f"HEIGHT-FOR-AGE (STUNTING) ANALYSIS — {year_label}"
    ws6["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws6["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 28

    ht_yr_hdrs = []
    for y in years_present:
        ht_yr_hdrs += [f"{y} Count", f"{y} %"]
    if len(years_present) > 1:
        ht_yr_hdrs.append("Combined")
    _header_row(ws6, 3, ["Height Category"] + ht_yr_hdrs)

    for i, cat in enumerate(["Stunted", "Normal", "Tall", "Unknown"], 4):
        row_vals = [cat]
        counts = []
        for y in years_present:
            sub_y = df_all[df_all["Tahun_Persekolahan"] == y]
            cnt = int((sub_y["Height_Category"] == cat).sum())
            tot = len(sub_y)
            row_vals += [cnt, f"{cnt/tot*100:.1f}%" if tot else "—"]
            counts.append(cnt)
        if len(years_present) > 1:
            row_vals.append(sum(counts))
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(row_vals, 1):
            _cell_style(ws6, i, c, v, bg=bg, align="center" if c > 1 else "left")

    _auto_col_width(ws6)

    # ── Sheet 7: Flagged Records (sample) ─────────────────────────────────
    ws7 = wb.create_sheet("Flagged Records")
    ws7.sheet_view.showGridLines = False

    ws7.merge_cells("A1:H1")
    ws7["A1"].value = "FLAGGED RECORDS — Sample (max 500 rows per issue type)"
    ws7["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws7["A1"].fill  = PatternFill("solid", fgColor="C00000")
    ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[1].height = 28

    flagged = df_all[df_all["Data_Quality_Flag"] != "Valid"].copy()

    sample_parts = []
    for flag_val in flagged["Data_Quality_Flag"].unique():
        sub = flagged[flagged["Data_Quality_Flag"] == flag_val]
        sample_parts.append(sub.head(min(500, len(sub))))
    sample_df = pd.concat(sample_parts, ignore_index=True) if sample_parts else pd.DataFrame()

    flag_cols = [
        "Tahun_Persekolahan", "Negeri", "ID_Murid",
        "Gender", "Tarikh_Lahir", "Age_At_Measurement",
        "Tarikh_Pengukuran", "Berat_kg", "Tinggi_cm",
        "BMI", "BMI_Category", "Data_Quality_Flag",
    ]

    _header_row(ws7, 3, flag_cols, bg="C00000")
    for i, (_, row) in enumerate(sample_df.iterrows(), 4):
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, col in enumerate(flag_cols, 1):
            val = row[col]
            if hasattr(val, "date"):
                val = val.date() if not pd.isna(val) else None
            _cell_style(ws7, i, c, val, bg=bg)

    _auto_col_width(ws7)

    return wb


# ---------------------------------------------------------------------------
# Cleaned data Excel builder  (pandas ExcelWriter for speed on large data)
# ---------------------------------------------------------------------------

def _apply_header_style(ws, n_cols: int):
    """Style the first row of a worksheet as a formatted header."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 40)


def build_cleaned_csv(df_all: pd.DataFrame, out_folder: Path) -> list[Path]:
    """
    Write cleaned data as CSV files (one per year) — much faster than Excel
    for large datasets.  Returns the list of paths written.
    """
    # Convert bool columns to readable strings
    bool_cols = ["Has_Complete_Measurements", "Is_Valid_Age",
                 "Is_Valid_Measurement_Date", "Is_Duplicate_ID"]
    df_write = df_all.copy()
    for col in bool_cols:
        if col in df_write.columns:
            df_write[col] = df_write[col].astype(str)

    # Format date columns as YYYY-MM-DD strings (avoids locale issues in CSV)
    for col in ["Tarikh_Lahir", "Tarikh_Pengukuran"]:
        if col in df_write.columns:
            df_write[col] = pd.to_datetime(df_write[col], errors="coerce").dt.strftime("%Y-%m-%d")

    paths_written: list[Path] = []
    for year in [2024, 2025]:
        sub = df_write[df_write["Tahun_Persekolahan"] == year]
        csv_path = out_folder / f"KKM_BeratTinggi_Cleaned_{year}.csv"
        print(f"  Writing {len(sub):,} rows → {csv_path.name} ...")
        sub.to_csv(csv_path, index=False, encoding="utf-8-sig")   # utf-8-sig for Excel-friendly BOM
        paths_written.append(csv_path)
        print(f"  Saved : {csv_path}")

    return paths_written


# ---------------------------------------------------------------------------
# Interactive file-picker helpers
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    """Open a Tk file-open dialog and return the chosen path (or None)."""
    root = tk.Tk()
    root.withdraw()          # hide the blank Tk window
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title,
        initialdir=initial_dir,
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    """Open a Tk folder-select dialog and return the chosen path (or None)."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    default_dir = str(BASE_DIR)

    print("=" * 60)
    print("  KKM Berat & Tinggi — Data Cleaning Tool")
    print("=" * 60)
    print()

    # ── Step 1: Choose which years to process ──────────────────────────────
    print("Step 1: Which year(s) do you want to process?")
    print("  [1] 2024 only")
    print("  [2] 2025 only")
    print("  [3] Both 2024 and 2025")
    choice = input("  Enter choice (1/2/3): ").strip()
    if choice not in ("1", "2", "3"):
        print("  ✗ Invalid choice. Exiting.")
        return
    years_to_run = {"1": [2024], "2": [2025], "3": [2024, 2025]}[choice]
    print(f"  ✓ Processing: {years_to_run}")

    # ── Step 2: Select input files ─────────────────────────────────────────
    file_map: dict[int, Path] = {}
    for step_num, year in enumerate(years_to_run, 2):
        print()
        print(f"Step {step_num}: Select the {year} data file (.xlsx)")
        initial = str(file_map[years_to_run[0]].parent) if file_map else default_dir
        chosen = _pick_file(f"Select {year} data file", initial)
        if not chosen:
            print(f"  ✗ No file selected for {year}. Exiting.")
            return
        file_map[year] = chosen
        print(f"  ✓ {year} file : {chosen}")

    # ── Step 3: Select output folder ──────────────────────────────────────
    step_out = len(years_to_run) + 2
    print()
    print(f"Step {step_out}: Select the output folder for the generated reports")
    first_file = file_map[years_to_run[0]]
    out_folder = _pick_folder("Select output folder", str(first_file.parent))
    if not out_folder:
        out_folder = first_file.parent / "Report"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output dir: {out_folder}")

    year_tag     = "_".join(str(y) for y in years_to_run)
    output_report = out_folder / f"KKM_BeratTinggi_QualityReport_{year_tag}.xlsx"

    # ── Detect / pick sheet name ───────────────────────────────────────────
    REQUIRED_COLS = {"NEGERI", "JANTINA", "ID_MURID", "BERAT (KG)", "TINGGI (CM)"}

    def _get_sheet(path: Path, year: int) -> str | None:
        """
        Auto-detect the data sheet. If the sheet contains a year digit or
        'DATA' keyword it is preferred. If multiple candidates exist or the
        auto pick looks wrong, show a Tk listbox so the user can choose.
        Returns None if the user cancels.
        """
        xl = pd.ExcelFile(path)
        sheets = xl.sheet_names

        # Try to find a sheet whose name contains the year or is purely numeric
        candidates = [s for s in sheets
                      if str(year) in s or s.strip().isdigit()]
        if len(candidates) == 1:
            return candidates[0]

        # Fall back: validate each sheet by peeking at column names
        for s in sheets:
            try:
                peek = xl.parse(s, nrows=2)
                cols = set(peek.columns.str.replace(r"\s+", " ", regex=True).str.strip().str.upper())
                if REQUIRED_COLS.issubset(cols):
                    return s
            except Exception:
                continue

        # Nothing matched — ask the user to pick
        print(f"  Could not auto-detect sheet in {path.name}.")
        print(f"  Available sheets: {sheets}")
        choice = _pick_sheet_dialog(path.name, sheets)
        return choice

    def _pick_sheet_dialog(filename: str, sheets: list[str]) -> str | None:
        """Show a small Tk window with a listbox of sheet names."""
        result: list[str | None] = [None]

        root = tk.Tk()
        root.title(f"Select sheet — {filename}")
        root.attributes("-topmost", True)
        root.resizable(False, False)

        tk.Label(root, text=f"Could not auto-detect the data sheet in:\n{filename}\n\nPlease select the correct sheet:",
                 wraplength=400, justify="left", padx=10, pady=6).pack()

        lb = tk.Listbox(root, selectmode=tk.SINGLE, width=60, height=min(len(sheets), 12))
        for s in sheets:
            lb.insert(tk.END, s)
        lb.select_set(0)
        lb.pack(padx=10, pady=4)

        def _ok():
            sel = lb.curselection()
            if sel:
                result[0] = sheets[sel[0]]
            root.destroy()

        def _cancel():
            root.destroy()

        btn_frame = tk.Frame(root)
        tk.Button(btn_frame, text="OK", width=10, command=_ok).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_frame, text="Cancel", width=10, command=_cancel).pack(side=tk.LEFT, padx=4)
        btn_frame.pack(pady=6)

        root.mainloop()
        return result[0]

    # ── Process each year ──────────────────────────────────────────────────
    # 2025: drop invalid rows entirely
    # 2024: flag but keep invalid rows
    DROP_INVALID_YEARS = {2025}

    all_stats: list[dict] = []
    cleaned_frames: list[pd.DataFrame] = []

    for year in years_to_run:
        path  = file_map[year]
        sheet = _get_sheet(path, year)
        if sheet is None:
            print(f"  ✗ No sheet selected for {year}. Skipping.")
            continue
        drop  = year in DROP_INVALID_YEARS

        print()
        print(f"[{year}] Loading {path.name}  (sheet: {sheet}) ...")
        if drop:
            print(f"  Mode: DROP invalid rows")
        else:
            print(f"  Mode: FLAG invalid rows (keep all)")

        raw = pd.read_excel(path, sheet_name=sheet)
        print(f"  Raw rows: {len(raw):,}")

        # ── Validate columns before proceeding ────────────────────────────
        raw_cols = set(raw.columns.str.replace(r"\s+", " ", regex=True).str.strip().str.upper())
        missing = REQUIRED_COLS - raw_cols
        if missing:
            root2 = tk.Tk(); root2.withdraw()
            messagebox.showerror(
                "Wrong file",
                f"The selected file for {year} does not look like a KKM data file.\n\n"
                f"Missing columns: {', '.join(sorted(missing))}\n\n"
                f"File: {path.name}\nSheet: {sheet}\n\n"
                f"Please re-run and select the correct file\n"
                f"(e.g. 'Data berat dan tinggi {year}_7tahun_KKM.xlsx')."
            )
            root2.destroy()
            return

        cleaned, stats = clean_dataset(raw, year, drop_invalid=drop)
        all_stats.append(stats)
        cleaned_frames.append(cleaned)

        print(f"  Rows after cleaning : {stats['rows_after_clean']:,}")
        print(f"  Rows dropped total  : {stats['total_rows_dropped']:,}")
        if drop:
            print(f"    ├ bad date        : {stats['rows_dropped_bad_date']:,}")
            print(f"    ├ berat OOR       : {stats['rows_dropped_berat']:,}")
            print(f"    └ tinggi OOR      : {stats['rows_dropped_tinggi']:,}")
        print(f"  Valid records       : {stats['valid_records']:,}")
        print(f"  Flagged records     : {stats['flagged_records']:,}")
        print(f"  Berat out-of-range  : {stats['berat_out_of_range']:,}")
        print(f"  Tinggi out-of-range : {stats['tinggi_out_of_range']:,}")
        print(f"  Invalid dates       : {stats['unix_epoch_dates']:,} epoch  |  {stats['pre2020_dates']:,} pre-2020  |  {stats['future_dates']:,} future")
        print(f"  Duplicate IDs       : {stats['duplicate_ic']:,}")

    df_all = pd.concat(cleaned_frames, ignore_index=True)
    print(f"\nTotal rows in output: {len(df_all):,}")

    # ── Quality Report ─────────────────────────────────────────────────────
    print(f"\nBuilding quality report ...")
    qr_wb = build_quality_report(all_stats, df_all)
    qr_wb.save(output_report)
    print(f"  Saved: {output_report}")

    # ── Cleaned Data (CSV) ─────────────────────────────────────────────────
    print(f"\nSaving cleaned data as CSV ...")
    csv_paths = build_cleaned_csv(df_all, out_folder)

    print()
    print("=" * 60)
    print("  Done! Output files:")
    print(f"    {output_report}")
    for p in csv_paths:
        print(f"    {p}")
    print("=" * 60)

    # Completion popup
    file_list = "• " + output_report.name + "\n" + "\n".join(f"• {p.name}" for p in csv_paths)
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Done", f"Cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n{file_list}")
    root.destroy()


if __name__ == "__main__":
    main()
