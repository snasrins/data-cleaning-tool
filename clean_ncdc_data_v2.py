"""
NCDC (TASKA) Data Cleaning Script — Standardized Version v2
============================================================
Cleans Status Kesihatan Kanak-Kanak Kebangsaan (NCDC) data and generates:
  1. NCDC_Cleaned.csv             — Cleansed data with WHO z-scores & indicators
  2. NCDC_QualityReport.xlsx      — Data quality report

Source: Multi-sheet Excel file (one sheet per Negeri) with wide format
        Year columns: 2023/2024/2025 Berat (kg), Tinggi (cm), Tarikh Pengukuran

DATA OWNER CLARIFICATIONS IMPLEMENTED:
  Q-01: Keep non-standard MyKid (passport numbers for foreign children)
  Q-02: Exclude Pendapatan Keluarga = 'X' (tiada maklumat)
  Q-03: Rename 'P20' to 'Miskin Tegar'
  Q-05: Exclude age > 5 years (NCDC covers 0-5 only)
  Q-06: Recalculate BMI from Berat/Tinggi (fix impossible values)
  Q-07: Keep most recent record for duplicate MyKid (same child, multiple agencies)
  Q-08: One row per child — take most recent if multiple entries

Cleaning rules applied (DROP invalid rows):
  Rule 1  : Drop Jantina not in (Lelaki, Perempuan)
  Rule 2  : Drop biologically implausible weight (≤0.5 or >35 kg) / height (≤30 or >130 cm)
  Rule 3  : Drop age at assessment > 5 years (≥60 months) or < 0
  Rule 4  : Drop records where measurement date < DOB
  Rule 5  : Drop implausible BMI > 40
  Rule 6  : Drop records where both weight AND height are null
  Rule 7  : Drop records where ANY z-score is null
  Rule 8  : Drop duplicate MyKid (keep most recent measurement date)
  Rule 9  : Exclude Pendapatan Keluarga = 'X'

NCDC 4 Core Indicators (WHO standards):
  1. Kurang Berat Badan  — WAZ < -2SD (includes Teruk < -3SD)
  2. Bantut              — HAZ < -2SD (includes Teruk < -3SD)
  3. Susut               — BAZ < -2SD (includes Teruk < -3SD)
  4. Berlebihan BB/Obes  — BAZ > +1SD (Berisiko), > +2SD (Berlebihan), > +3SD (Obes)

2 Age Groups:
  - Bawah 2 Tahun (Under 2): 0-23 months
  - Bawah 5 Tahun (Under 5): 0-59 months

3 Geographic Levels:
  - Negeri (15 states)
  - Bahagian (for Sabah: 6, Sarawak: 12) — derived from Daerah
  - Daerah (160 districts)
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path for WHO z-score module ─────────────────────────────
_BACKEND = Path(__file__).parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore_daily import compute_zscore, classify_waz, classify_haz, classify_baz
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height bounds (biologically plausible for 0–5 years)
BERAT_MIN, BERAT_MAX = 0.5, 35.0      # kg
TINGGI_MIN, TINGGI_MAX = 30.0, 130.0  # cm
BMI_MAX = 40.0

# Age bounds
AGE_MAX_MONTHS = 60  # under 5 years

# WHO BIV thresholds
BIV = {
    "WAZ": (-6.0, +5.0),
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}
_BIV_LO_WAZ, _BIV_HI_WAZ = BIV["WAZ"]
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]

# Gender mapping (standardize to English)
GENDER_MAP = {
    "LELAKI": "Male",
    "PEREMPUAN": "Female",
    "L": "Male",
    "P": "Female",
    "M": "Male",
    "F": "Female",
    "MALE": "Male",
    "FEMALE": "Female",
}

# Pendapatan Keluarga mapping (Q-03: P20 → Miskin Tegar)
PENDAPATAN_MAP = {
    "B40": "B40",
    "M40": "M40",
    "T20": "T20",
    "P20": "Miskin Tegar",  # Q-03 clarification
}

# Bahagian mapping for Sabah (6 divisions)
SABAH_BAHAGIAN = {
    # Bahagian Pantai Barat
    "KOTA KINABALU": "Pantai Barat", "PENAMPANG": "Pantai Barat", "PUTATAN": "Pantai Barat",
    "PAPAR": "Pantai Barat", "TUARAN": "Pantai Barat", "KOTA BELUD": "Pantai Barat",
    # Bahagian Pedalaman
    "RANAU": "Pedalaman", "KUNDASANG": "Pedalaman", "TAMBUNAN": "Pedalaman",
    "KENINGAU": "Pedalaman", "TENOM": "Pedalaman", "NABAWAN": "Pedalaman",
    "SOOK": "Pedalaman", "KEMABONG": "Pedalaman",
    # Bahagian Kudat
    "KUDAT": "Kudat", "KOTA MARUDU": "Kudat", "PITAS": "Kudat", "MATUNGGONG": "Kudat",
    # Bahagian Sandakan
    "SANDAKAN": "Sandakan", "KINABATANGAN": "Sandakan", "BELURAN": "Sandakan",
    "TONGOD": "Sandakan", "TELUPID": "Sandakan",
    # Bahagian Tawau
    "TAWAU": "Tawau", "LAHAD DATU": "Tawau", "SEMPORNA": "Tawau", "KUNAK": "Tawau",
    # Bahagian Beaufort
    "BEAUFORT": "Beaufort", "KUALA PENYU": "Beaufort", "SIPITANG": "Beaufort", "MEMBAKUT": "Beaufort",
}

# Bahagian mapping for Sarawak (12 divisions)
SARAWAK_BAHAGIAN = {
    # Bahagian Kuching
    "KUCHING": "Kuching", "BAU": "Kuching", "LUNDU": "Kuching", "PADAWAN": "Kuching",
    # Bahagian Samarahan
    "KOTA SAMARAHAN": "Samarahan", "SAMARAHAN": "Samarahan", "ASAJAYA": "Samarahan", 
    "SIMUNJAN": "Samarahan",
    # Bahagian Serian
    "SERIAN": "Serian", "TEBEDU": "Serian",
    # Bahagian Sri Aman
    "SRI AMAN": "Sri Aman", "LUBOK ANTU": "Sri Aman", "ENGKILILI": "Sri Aman",
    # Bahagian Betong
    "BETONG": "Betong", "PUSA": "Betong", "SARATOK": "Betong", "KABONG": "Betong",
    # Bahagian Sarikei
    "SARIKEI": "Sarikei", "JULAU": "Sarikei", "MERADONG": "Sarikei", "PAKAN": "Sarikei",
    # Bahagian Sibu
    "SIBU": "Sibu", "SELANGAU": "Sibu", "KANOWIT": "Sibu",
    # Bahagian Mukah
    "MUKAH": "Mukah", "DALAT": "Mukah", "DARO": "Mukah", "MATU": "Mukah", "TANJUNG MANIS": "Mukah",
    # Bahagian Kapit
    "KAPIT": "Kapit", "SONG": "Kapit", "BELAGA": "Kapit", "BUKIT MABONG": "Kapit",
    # Bahagian Bintulu
    "BINTULU": "Bintulu", "SEBAUH": "Bintulu", "TATAU": "Bintulu",
    # Bahagian Miri
    "MIRI": "Miri", "MARUDI": "Miri", "SUBIS": "Miri", "TELANG USAN": "Miri", "BELURU": "Miri",
    # Bahagian Limbang
    "LIMBANG": "Limbang", "LAWAS": "Limbang",
}


# ---------------------------------------------------------------------------
# Label functions (standardized Malay)
# ---------------------------------------------------------------------------

def _waz_label(z):
    """WAZ status label per WHO guidelines."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Kurang Berat Badan Teruk"
    if z < -2:  return "Kurang Berat Badan"
    if z <= 2:  return "Berat Badan Normal"
    return "Mungkin Masalah Pertumbuhan"


def _haz_label(z):
    """HAZ status label per WHO guidelines."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Bantut Teruk"
    if z < -2:  return "Bantut"
    if z <= 3:  return "Panjang/Tinggi Normal"
    return "Mungkin Masalah Endokrin"


def _baz_label(z):
    """BAZ status label per WHO guidelines."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Susut Teruk"
    if z < -2:  return "Susut"
    if z <= 1:  return "Berat Badan Normal"
    if z <= 2:  return "Berisiko Berlebihan Berat Badan"
    if z <= 3:  return "Berlebihan Berat Badan"
    return "Obes"


def _get_bahagian(negeri: str, daerah: str) -> str | None:
    """Derive Bahagian from Negeri and Daerah for Sabah/Sarawak."""
    negeri_upper = str(negeri).upper().strip()
    daerah_upper = str(daerah).upper().strip()
    
    if negeri_upper == "SABAH":
        return SABAH_BAHAGIAN.get(daerah_upper, "Unknown")
    elif negeri_upper == "SARAWAK":
        return SARAWAK_BAHAGIAN.get(daerah_upper, "Unknown")
    return None  # Not applicable for other states


def _get_age_group(age_months: float) -> str:
    """Classify age into Bawah 2 Tahun or Bawah 5 Tahun."""
    if age_months is None or math.isnan(age_months):
        return "Unknown"
    if age_months < 24:
        return "Bawah 2 Tahun"
    elif age_months < 60:
        return "2-5 Tahun"
    return "Lebih 5 Tahun"


# ---------------------------------------------------------------------------
# Load NCDC data (ALL years - reshape wide to long)
# ---------------------------------------------------------------------------

def load_ncdc_all_years(file_path: Path) -> pd.DataFrame:
    """
    Load NCDC Excel file and extract ALL years (2023, 2024, 2025).
    Wide format columns: "YYYY Berat (kg)", "YYYY Tinggi (cm)", "YYYY Tarikh Pengukuran"
    Convert to long format with Year column.
    """
    xl = pd.ExcelFile(file_path)
    all_frames = []
    years_to_extract = [2023, 2024, 2025]

    print(f"  Found {len(xl.sheet_names)} sheets")

    for sheet_name in xl.sheet_names:
        print(f"    Loading sheet: {sheet_name} ...")
        df_sheet = xl.parse(sheet_name)
        if df_sheet.empty:
            continue

        # Standardize column names
        df_sheet.columns = df_sheet.columns.str.strip()
        
        print(f"      Columns: {list(df_sheet.columns)}")
        print(f"      Rows: {len(df_sheet):,}")

        # Base columns mapping (non-year specific)
        base_col_map = {
            "Agensi": "Agensi",
            "Negeri": "Negeri",
            "Daerah": "Daerah",
            "Nama TASKA": "Nama_Taska",
            "Nama Anak": "Nama_Anak",
            "No. Mykid": "MyKid",
            "Pendapatan Keluarga": "Pendapatan_Keluarga",
            "Kumpulan Umur": "Kumpulan_Umur_Source",
            "Tarikh Lahir": "Tarikh_Lahir",
            "Jantina": "Jantina_Raw",
        }
        
        # Find year-specific columns
        # Format: "2023 Berat (kg)", "2024 Berat (kg)", "2025 Berat (kg)"
        year_cols = {}
        for year in years_to_extract:
            year_cols[year] = {"berat": None, "tinggi": None, "tarikh": None}
            for col in df_sheet.columns:
                col_str = str(col)
                # Check if column starts with year
                if col_str.startswith(str(year)):
                    col_lower = col_str.lower()
                    if "berat" in col_lower and "status" not in col_lower:
                        year_cols[year]["berat"] = col
                    elif "tinggi" in col_lower and "status" not in col_lower:
                        year_cols[year]["tinggi"] = col
                    elif "tarikh" in col_lower:
                        year_cols[year]["tarikh"] = col

        # Show found columns
        for year in years_to_extract:
            cols = year_cols[year]
            if cols["berat"] or cols["tinggi"]:
                print(f"        {year}: Berat={cols['berat']}, Tinggi={cols['tinggi']}, Tarikh={cols['tarikh']}")

        # Build records for each year
        sheet_records = []
        for _, row in df_sheet.iterrows():
            # Extract base info once
            base_rec = {}
            for src, dst in base_col_map.items():
                if src in df_sheet.columns:
                    base_rec[dst] = row[src]
            
            # For each year, create a record if there's data
            for year in years_to_extract:
                cols = year_cols[year]
                
                # Get measurements for this year
                berat_val = row[cols["berat"]] if cols["berat"] else None
                tinggi_val = row[cols["tinggi"]] if cols["tinggi"] else None
                tarikh_val = row[cols["tarikh"]] if cols["tarikh"] else None
                
                # Check if valid measurements exist
                has_berat = pd.notna(berat_val) and berat_val != "" and berat_val != 0
                has_tinggi = pd.notna(tinggi_val) and tinggi_val != "" and tinggi_val != 0
                
                # Skip if no measurements for this year
                if not has_berat and not has_tinggi:
                    continue
                
                # Create record
                rec = {**base_rec, "Year": year}
                
                if has_berat:
                    rec["Berat_kg"] = berat_val
                if has_tinggi:
                    rec["Tinggi_cm"] = tinggi_val
                if pd.notna(tarikh_val):
                    rec["Tarikh_Pengukuran"] = tarikh_val
                else:
                    # Default measurement date to end of year if missing
                    rec["Tarikh_Pengukuran"] = f"{year}-12-31"
                
                sheet_records.append(rec)
        
        if sheet_records:
            df_out = pd.DataFrame(sheet_records)
            all_frames.append(df_out)
            
            # Count by year
            year_counts = df_out["Year"].value_counts().sort_index()
            print(f"      → {len(df_out):,} total records: {year_counts.to_dict()}")
        else:
            print(f"      → 0 records found")

    if not all_frames:
        raise ValueError("No data found in any sheet. Check column names.")

    combined = pd.concat(all_frames, ignore_index=True)
    
    # Summary by year
    print(f"\n  Total records loaded: {len(combined):,}")
    year_summary = combined["Year"].value_counts().sort_index()
    for year, count in year_summary.items():
        print(f"    {year}: {count:,} records")
    
    return combined


# ---------------------------------------------------------------------------
# Cleaning function
# ---------------------------------------------------------------------------

def clean_ncdc(raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Clean NCDC data and compute WHO z-scores."""
    stats: dict = {"raw_count": len(raw)}
    df = raw.copy()

    # ── Rule 1: Standardize gender, drop invalid ──────────────────────────
    df["Jantina_Raw"] = df["Jantina_Raw"].astype(str).str.upper().str.strip()
    df["Gender"] = df["Jantina_Raw"].map(GENDER_MAP)

    before = len(df)
    invalid_gender = df["Gender"].isna()
    stats["dropped_invalid_gender"] = int(invalid_gender.sum())
    df = df[~invalid_gender].copy()
    print(f"    Rule 1: Dropped {stats['dropped_invalid_gender']:,} invalid gender")

    # Jantina in Malay
    df["Jantina"] = df["Gender"].map({"Male": "Lelaki", "Female": "Perempuan"})

    # ── Rule 9: Exclude Pendapatan = 'X' (Q-02) ───────────────────────────
    if "Pendapatan_Keluarga" in df.columns:
        df["Pendapatan_Keluarga"] = df["Pendapatan_Keluarga"].astype(str).str.upper().str.strip()
        before = len(df)
        pendapatan_x = df["Pendapatan_Keluarga"] == "X"
        stats["dropped_pendapatan_x"] = int(pendapatan_x.sum())
        df = df[~pendapatan_x].copy()
        print(f"    Rule 9: Dropped {stats['dropped_pendapatan_x']:,} Pendapatan = 'X'")
        
        # Map P20 to Miskin Tegar (Q-03)
        df["Pendapatan_Keluarga"] = df["Pendapatan_Keluarga"].map(
            lambda x: PENDAPATAN_MAP.get(x, x)
        )
        p20_renamed = (df["Pendapatan_Keluarga"] == "Miskin Tegar").sum()
        print(f"    Q-03: Renamed {p20_renamed:,} 'P20' to 'Miskin Tegar'")

    # ── Standardize text fields ───────────────────────────────────────────
    for col in ["Negeri", "Daerah", "Agensi"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().str.strip()

    for col in ["Nama_Taska", "Nama_Anak", "MyKid"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # ── Parse dates ───────────────────────────────────────────────────────
    df["Tarikh_Lahir"] = pd.to_datetime(df["Tarikh_Lahir"], dayfirst=True, errors="coerce")
    
    if "Tarikh_Pengukuran" in df.columns:
        df["Tarikh_Pengukuran"] = pd.to_datetime(df["Tarikh_Pengukuran"], dayfirst=True, errors="coerce")
    else:
        # Default to end of 2025 if no measurement date
        df["Tarikh_Pengukuran"] = pd.Timestamp("2025-12-31")

    # Remove time component
    df["Tarikh_Lahir"] = df["Tarikh_Lahir"].dt.normalize()
    df["Tarikh_Pengukuran"] = df["Tarikh_Pengukuran"].dt.normalize()

    # ── Drop rows with null DOB ───────────────────────────────────────────
    before = len(df)
    df = df[df["Tarikh_Lahir"].notna()].copy()
    stats["dropped_null_dob"] = before - len(df)
    print(f"    Dropped {stats['dropped_null_dob']:,} null DOB")

    # ── Rule 4: Drop where measurement date < DOB ─────────────────────────
    before = len(df)
    bad_date = df["Tarikh_Pengukuran"] < df["Tarikh_Lahir"]
    stats["dropped_date_before_dob"] = int(bad_date.sum())
    df = df[~bad_date].copy()
    print(f"    Rule 4: Dropped {stats['dropped_date_before_dob']:,} measurement before DOB")

    # ── Compute age at measurement (in DAYS for WHO z-score calculation) ─
    df["Age_Days"] = (df["Tarikh_Pengukuran"] - df["Tarikh_Lahir"]).dt.days
    df["Age_Months"] = (df["Age_Days"] / 30.4375).round(2)

    # ── Rule 3: Drop age >= 60 months (Q-05: exclude > 5 years) ───────────
    before = len(df)
    age_invalid = (df["Age_Days"] < 0) | (df["Age_Months"] >= (AGE_MAX_MONTHS))
    stats["dropped_age_invalid"] = int(age_invalid.sum())
    df = df[~age_invalid].copy()
    print(f"    Rule 3: Dropped {stats['dropped_age_invalid']:,} age >= 5 years or < 0")

    # Add age group
    df["Kumpulan_Umur"] = df["Age_Months"].apply(_get_age_group)

    # ── Numeric measurements ──────────────────────────────────────────────
    df["Berat_kg"] = pd.to_numeric(df["Berat_kg"], errors="coerce")
    df["Tinggi_cm"] = pd.to_numeric(df["Tinggi_cm"], errors="coerce")

    # ── Rule 2: Drop measurement outliers ─────────────────────────────────
    berat_bad = (df["Berat_kg"] < BERAT_MIN) | (df["Berat_kg"] > BERAT_MAX)
    tinggi_bad = (df["Tinggi_cm"] < TINGGI_MIN) | (df["Tinggi_cm"] > TINGGI_MAX)

    outlier_mask = (berat_bad & df["Berat_kg"].notna()) | (tinggi_bad & df["Tinggi_cm"].notna())
    stats["dropped_measurement_outlier"] = int(outlier_mask.sum())
    stats["berat_out_of_range"] = int((berat_bad & df["Berat_kg"].notna()).sum())
    stats["tinggi_out_of_range"] = int((tinggi_bad & df["Tinggi_cm"].notna()).sum())
    df = df[~outlier_mask].copy()
    print(f"    Rule 2: Dropped {stats['dropped_measurement_outlier']:,} measurement outliers")

    df["Berat_kg"] = df["Berat_kg"].round(1)
    df["Tinggi_cm"] = df["Tinggi_cm"].round(1)

    # ── Rule 6: Drop rows where both measurements are null ────────────────
    before = len(df)
    no_meas = df["Berat_kg"].isna() & df["Tinggi_cm"].isna()
    stats["dropped_no_measurement"] = int(no_meas.sum())
    df = df[~no_meas].copy()
    print(f"    Rule 6: Dropped {stats['dropped_no_measurement']:,} no measurements")

    # ── Recalculate BMI (Q-06: fix impossible values) ─────────────────────
    valid_both = df["Berat_kg"].notna() & df["Tinggi_cm"].notna() & (df["Tinggi_cm"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["Berat_kg"] / ((df["Tinggi_cm"] / 100) ** 2)).round(2),
        np.nan,
    )

    # ── Rule 5: Drop implausible BMI > 40 ─────────────────────────────────
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()
    print(f"    Rule 5: Dropped {stats['dropped_bmi_outlier']:,} BMI > 40")

    # ── Rule 8: Remove duplicate MyKid, keep most recent (Q-07, Q-08) ─────
    if "MyKid" in df.columns:
        before = len(df)
        df = df.sort_values("Tarikh_Pengukuran", ascending=False)
        df = df.drop_duplicates(subset=["MyKid"], keep="first")
        stats["dropped_duplicate_mykid"] = before - len(df)
        print(f"    Rule 8: Dropped {stats['dropped_duplicate_mykid']:,} duplicate MyKid (kept most recent)")
    else:
        stats["dropped_duplicate_mykid"] = 0

    # ── Add Bahagian for Sabah/Sarawak ────────────────────────────────────
    df["Bahagian"] = df.apply(
        lambda r: _get_bahagian(r.get("Negeri", ""), r.get("Daerah", "")), axis=1
    )

    # ── WHO Z-scores (using DAILY LMS tables) ────────────────────────────
    if _ZSCORE_AVAILABLE:
        print("    Computing WHO z-scores (daily LMS) ...")
        sex_s = df["Gender"].astype(str)
        age_s = df["Age_Days"]  # Use days instead of months
        berat_s = df["Berat_kg"]
        tinggi_s = df["Tinggi_cm"]
        bmi_s = df["BMI"]

        n = len(df)
        waz_vals, haz_vals, baz_vals = [], [], []

        for i in range(n):
            age_d = age_s.iat[i]
            sex = sex_s.iat[i]
            w = berat_s.iat[i]
            h = tinggi_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age_d) or pd.isna(sex):
                waz_vals.append(None)
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            # New signature: compute_zscore(indicator, value, age_days, sex)
            waz = compute_zscore("WAZ", float(w), int(age_d), sex) if pd.notna(w) else None
            haz = compute_zscore("HAZ", float(h), int(age_d), sex) if pd.notna(h) else None
            baz = compute_zscore("BAZ", float(bmi), int(age_d), sex) if pd.notna(bmi) else None

            # Apply BIV bounds
            if waz is not None and not (_BIV_LO_WAZ <= waz <= _BIV_HI_WAZ): waz = None
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            waz_vals.append(round(waz, 3) if waz is not None else None)
            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 5000 == 0:
                print(f"      Z-scores: {i+1:,}/{n:,} ...")

        df["WAZ"] = waz_vals
        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["WAZ_Status"] = df["WAZ"].apply(classify_waz)
        df["HAZ_Status"] = df["HAZ"].apply(classify_haz)
        df["BAZ_Status"] = df["BAZ"].apply(classify_baz)

        # Debug: Check gender distribution before dropping null z-scores
        print(f"    Gender distribution before Rule 7:")
        print(f"      Male: {(df['Gender'] == 'Male').sum():,}")
        print(f"      Female: {(df['Gender'] == 'Female').sum():,}")

        # NCDC 4 Core Indicator flags
        df["Ind_Kurang_Berat_Badan"] = df["WAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Bantut"] = df["HAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Susut"] = df["BAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Berlebihan_BB"] = df["BAZ"].apply(lambda z: z is not None and z > 1)
        df["Ind_Obes"] = df["BAZ"].apply(lambda z: z is not None and z > 2)

        # Ind_Normal: Not Kurang BB, Not Bantut, Not Susut, Not Berlebihan
        df["Ind_Normal"] = ~(
            df["Ind_Kurang_Berat_Badan"] | df["Ind_Bantut"] | 
            df["Ind_Susut"] | df["Ind_Berlebihan_BB"]
        )

        # ── Rule 7: Drop rows where ANY z-score is null ───────────────────
        before = len(df)
        null_zscore = df["WAZ"].isna() | df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore.sum())
        df = df[~null_zscore].copy()
        print(f"    Rule 7: Dropped {stats['dropped_null_zscore']:,} null z-scores")
        
        # Debug: Check gender distribution after Rule 7
        print(f"    Gender distribution after Rule 7:")
        print(f"      Male: {(df['Gender'] == 'Male').sum():,}")
        print(f"      Female: {(df['Gender'] == 'Female').sum():,}")

    else:
        for col in ["WAZ", "HAZ", "BAZ", "WAZ_Status", "HAZ_Status", "BAZ_Status",
                    "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", 
                    "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["ind_kurang_berat"] = int(df["Ind_Kurang_Berat_Badan"].sum())
        stats["ind_bantut"] = int(df["Ind_Bantut"].sum())
        stats["ind_susut"] = int(df["Ind_Susut"].sum())
        stats["ind_berlebihan_bb"] = int(df["Ind_Berlebihan_BB"].sum())
        stats["ind_obes"] = int(df["Ind_Obes"].sum())
        stats["ind_normal"] = int(df["Ind_Normal"].sum())

        # Age group stats
        stats["bawah_2_tahun"] = int((df["Kumpulan_Umur"] == "Bawah 2 Tahun").sum())
        stats["bawah_5_tahun"] = len(df)  # All are < 5 years after Rule 3

    # Year breakdown
    if "Year" in df.columns:
        stats["year_breakdown"] = df["Year"].value_counts().sort_index().to_dict()
    
    # Gender breakdown
    if "Gender" in df.columns:
        stats["gender_breakdown"] = df["Gender"].value_counts().to_dict()
        print(f"\n  Final Gender Distribution:")
        for gender, count in stats["gender_breakdown"].items():
            print(f"    {gender}: {count:,}")

    return df, stats


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr


def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(stats: dict, df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "NCDC (TASKA) — DATA QUALITY REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:D3")
    ws1["A3"].value = (
        "Data Source: Status Kesihatan Kanak-Kanak Kebangsaan (NCDC/TASKA) | "
        "All years (2023, 2024, 2025)"
    )
    ws1["A3"].font = Font(italic=True, size=9, color="7F7F7F")
    ws1["A3"].alignment = Alignment(horizontal="left", wrap_text=True)

    _hdr(ws1, 5, ["Metric", "Count", "Notes"])

    summary_rows = [
        ("Raw records (all years)", stats["raw_count"], "Source Excel"),
        ("Dropped — Invalid Gender", stats.get("dropped_invalid_gender", 0), "Rule 1"),
        ("Dropped — Pendapatan = 'X'", stats.get("dropped_pendapatan_x", 0), "Rule 9 (Q-02)"),
        ("Dropped — Null DOB", stats.get("dropped_null_dob", 0), "Required for age"),
        ("Dropped — Measurement before DOB", stats.get("dropped_date_before_dob", 0), "Rule 4"),
        ("Dropped — Age invalid (≥5 yrs)", stats.get("dropped_age_invalid", 0), "Rule 3 (Q-05)"),
        ("Dropped — Measurement outliers", stats.get("dropped_measurement_outlier", 0), "Rule 2"),
        ("Dropped — No measurement", stats.get("dropped_no_measurement", 0), "Rule 6"),
        ("Dropped — BMI > 40", stats.get("dropped_bmi_outlier", 0), "Rule 5 (Q-06)"),
        ("Dropped — Duplicate MyKid", stats.get("dropped_duplicate_mykid", 0), "Rule 8 (Q-07)"),
        ("Dropped — Null z-score(s)", stats.get("dropped_null_zscore", 0), "Rule 7"),
        ("Final cleaned records", stats["final_count"], "After all rules"),
    ]
    for i, (label, val, note) in enumerate(summary_rows, 6):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        _cell(ws1, i, 1, label, bg=bg)
        _cell(ws1, i, 2, val, bg=bg, align="right")
        _cell(ws1, i, 3, note, bg=bg)

    row = len(summary_rows) + 7

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["NCDC 4 Core Indicators", "Total", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        final = stats["final_count"]
        ind_rows = [
            ("1. Kurang Berat Badan (WAZ < -2SD)", stats.get("ind_kurang_berat", 0),
             f"{stats.get('ind_kurang_berat', 0)/max(final,1)*100:.1f}%", "WAZ < -2"),
            ("2. Bantut (HAZ < -2SD)", stats.get("ind_bantut", 0),
             f"{stats.get('ind_bantut', 0)/max(final,1)*100:.1f}%", "HAZ < -2"),
            ("3. Susut (BAZ < -2SD)", stats.get("ind_susut", 0),
             f"{stats.get('ind_susut', 0)/max(final,1)*100:.1f}%", "BAZ < -2"),
            ("4. Berlebihan BB (BAZ > +1SD)", stats.get("ind_berlebihan_bb", 0),
             f"{stats.get('ind_berlebihan_bb', 0)/max(final,1)*100:.1f}%", "BAZ > +1 (includes Obes)"),
            ("   └─ Obes (BAZ > +2SD)", stats.get("ind_obes", 0),
             f"{stats.get('ind_obes', 0)/max(final,1)*100:.1f}%", "BAZ > +2"),
            ("Normal (healthy child)", stats.get("ind_normal", 0),
             f"{stats.get('ind_normal', 0)/max(final,1)*100:.1f}%", "All indicators FALSE"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

        row = row + len(ind_rows) + 2
        _hdr(ws1, row, ["Age Group", "Count", "% of Total"], bg="548235")
        row += 1
        age_rows = [
            ("Bawah 2 Tahun (0-23 months)", stats.get("bawah_2_tahun", 0),
             f"{stats.get('bawah_2_tahun', 0)/max(final,1)*100:.1f}%"),
            ("Bawah 5 Tahun (0-59 months)", final, "100.0%"),
        ]
        for i, r in enumerate(age_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

    _auto_width(ws1)

    # ── Sheet 2: By Negeri ────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Negeri")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:H1")
    ws2["A1"].value = "DISTRIBUTION BY NEGERI"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _hdr(ws2, 3, ["Negeri", "Total", "Kurang BB", "Bantut", "Susut", "Berlebihan BB", "Obes", "Normal"])

    if "Negeri" in df.columns:
        grp = df.groupby("Negeri")
        for i, (neg, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                neg, len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                int(sub["Ind_Obes"].sum()) if "Ind_Obes" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws2, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws2)

    # ── Sheet 3: By Bahagian (Sabah/Sarawak) ──────────────────────────────
    ws3 = wb.create_sheet("By Bahagian")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    ws3["A1"].value = "DISTRIBUTION BY BAHAGIAN (Sabah & Sarawak only)"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _hdr(ws3, 3, ["Negeri", "Bahagian", "Total", "Kurang BB", "Bantut", "Susut", "Berlebihan BB", "Normal"])

    if "Bahagian" in df.columns:
        df_bah = df[df["Bahagian"].notna() & (df["Bahagian"] != "")]
        if len(df_bah) > 0:
            grp = df_bah.groupby(["Negeri", "Bahagian"])
            row_num = 4
            for (neg, bah), sub in sorted(grp, key=lambda x: (x[0][0], -len(x[1]))):
                bg = ALT if row_num % 2 == 0 else "FFFFFF"
                row_vals = [
                    neg, bah, len(sub),
                    int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                    int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                    int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                    int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                    int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
                ]
                for c, v in enumerate(row_vals, 1):
                    _cell(ws3, row_num, c, v, bg=bg, align="right" if c > 2 else "left")
                row_num += 1

    _auto_width(ws3)

    # ── Sheet 4: By Age Group ─────────────────────────────────────────────
    ws4 = wb.create_sheet("By Age Group")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:H1")
    ws4["A1"].value = "DISTRIBUTION BY AGE GROUP"
    ws4["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    _hdr(ws4, 3, ["Kumpulan Umur", "Total", "Kurang BB", "Bantut", "Susut", "Berlebihan BB", "Obes", "Normal"])

    if "Kumpulan_Umur" in df.columns:
        grp = df.groupby("Kumpulan_Umur")
        row_num = 4
        for ku, sub in sorted(grp, key=lambda x: x[0]):
            bg = ALT if row_num % 2 == 0 else "FFFFFF"
            row_vals = [
                ku, len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                int(sub["Ind_Obes"].sum()) if "Ind_Obes" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws4, row_num, c, v, bg=bg, align="right" if c > 1 else "left")
            row_num += 1

    _auto_width(ws4)

    # ── Sheet 5: By Year ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("By Year")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:H1")
    ws5["A1"].value = "DISTRIBUTION BY YEAR"
    ws5["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws5["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    _hdr(ws5, 3, ["Year", "Total", "Kurang BB", "Bantut", "Susut", "Berlebihan BB", "Obes", "Normal"])

    if "Year" in df.columns:
        grp = df.groupby("Year")
        row_num = 4
        for year, sub in sorted(grp, key=lambda x: x[0]):
            bg = ALT if row_num % 2 == 0 else "FFFFFF"
            row_vals = [
                int(year), len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                int(sub["Ind_Obes"].sum()) if "Ind_Obes" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws5, row_num, c, v, bg=bg, align="right" if c > 1 else "left")
            row_num += 1

    _auto_width(ws5)

    return wb


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def save_cleaned_csv(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()

    # Format dates as YYYY-MM-DD
    for col in ["Tarikh_Lahir", "Tarikh_Pengukuran"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Booleans as strings
    for col in ["Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", 
                "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)

    # Select output columns in standardized order
    out_cols = [
        "Year", "Agensi", "Negeri", "Bahagian", "Daerah", "Nama_Taska",
        "MyKid", "Nama_Anak", "Gender", "Jantina",
        "Pendapatan_Keluarga", "Kumpulan_Umur",
        "Tarikh_Lahir", "Tarikh_Pengukuran", "Age_Days", "Age_Months",
        "Berat_kg", "Tinggi_cm", "BMI",
        "WAZ", "HAZ", "BAZ",
        "WAZ_Status", "HAZ_Status", "BAZ_Status",
        "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut",
        "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal",
    ]
    out_cols = [c for c in out_cols if c in df_out.columns]
    df_out = df_out[out_cols]

    # Sort by Gender, Year, Negeri for organized output
    sort_cols = []
    if "Gender" in df_out.columns:
        sort_cols.append("Gender")
    if "Year" in df_out.columns:
        sort_cols.append("Year")
    if "Negeri" in df_out.columns:
        sort_cols.append("Negeri")
    
    if sort_cols:
        df_out = df_out.sort_values(sort_cols)

    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  Saved: {out_path}")
    print(f"  Records by Gender: {df_out['Gender'].value_counts().to_dict() if 'Gender' in df_out.columns else 'N/A'}")


# ---------------------------------------------------------------------------
# File picker
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=initial_dir,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("  NCDC (TASKA) Data Cleaning Tool v2")
    print("  Status Kesihatan Kanak-Kanak Kebangsaan — All Years (2023-2025)")
    print("=" * 70)

    # Step 1: Select input file
    print("\nStep 1: Select the NCDC Excel file (.xlsx)")
    chosen = _pick_file("Select NCDC data file", str(BASE_DIR))
    if not chosen:
        print("  ✗ No file selected. Exiting.")
        return
    print(f"  ✓ File: {chosen}")

    # Step 2: Select output folder
    print("\nStep 2: Select the output folder")
    out_folder = _pick_folder("Select output folder", str(chosen.parent))
    if not out_folder:
        out_folder = chosen.parent / "NCDC_Cleaned"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output: {out_folder}")

    # Load data (ALL years)
    print(f"\n{'='*70}")
    print("Loading NCDC data (all years: 2023, 2024, 2025) ...")
    raw = load_ncdc_all_years(chosen)

    # Clean data
    print(f"\n{'='*70}")
    print("Applying cleaning rules ...")
    df, stats = clean_ncdc(raw)

    print(f"\n  ───────────────────────────────────────────────────────────")
    print(f"  Total dropped : {stats['total_dropped']:,}  ({stats['total_dropped']/stats['raw_count']*100:.1f}%)")
    print(f"  Final records : {stats['final_count']:,}")

    if _ZSCORE_AVAILABLE:
        print(f"\n  NCDC 4 Core Indicators:")
        print(f"    1. Kurang Berat Badan (WAZ < -2) : {stats.get('ind_kurang_berat', 0):,}")
        print(f"    2. Bantut (HAZ < -2)             : {stats.get('ind_bantut', 0):,}")
        print(f"    3. Susut (BAZ < -2)              : {stats.get('ind_susut', 0):,}")
        print(f"    4. Berlebihan BB (BAZ > +1)      : {stats.get('ind_berlebihan_bb', 0):,}")
        print(f"       └─ Obes (BAZ > +2)            : {stats.get('ind_obes', 0):,}")
        print(f"    Normal                           : {stats.get('ind_normal', 0):,}")
        print(f"\n  Age Groups:")
        print(f"    Bawah 2 Tahun (0-23 mo) : {stats.get('bawah_2_tahun', 0):,}")
        print(f"    Bawah 5 Tahun (0-59 mo) : {stats['final_count']:,}")
        
        # Year breakdown
        if "year_breakdown" in stats:
            print(f"\n  By Year:")
            for year, count in sorted(stats["year_breakdown"].items()):
                print(f"    {year}: {count:,}")

    # Save quality report
    print(f"\n{'='*70}")
    print("Building quality report ...")
    out_report = out_folder / "NCDC_QualityReport.xlsx"
    qr = build_quality_report(stats, df)
    qr.save(out_report)
    print(f"  Saved: {out_report}")

    # Save cleaned CSV
    print("\nSaving cleaned CSV ...")
    out_csv = out_folder / "NCDC_Cleaned.csv"
    save_cleaned_csv(df, out_csv)

    print()
    print("=" * 70)
    print("  Done! Files saved to:")
    print(f"    {out_report}")
    print(f"    {out_csv}")
    print("=" * 70)

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(
        "Done",
        f"NCDC cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n"
        f"• NCDC_Cleaned.csv\n"
        f"• NCDC_QualityReport.xlsx",
    )
    root.destroy()


if __name__ == "__main__":
    main()
