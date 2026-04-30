"""
KPM (KKM) Data Cleaning Script — Standardized Version v2
========================================================
Cleans KPM 2024/2025 weight/height datasets and generates:
  1. KPM_Cleaned_YYYY.csv         — Cleansed data with WHO z-scores & indicators
  2. KPM_QualityReport_YYYY.xlsx  — Data quality report

Source: KPM school-based weight/height data for Tahun Satu (7-year-olds)

Cleaning rules applied (DROP invalid rows):
  Rule 1  : Drop Jantina = 'RAGU' or invalid gender
  Rule 2  : Drop THN_TING not 'TAHUN SATU' (only Tahun 1 students)
  Rule 3  : Drop biologically implausible weight (<12 or >50 kg) / height (<100 or >160 cm)
  Rule 4  : Drop age at assessment outside 6–8 years (72–96 months)
  Rule 5  : Drop implausible BMI > 40
  Rule 6  : Drop records where both weight AND height are null
  Rule 7  : Drop duplicate ID_MURID (keep first occurrence)
  Rule 8  : Drop records where ANY z-score is null (ensures fully clean output)

Special handling for 2024:
  - TARIKH PENGUKURAN is NULL → Set to 31/12/2024 per data owner clarification

WHO Indicators computed (WHO 2007 for school-age children):
  HAZ  — Height-for-Age Z-score  → Bantut (HAZ < -2SD)
  BAZ  — BMI-for-Age Z-score     → Kurus (BAZ < -2SD), Berlebihan Berat Badan (BAZ > +1), Obes (BAZ > +2)
  
  Note: WAZ not used for KPM (school-age children 7 years) per WHO guidelines

KPM 3 Core Indicators:
  1. Prevalen Bantut                    — HAZ < -2SD
  2. Prevalen Kurus                     — BAZ < -2SD  
  3. Prevalen Berlebihan Berat Badan    — BAZ > +1SD (includes Obes BAZ > +2SD)

Standardization (consistent with NCDC/MyVASS):
  - Gender: Male / Female (English)
  - Jantina: Lelaki / Perempuan (Malay)
  - Dates: YYYY-MM-DD format
  - Status labels: Malay
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path for WHO z-score module ─────────────────────────────
_BACKEND = Path(__file__).parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore import compute_zscore
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height bounds (biologically plausible for 6–8 year olds)
BERAT_MIN, BERAT_MAX = 12.0, 50.0      # kg
TINGGI_MIN, TINGGI_MAX = 100.0, 160.0  # cm
BMI_MAX = 40.0

# Age bounds (6–8 years = 72–96 months)
AGE_MIN_MONTHS = 72   # 6 years
AGE_MAX_MONTHS = 96   # 8 years

# WHO BIV thresholds (KPM uses HAZ and BAZ only — no WAZ for school-age)
BIV = {
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]

# Gender mapping
GENDER_MAP = {
    "LELAKI": "Male",
    "PEREMPUAN": "Female",
    "L": "Male",
    "P": "Female",
    "MALE": "Male",
    "FEMALE": "Female",
}

# Valid THN_TING for Tahun Satu
VALID_THN_TING = {"TAHUN SATU", "TAHUN 1", "1", "TAHUN 1A", "TAHUN 1B", "TAHUN 1C"}

# ---------------------------------------------------------------------------
# Label functions (standardized Malay)
# ---------------------------------------------------------------------------

def _haz_label(z):
    """HAZ status label per WHO guidelines."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Bantut Teruk"
    if z < -2:  return "Bantut"
    if z <= 3:  return "Tinggi Normal"
    return "Mungkin Masalah Endokrin"


def _baz_label(z):
    """BAZ status label per WHO guidelines."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Kurus Teruk"
    if z < -2:  return "Kurus"
    if z <= 1:  return "Berat Badan Normal"
    if z <= 2:  return "Berisiko Berlebihan Berat Badan"
    if z <= 3:  return "Berlebihan Berat Badan"
    return "Obes"


# ---------------------------------------------------------------------------
# Cleaning function
# ---------------------------------------------------------------------------

def clean_kpm(raw: pd.DataFrame, year: int) -> tuple[pd.DataFrame, dict]:
    """Clean KPM data and compute WHO z-scores (HAZ, BAZ only)."""
    stats: dict = {"raw_count": len(raw), "year": year}
    df = raw.copy()

    # ── Normalise column names ────────────────────────────────────────────
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    print(f"    Columns detected: {list(df.columns)}")

    # ── Rename columns to standardized names ──────────────────────────────
    rename_map = {
        "TAHUN PERSEKOLAHAN": "Tahun_Persekolahan",
        "NEGERI": "Negeri",
        "DAERAH": "Daerah",
        "NAMA SEKOLAH": "Nama_Sekolah",
        "LOKASI (BANDAR/LUAR BANDAR)": "Lokasi",
        "JENIS SEKOLAH": "Jenis_Sekolah",
        "THN_TING": "Thn_Ting",
        "JANTINA": "Jantina_Raw",
        "ID_MURID": "ID_Murid",
        "TARIKH LAHIR": "Tarikh_Lahir",
        "TARIKH PENGUKURAN BERAT/ TINGGI": "Tarikh_Pengukuran",
        "BERAT (kg)": "Berat_kg",
        "TINGGI (cm)": "Tinggi_cm",
    }
    df = df.rename(columns=rename_map)

    # Add Year column
    df["Year"] = year

    # ── Standardize text fields ───────────────────────────────────────────
    for col in ["Negeri", "Daerah", "Jenis_Sekolah", "Thn_Ting", "Lokasi"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().str.strip()

    for col in ["Nama_Sekolah", "ID_Murid"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # ── Rule 1: Standardize gender, drop RAGU and invalid ─────────────────
    df["Jantina_Raw"] = df["Jantina_Raw"].astype(str).str.upper().str.strip()
    df["Gender"] = df["Jantina_Raw"].map(GENDER_MAP)

    ragu_mask = df["Jantina_Raw"] == "RAGU"
    invalid_gender_mask = df["Gender"].isna()
    stats["dropped_ragu_gender"] = int(ragu_mask.sum())
    stats["dropped_invalid_gender"] = int((invalid_gender_mask & ~ragu_mask).sum())
    df = df[~(invalid_gender_mask | ragu_mask)].copy()

    # Jantina in Malay for compatibility
    df["Jantina"] = df["Gender"].map({"Male": "Lelaki", "Female": "Perempuan"})

    # ── Rule 2: Drop non-Tahun Satu (THN_TING filter) ─────────────────────
    before = len(df)
    tahun_satu_mask = df["Thn_Ting"].isin(VALID_THN_TING)
    stats["dropped_non_tahun_satu"] = int((~tahun_satu_mask).sum())
    non_tahun_satu_values = df.loc[~tahun_satu_mask, "Thn_Ting"].value_counts().to_dict()
    stats["non_tahun_satu_breakdown"] = non_tahun_satu_values
    df = df[tahun_satu_mask].copy()
    print(f"    Rule 2: Dropped {stats['dropped_non_tahun_satu']:,} non-Tahun Satu records")
    if non_tahun_satu_values:
        print(f"      Breakdown: {non_tahun_satu_values}")

    # ── Rule 7: Remove duplicate ID_MURID (keep first) ────────────────────
    before = len(df)
    df = df.drop_duplicates(subset=["ID_Murid"], keep="first")
    stats["dropped_duplicate_id"] = before - len(df)
    print(f"    Rule 7: Dropped {stats['dropped_duplicate_id']:,} duplicate ID_MURID")

    # ── Parse dates ───────────────────────────────────────────────────────
    df["Tarikh_Lahir"] = pd.to_datetime(df["Tarikh_Lahir"], dayfirst=True, errors="coerce")
    
    # Special handling for 2024: Set measurement date to 31/12/2024 if NULL
    if year == 2024:
        df["Tarikh_Pengukuran"] = pd.Timestamp("2024-12-31")
        stats["measurement_date_set"] = "31/12/2024 (all records)"
        print(f"    2024 Special: Set all TARIKH PENGUKURAN to 31/12/2024")
    else:
        df["Tarikh_Pengukuran"] = pd.to_datetime(df["Tarikh_Pengukuran"], dayfirst=True, errors="coerce")

    # Remove time component
    df["Tarikh_Lahir"] = df["Tarikh_Lahir"].dt.normalize()
    if year != 2024:
        df["Tarikh_Pengukuran"] = df["Tarikh_Pengukuran"].dt.normalize()

    stats["null_tarikh_lahir"] = int(df["Tarikh_Lahir"].isna().sum())

    # ── Drop rows with null Tarikh_Lahir (needed for age) ─────────────────
    before = len(df)
    df = df[df["Tarikh_Lahir"].notna()].copy()
    stats["dropped_null_dob"] = before - len(df)

    # ── Compute age at measurement ────────────────────────────────────────
    df["Age_Months"] = ((df["Tarikh_Pengukuran"] - df["Tarikh_Lahir"]).dt.days / 30.4375).round(2)
    df["Age_Years"] = (df["Age_Months"] / 12).round(2)

    # ── Rule 4: Drop age outside 6–8 years ────────────────────────────────
    before = len(df)
    age_invalid = (df["Age_Months"] < AGE_MIN_MONTHS) | (df["Age_Months"] > AGE_MAX_MONTHS)
    stats["dropped_age_invalid"] = int(age_invalid.sum())
    df = df[~age_invalid].copy()
    print(f"    Rule 4: Dropped {stats['dropped_age_invalid']:,} records outside 6–8 years age range")

    # ── Numeric measurements ──────────────────────────────────────────────
    df["Berat_kg"] = pd.to_numeric(df["Berat_kg"], errors="coerce")
    df["Tinggi_cm"] = pd.to_numeric(df["Tinggi_cm"], errors="coerce")

    # ── Rule 3: Drop measurement outliers ─────────────────────────────────
    berat_bad = (df["Berat_kg"] < BERAT_MIN) | (df["Berat_kg"] > BERAT_MAX)
    tinggi_bad = (df["Tinggi_cm"] < TINGGI_MIN) | (df["Tinggi_cm"] > TINGGI_MAX)

    outlier_mask = (berat_bad & df["Berat_kg"].notna()) | (tinggi_bad & df["Tinggi_cm"].notna())
    stats["dropped_measurement_outlier"] = int(outlier_mask.sum())
    stats["berat_out_of_range"] = int((berat_bad & df["Berat_kg"].notna()).sum())
    stats["tinggi_out_of_range"] = int((tinggi_bad & df["Tinggi_cm"].notna()).sum())
    df = df[~outlier_mask].copy()
    print(f"    Rule 3: Dropped {stats['dropped_measurement_outlier']:,} measurement outliers")

    df["Berat_kg"] = df["Berat_kg"].round(1)
    df["Tinggi_cm"] = df["Tinggi_cm"].round(1)

    # ── Rule 6: Drop rows where both measurements are null ────────────────
    before = len(df)
    no_meas_mask = df["Berat_kg"].isna() & df["Tinggi_cm"].isna()
    stats["dropped_no_measurement"] = int(no_meas_mask.sum())
    df = df[~no_meas_mask].copy()

    # ── Compute BMI ───────────────────────────────────────────────────────
    valid_both = df["Berat_kg"].notna() & df["Tinggi_cm"].notna() & (df["Tinggi_cm"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["Berat_kg"] / ((df["Tinggi_cm"] / 100) ** 2)).round(2),
        np.nan,
    )

    # ── Rule 5: Drop implausible BMI > 40 ─────────────────────────────────
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()

    # ── Lokasi mapping ────────────────────────────────────────────────────
    lokasi_map = {"BANDAR": "Urban", "LUAR BANDAR": "Rural"}
    df["Lokasi_EN"] = df["Lokasi"].map(lokasi_map).fillna("Unknown")

    # ── WHO Z-scores (HAZ and BAZ only for school-age) ────────────────────
    if _ZSCORE_AVAILABLE:
        print("    Computing WHO z-scores (HAZ, BAZ for school-age) ...")
        sex_s = df["Gender"].astype(str)
        age_s = df["Age_Months"]
        tinggi_s = df["Tinggi_cm"]
        bmi_s = df["BMI"]

        n = len(df)
        haz_vals, baz_vals = [], []

        for i in range(n):
            age = age_s.iat[i]
            sex = sex_s.iat[i]
            h = tinggi_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age) or pd.isna(sex):
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            haz = compute_zscore(float(h), sex, float(age), "HAZ") if pd.notna(h) else None
            baz = compute_zscore(float(bmi), sex, float(age), "BAZ") if pd.notna(bmi) else None

            # Apply BIV
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 100000 == 0:
                print(f"      Z-scores: {i+1:,}/{n:,} ...")

        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["HAZ_Status"] = df["HAZ"].apply(_haz_label)
        df["BAZ_Status"] = df["BAZ"].apply(_baz_label)

        # KPM 3 Core Indicator flags
        df["Ind_Bantut"] = df["HAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Kurus"] = df["BAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Berlebihan_BB"] = df["BAZ"].apply(lambda z: z is not None and z > 1)
        df["Ind_Obes"] = df["BAZ"].apply(lambda z: z is not None and z > 2)

        # Ind_Normal: Not Bantut, Not Kurus, Not Berlebihan
        df["Ind_Normal"] = ~(df["Ind_Bantut"] | df["Ind_Kurus"] | df["Ind_Berlebihan_BB"])

        # ── Rule 8: Drop rows where ANY z-score is null ───────────────────
        before = len(df)
        null_zscore_mask = df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore_mask.sum())
        df = df[~null_zscore_mask].copy()
        print(f"    Rule 8: Dropped {stats['dropped_null_zscore']:,} rows with null z-scores")

    else:
        for col in ["HAZ", "BAZ", "HAZ_Status", "BAZ_Status",
                    "Ind_Bantut", "Ind_Kurus", "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["ind_bantut"] = int(df["Ind_Bantut"].sum())
        stats["ind_kurus"] = int(df["Ind_Kurus"].sum())
        stats["ind_berlebihan_bb"] = int(df["Ind_Berlebihan_BB"].sum())
        stats["ind_obes"] = int(df["Ind_Obes"].sum())
        stats["ind_normal"] = int(df["Ind_Normal"].sum())

    return df, stats


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr


def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(all_stats: list[dict], df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # Build stats by year
    stats_by_year = {s["year"]: s for s in all_stats}
    years_present = sorted(stats_by_year.keys())

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:F1")
    ws1["A1"].value = "KPM (Kementerian Pendidikan Malaysia) — DATA QUALITY REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:F3")
    ws1["A3"].value = (
        "Data Source: Kanak-Kanak Tahun Satu (7 tahun) | "
        "Indicators: HAZ (Height-for-Age), BAZ (BMI-for-Age) | "
        "WAZ not applicable for school-age children"
    )
    ws1["A3"].font = Font(italic=True, size=9, color="7F7F7F")
    ws1["A3"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Build headers
    year_headers = [str(y) for y in years_present]
    if len(years_present) > 1:
        year_headers.append("Combined")
    _hdr(ws1, 5, ["Metric"] + year_headers + ["Notes"])

    def _total(key):
        return sum(stats_by_year[y].get(key, 0) for y in years_present)

    def _y_vals(key):
        vals = [stats_by_year[y].get(key, "—") for y in years_present]
        if len(years_present) > 1:
            vals.append(_total(key))
        return vals

    summary_rows = [
        ("Raw records (before cleaning)", _y_vals("raw_count"), "Source Excel"),
        ("Dropped — RAGU Gender", _y_vals("dropped_ragu_gender"), "Rule 1"),
        ("Dropped — Invalid Gender", _y_vals("dropped_invalid_gender"), "Rule 1"),
        ("Dropped — Non-Tahun Satu", _y_vals("dropped_non_tahun_satu"), "Rule 2"),
        ("Dropped — Duplicate ID_MURID", _y_vals("dropped_duplicate_id"), "Rule 7"),
        ("Dropped — Null DOB", _y_vals("dropped_null_dob"), "Required for age calc"),
        ("Dropped — Age invalid (6-8 yrs)", _y_vals("dropped_age_invalid"), "Rule 4"),
        ("Dropped — Measurement outliers", _y_vals("dropped_measurement_outlier"), "Rule 3"),
        ("Dropped — BMI > 40", _y_vals("dropped_bmi_outlier"), "Rule 5"),
        ("Dropped — No measurement", _y_vals("dropped_no_measurement"), "Rule 6"),
        ("Dropped — Null z-score(s)", _y_vals("dropped_null_zscore"), "Rule 8"),
        ("Final cleaned records", _y_vals("final_count"), "After all rules"),
    ]
    for i, (label, vals, note) in enumerate(summary_rows, 6):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        _cell(ws1, i, 1, label, bg=bg)
        for j, v in enumerate(vals, 2):
            _cell(ws1, i, j, v, bg=bg, align="right")
        _cell(ws1, i, len(vals) + 2, note, bg=bg)

    row = len(summary_rows) + 7

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["KPM Core Indicator", "Total", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        final_total = _total("final_count")
        ind_rows = [
            ("1. Prevalen Bantut (HAZ < -2SD)", _total("ind_bantut"),
             f"{_total('ind_bantut')/max(final_total,1)*100:.1f}%", "HAZ < -2"),
            ("2. Prevalen Kurus (BAZ < -2SD)", _total("ind_kurus"),
             f"{_total('ind_kurus')/max(final_total,1)*100:.1f}%", "BAZ < -2"),
            ("3. Prevalen Berlebihan BB (BAZ > +1SD)", _total("ind_berlebihan_bb"),
             f"{_total('ind_berlebihan_bb')/max(final_total,1)*100:.1f}%", "BAZ > +1 (includes Obes)"),
            ("   └─ Obes (BAZ > +2SD)", _total("ind_obes"),
             f"{_total('ind_obes')/max(final_total,1)*100:.1f}%", "BAZ > +2"),
            ("Normal (healthy child)", _total("ind_normal"),
             f"{_total('ind_normal')/max(final_total,1)*100:.1f}%", "Not Bantut, Not Kurus, Not BB"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

    _auto_width(ws1)

    # ── Sheet 2: By Year ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Year")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "DISTRIBUTION BY YEAR"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _hdr(ws2, 3, ["Year", "Total", "Bantut", "Kurus", "Berlebihan BB", "Obes", "Normal"])

    for i, year in enumerate(years_present, 4):
        s = stats_by_year[year]
        bg = ALT if i % 2 == 0 else "FFFFFF"
        row_vals = [
            year, s.get("final_count", 0),
            s.get("ind_bantut", 0),
            s.get("ind_kurus", 0),
            s.get("ind_berlebihan_bb", 0),
            s.get("ind_obes", 0),
            s.get("ind_normal", 0),
        ]
        for c, v in enumerate(row_vals, 1):
            _cell(ws2, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws2)

    # ── Sheet 3: By Negeri ────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Negeri")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:G1")
    ws3["A1"].value = "DISTRIBUTION BY NEGERI"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _hdr(ws3, 3, ["Negeri", "Total", "Bantut", "Kurus", "Berlebihan BB", "Obes", "Normal"])

    if "Negeri" in df.columns:
        grp = df.groupby("Negeri")
        for i, (neg, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                neg, len(sub),
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Kurus"].sum()) if "Ind_Kurus" in sub else 0,
                int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                int(sub["Ind_Obes"].sum()) if "Ind_Obes" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws3, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws3)

    # ── Sheet 4: By Jenis Sekolah ─────────────────────────────────────────
    ws4 = wb.create_sheet("By Jenis Sekolah")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:G1")
    ws4["A1"].value = "DISTRIBUTION BY JENIS SEKOLAH"
    ws4["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    _hdr(ws4, 3, ["Jenis Sekolah", "Total", "Bantut", "Kurus", "Berlebihan BB", "Obes", "Normal"])

    if "Jenis_Sekolah" in df.columns:
        grp = df.groupby("Jenis_Sekolah")
        for i, (js, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                js, len(sub),
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Kurus"].sum()) if "Ind_Kurus" in sub else 0,
                int(sub["Ind_Berlebihan_BB"].sum()) if "Ind_Berlebihan_BB" in sub else 0,
                int(sub["Ind_Obes"].sum()) if "Ind_Obes" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws4, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws4)

    return wb


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def save_cleaned_csv(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()

    # Format dates as YYYY-MM-DD
    for col in ["Tarikh_Lahir", "Tarikh_Pengukuran"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Booleans as strings
    for col in ["Ind_Bantut", "Ind_Kurus", "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)

    # Select output columns in standardized order
    out_cols = [
        "Year", "Tahun_Persekolahan", "Negeri", "Daerah", "Nama_Sekolah", 
        "Jenis_Sekolah", "Lokasi", "Lokasi_EN", "Thn_Ting",
        "ID_Murid", "Gender", "Jantina",
        "Tarikh_Lahir", "Tarikh_Pengukuran", "Age_Months", "Age_Years",
        "Berat_kg", "Tinggi_cm", "BMI",
        "HAZ", "BAZ",
        "HAZ_Status", "BAZ_Status",
        "Ind_Bantut", "Ind_Kurus", "Ind_Berlebihan_BB", "Ind_Obes", "Ind_Normal",
    ]
    out_cols = [c for c in out_cols if c in df_out.columns]
    df_out = df_out[out_cols]

    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# File picker helpers
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=initial_dir,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


def _pick_sheet_dialog(filename: str, sheets: list[str]) -> str | None:
    """Show a small Tk window with a listbox of sheet names."""
    result: list[str | None] = [None]

    root = tk.Tk()
    root.title(f"Select sheet — {filename}")
    root.attributes("-topmost", True)
    root.resizable(False, False)

    tk.Label(root, text=f"Select the data sheet in:\n{filename}",
             wraplength=400, justify="left", padx=10, pady=6).pack()

    lb = tk.Listbox(root, selectmode=tk.SINGLE, width=60, height=min(len(sheets), 12))
    for s in sheets:
        lb.insert(tk.END, s)
    lb.select_set(0)
    lb.pack(padx=10, pady=4)

    def _ok():
        sel = lb.curselection()
        if sel:
            result[0] = sheets[sel[0]]
        root.destroy()

    def _cancel():
        root.destroy()

    btn_frame = tk.Frame(root)
    tk.Button(btn_frame, text="OK", width=10, command=_ok).pack(side=tk.LEFT, padx=4)
    tk.Button(btn_frame, text="Cancel", width=10, command=_cancel).pack(side=tk.LEFT, padx=4)
    btn_frame.pack(pady=6)

    root.mainloop()
    return result[0]


REQUIRED_COLS = {"NEGERI", "JANTINA", "ID_MURID", "BERAT (KG)", "TINGGI (CM)"}


def _get_sheet(path: Path, year: int) -> str | None:
    """Auto-detect the data sheet."""
    xl = pd.ExcelFile(path)
    sheets = xl.sheet_names

    # Try to find sheet with year in name
    candidates = [s for s in sheets if str(year) in s or s.strip().isdigit()]
    if len(candidates) == 1:
        return candidates[0]

    # Validate each sheet by columns
    for s in sheets:
        try:
            peek = xl.parse(s, nrows=2)
            cols = set(peek.columns.str.replace(r"\s+", " ", regex=True).str.strip().str.upper())
            if REQUIRED_COLS.issubset(cols):
                return s
        except Exception:
            continue

    # Ask user to pick
    return _pick_sheet_dialog(path.name, sheets)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("  KPM (Kementerian Pendidikan Malaysia) Data Cleaning Tool v2")
    print("  Tahun Satu (7-year-old) Weight/Height Data")
    print("=" * 70)

    # Step 1: Choose years
    print("\nStep 1: Which year(s) do you want to process?")
    print("  [1] 2024 only")
    print("  [2] 2025 only")
    print("  [3] Both 2024 and 2025")
    choice = input("  Enter choice (1/2/3): ").strip()
    if choice not in ("1", "2", "3"):
        print("  ✗ Invalid choice. Exiting.")
        return
    years_to_run = {"1": [2024], "2": [2025], "3": [2024, 2025]}[choice]
    print(f"  ✓ Processing: {years_to_run}")

    # Step 2: Select input files
    file_map: dict[int, Path] = {}
    for step_num, year in enumerate(years_to_run, 2):
        print(f"\nStep {step_num}: Select the {year} data file (.xlsx)")
        initial = str(file_map[years_to_run[0]].parent) if file_map else str(BASE_DIR)
        chosen = _pick_file(f"Select {year} data file", initial)
        if not chosen:
            print(f"  ✗ No file selected for {year}. Exiting.")
            return
        file_map[year] = chosen
        print(f"  ✓ {year} file: {chosen}")

    # Step 3: Select output folder
    step_out = len(years_to_run) + 2
    print(f"\nStep {step_out}: Select the output folder")
    first_file = file_map[years_to_run[0]]
    out_folder = _pick_folder("Select output folder", str(first_file.parent))
    if not out_folder:
        out_folder = first_file.parent / "KPM_Cleaned"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output: {out_folder}")

    # Process each year
    all_stats: list[dict] = []
    cleaned_frames: list[pd.DataFrame] = []

    for year in years_to_run:
        path = file_map[year]
        sheet = _get_sheet(path, year)
        if sheet is None:
            print(f"  ✗ No sheet selected for {year}. Skipping.")
            continue

        print(f"\n{'='*70}")
        print(f"[{year}] Loading {path.name} (sheet: {sheet}) ...")
        raw = pd.read_excel(path, sheet_name=sheet)
        print(f"  Raw rows: {len(raw):,}")

        # Clean
        print(f"\n  Applying cleaning rules ...")
        df, stats = clean_kpm(raw, year)
        all_stats.append(stats)
        cleaned_frames.append(df)

        print(f"\n  ───────────────────────────────────────────────────────────")
        print(f"  Total dropped : {stats['total_dropped']:,}  ({stats['total_dropped']/stats['raw_count']*100:.1f}%)")
        print(f"  Final records : {stats['final_count']:,}")

        if _ZSCORE_AVAILABLE:
            print(f"\n  KPM Core Indicators:")
            print(f"    1. Bantut (HAZ < -2)         : {stats.get('ind_bantut', 0):,}")
            print(f"    2. Kurus (BAZ < -2)          : {stats.get('ind_kurus', 0):,}")
            print(f"    3. Berlebihan BB (BAZ > +1)  : {stats.get('ind_berlebihan_bb', 0):,}")
            print(f"       └─ Obes (BAZ > +2)        : {stats.get('ind_obes', 0):,}")
            print(f"    Normal                       : {stats.get('ind_normal', 0):,}")

    if not cleaned_frames:
        print("  ✗ No data to save.")
        return

    df_all = pd.concat(cleaned_frames, ignore_index=True)
    print(f"\n{'='*70}")
    print(f"Total rows in combined output: {len(df_all):,}")

    # Save quality report
    print("\nBuilding quality report ...")
    year_tag = "_".join(str(y) for y in years_to_run)
    out_report = out_folder / f"KPM_QualityReport_{year_tag}.xlsx"
    qr = build_quality_report(all_stats, df_all)
    qr.save(out_report)
    print(f"  Saved: {out_report}")

    # Save cleaned CSV
    print("\nSaving cleaned CSV ...")
    out_csv = out_folder / f"KPM_Cleaned_{year_tag}.csv"
    save_cleaned_csv(df_all, out_csv)

    print()
    print("=" * 70)
    print("  Done! Files saved to:")
    print(f"    {out_report}")
    print(f"    {out_csv}")
    print("=" * 70)

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(
        "Done",
        f"KPM cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n"
        f"• KPM_Cleaned_{year_tag}.csv\n"
        f"• KPM_QualityReport_{year_tag}.xlsx",
    )
    root.destroy()


if __name__ == "__main__":
    main()
