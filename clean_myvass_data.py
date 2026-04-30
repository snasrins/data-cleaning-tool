"""
MyVASS Data Cleaning Script
========================================
Cleans PenyediaanGISStatusPemakananKanakkanak.csv and generates:
  1. MyVASS_Cleaned.csv          — Cleansed data with WHO z-scores & indicators
  2. MyVASS_QualityReport.xlsx   — Data quality report

Cleaning rules applied (drop invalid rows):
  Rule 1  : Drop Jantina = 'Others'
  Rule 2  : Drop biologically implausible weight (≤0.5 or >35 kg) / height (≤30 or >130 cm)
  Rule 3  : Drop age at assessment > 5 years (≥60 months)
  Rule 4  : Drop records where Tarikh Antropometri < Tarikh Lahir (assessment before birth)
  Rule 5  : Drop implausible BMI > 40 (weight/height combination impossible)
  Rule 6  : Drop records where both weight AND height are null (no measurement)
  Rule 7  : Drop records where ANY z-score is null (ensures fully clean data)

WHO Indicators computed:
  WAZ  — Weight-for-Age Z-score  → Kurang Berat Badan (WAZ < -2SD)
  HAZ  — Height-for-Age Z-score  → Bantut (HAZ < -2SD)
  BAZ  — BMI-for-Age Z-score     → Susut (BAZ < -2SD)
  Ind_Normal — TRUE when child is healthy (all above indicators FALSE)

Age groups:
  Bawah 2 Tahun : 0–23 months
  Bawah 5 Tahun : 0–59 months (includes Bawah 2 Tahun)

Geographic:
  Bahagian derived for Sabah (6 Bahagian) and Sarawak (12 Bahagian)
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path so we can reuse the WHO z-score module ─────────────
_BACKEND = Path(__file__).parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore_daily import compute_zscore, classify_waz, classify_haz, classify_baz
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height hard outlier bounds (biologically implausible for 0–5 yr)
BERAT_MIN,  BERAT_MAX  = 0.5,  35.0   # kg
PANJANG_MIN, PANJANG_MAX = 30.0, 130.0  # cm
BMI_MAX = 40.0  # Rule 5: combined weight/height implausibility check

# Age bounds
AGE_MAX_MONTHS = 60   # under 5 years

# WHO BIV (Biologically Implausible Values) thresholds for z-scores
BIV = {
    "WAZ": (-6.0, +5.0),
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}

# Output column names for kept fields
KEEP_COLS_MAP = {
    "Jantina":                  "Jantina",
    "Tarikh Lahir":             "Tarikh_Lahir",
    "Panjang Lahir (kg)":       "Panjang_cm",
    "Berat Lahir (kg)":         "Berat_kg",
    "Negeri":                   "Negeri",
    "Nama Klinik":              "Nama_Klinik",
    "Daerah":                   "Daerah",
    "VACCINE_NAME":             "Vaccine_Name",
    "Tarikh Antropometri":      "Tarikh_Antropometri",
}

# ---------------------------------------------------------------------------
# Bahagian lookup (Sabah: 6, Sarawak: 12)
# ---------------------------------------------------------------------------
BAHAGIAN_MAP: dict[str, str] = {
    # SABAH — 6 Bahagian
    "Kudat":         "Kudat",
    "Kota Marudu":   "Kudat",
    "Pitas":         "Kudat",
    "Kota Kinabalu": "Pantai Barat",
    "Penampang":     "Pantai Barat",
    "Putatan":       "Pantai Barat",
    "Papar":         "Pantai Barat",
    "Tuaran":        "Pantai Barat",
    "Kota Belud":    "Pantai Barat",
    "Ranau":         "Pantai Barat",
    "Beaufort":      "Pantai Barat",
    "Kuala Penyu":   "Pantai Barat",
    "Sipitang":      "Pantai Barat",
    "Keningau":      "Pedalaman",
    "Tambunan":      "Pedalaman",
    "Tenom":         "Pedalaman",
    "Nabawan":       "Pedalaman",
    "Sandakan":      "Sandakan",
    "Kinabatangan":  "Sandakan",
    "Beluran":       "Sandakan",
    "Tongod":        "Sandakan",
    "Telupid":       "Sandakan",
    "Lahad Datu":    "Lahad Datu",
    "Semporna":      "Lahad Datu",
    "Kunak":         "Lahad Datu",
    "Tawau":         "Tawau",
    # SARAWAK — 12 Bahagian
    "Kuching":       "Kuching",
    "Bau":           "Kuching",
    "Lundu":         "Kuching",
    "Asajaya":       "Kota Samarahan",
    "Sri Aman":      "Sri Aman",
    "Lubok Antu":    "Sri Aman",
    "Betong":        "Betong",
    "Saratok":       "Betong",
    "Sarikei":       "Sarikei",
    "Meradong":      "Sarikei",
    "Julau":         "Sarikei",
    "Sibu":          "Sibu",
    "Kanowit":       "Sibu",
    "Mukah":         "Mukah",
    "Bintulu":       "Bintulu",
    "Miri":          "Miri",
    "Marudi":        "Miri",
    "Subis":         "Miri",
    "Beluru":        "Miri",
    "Limbang":       "Limbang",
    "Lawas":         "Limbang",
    "Serian":        "Serian",
}

SABAH_SARAWAK = {"Sabah", "Sarawak"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(series: pd.Series) -> pd.Series:
    """Try multiple date formats; return datetime series."""
    return pd.to_datetime(series, dayfirst=True, errors="coerce")


def _biv_flag(z: float | None, indicator: str) -> bool:
    """Return True if z-score is within the BIV-accepted range."""
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return False
    lo, hi = BIV[indicator]
    return lo <= z <= hi


def _waz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Kurang Berat Badan Teruk"
    if z < -2:  return "Kurang Berat Badan"
    if z <= 2:  return "Berat Badan Normal"
    return "Mungkin Masalah Pertumbuhan"


def _haz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Bantut Teruk"
    if z < -2:  return "Bantut"
    if z <= 3:  return "Panjang/Tinggi Normal"
    return "Mungkin Masalah Endokrin"


def _baz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Susut Teruk"
    if z < -2:  return "Susut"
    if z <= 1:  return "Berat Badan Normal"
    if z <= 2:  return "Berisiko Berlebihan Berat Badan"
    if z <= 3:  return "Berlebihan Berat Badan"
    return "Obes"


# ---------------------------------------------------------------------------
# Main cleaning function
# ---------------------------------------------------------------------------

def clean_myvass(raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    stats: dict = {"raw_count": len(raw)}
    df = raw.copy()

    # ── Normalise column names ────────────────────────────────────────────
    df.columns = df.columns.str.strip()

    # ── Keep only required columns ────────────────────────────────────────
    missing_src = [c for c in KEEP_COLS_MAP if c not in df.columns]
    if missing_src:
        raise KeyError(f"Missing expected columns: {missing_src}")
    df = df[list(KEEP_COLS_MAP.keys())].rename(columns=KEEP_COLS_MAP)

    # ── Rule 1: Drop Jantina = Others ─────────────────────────────────────
    before = len(df)
    df = df[df["Jantina"] != "Others"].copy()
    stats["dropped_others_jantina"] = before - len(df)

    # ── Parse dates ───────────────────────────────────────────────────────
    df["Tarikh_Lahir"]       = _parse_date(df["Tarikh_Lahir"])
    df["Tarikh_Antropometri"]= _parse_date(df["Tarikh_Antropometri"])
    stats["null_tarikh_lahir"]        = int(df["Tarikh_Lahir"].isna().sum())
    stats["null_tarikh_antropometri"] = int(df["Tarikh_Antropometri"].isna().sum())

    # ── Rule 4 (assessment before DOB) ────────────────────────────────────
    before = len(df)
    bad_date_mask = (
        df["Tarikh_Lahir"].notna() &
        df["Tarikh_Antropometri"].notna() &
        (df["Tarikh_Antropometri"] < df["Tarikh_Lahir"])
    )
    stats["dropped_assessment_before_dob"] = int(bad_date_mask.sum())
    df = df[~bad_date_mask].copy()

    # ── Compute age at assessment (in DAYS for WHO z-score calculation) ─
    has_both = df["Tarikh_Lahir"].notna() & df["Tarikh_Antropometri"].notna()
    df["Age_Days"] = np.where(
        has_both,
        (df["Tarikh_Antropometri"] - df["Tarikh_Lahir"]).dt.days,
        np.nan,
    )
    df["Age_Months"] = np.where(
        has_both,
        df["Age_Days"] / 30.4375,
        np.nan,
    )
    df["Age_Months"] = df["Age_Months"].round(2)

    # ── Rule 3: Drop age > 5 years (≥ 60 months) ─────────────────────────
    before = len(df)
    over5_mask = df["Age_Months"].notna() & (df["Age_Months"] >= AGE_MAX_MONTHS)
    stats["dropped_age_over5"] = int(over5_mask.sum())
    df = df[~over5_mask].copy()

    # ── Rule 2: Drop weight/height outliers ───────────────────────────────
    df["Berat_kg"]  = pd.to_numeric(df["Berat_kg"],  errors="coerce")
    df["Panjang_cm"]= pd.to_numeric(df["Panjang_cm"],errors="coerce")

    berat_bad  = (df["Berat_kg"] <= BERAT_MIN)   | (df["Berat_kg"] > BERAT_MAX)
    panjang_bad= (df["Panjang_cm"] <= PANJANG_MIN)| (df["Panjang_cm"] > PANJANG_MAX)

    # Count before dropping
    stats["dropped_berat_outlier"]  = int((berat_bad   & df["Berat_kg"].notna()).sum())
    stats["dropped_panjang_outlier"]= int((panjang_bad & df["Panjang_cm"].notna()).sum())

    before = len(df)
    df = df[~(berat_bad | panjang_bad)].copy()
    stats["dropped_measurement_outlier"] = before - len(df)

    # Round cleaned values
    df["Berat_kg"]  = df["Berat_kg"].round(2)
    df["Panjang_cm"]= df["Panjang_cm"].round(1)

    # ── Rule 5: Drop missing measurements ────────────────────────────────
    # Records where both berat AND panjang are null have no measurement value
    before = len(df)
    no_meas_mask = df["Berat_kg"].isna() & df["Panjang_cm"].isna()
    stats["dropped_no_measurement"] = int(no_meas_mask.sum())
    df = df[~no_meas_mask].copy()

    # ── Compute BMI ───────────────────────────────────────────────────────
    valid_both = df["Berat_kg"].notna() & df["Panjang_cm"].notna() & (df["Panjang_cm"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["Berat_kg"] / ((df["Panjang_cm"] / 100) ** 2)).round(2),
        np.nan,
    )

    # ── Rule 6: Drop implausible BMI > 40 ────────────────────────────────
    # Individual weight/height may be within bounds but their combination is
    # impossible (e.g. 32 kg at 50 cm → BMI 128 — birth-length paired with
    # a much-later weight, or severe data entry error).
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()

    # ── Age group ─────────────────────────────────────────────────────────
    def _age_group(m):
        if pd.isna(m):    return None
        if m < 24:        return "Bawah 2 Tahun"
        if m < 60:        return "Bawah 5 Tahun"
        return None
    df["Kumpulan_Umur"] = df["Age_Months"].apply(_age_group)

    # ── Bahagian (Sabah & Sarawak) ────────────────────────────────────────
    df["Bahagian"] = df.apply(
        lambda r: BAHAGIAN_MAP.get(r["Daerah"], "Lain-lain")
        if r["Negeri"] in SABAH_SARAWAK else None,
        axis=1,
    )

    # ── WHO Z-scores (using DAILY LMS tables) ──────────────────────────────────────
    if _ZSCORE_AVAILABLE:
        sex_s  = df["Jantina"].astype(str)
        age_s  = df["Age_Days"]  # Use days instead of months
        berat_s = df["Berat_kg"]
        panjang_s = df["Panjang_cm"]
        bmi_s  = df["BMI"]

        n = len(df)
        waz_vals, haz_vals, baz_vals = [], [], []

        for i in range(n):
            age_d = age_s.iat[i]
            sex = sex_s.iat[i]
            w   = berat_s.iat[i]
            h   = panjang_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age_d) or pd.isna(sex):
                waz_vals.append(None)
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            # New signature: compute_zscore(indicator, value, age_days, sex)
            waz = compute_zscore("WAZ", float(w),   int(age_d), sex) if pd.notna(w)   else None
            haz = compute_zscore("HAZ", float(h),   int(age_d), sex) if pd.notna(h)   else None
            baz = compute_zscore("BAZ", float(bmi), int(age_d), sex) if pd.notna(bmi) else None

            # Apply BIV — set to None if outside plausible range
            if waz is not None and not (_BIV_LO_WAZ <= waz <= _BIV_HI_WAZ): waz = None
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            waz_vals.append(round(waz, 3) if waz is not None else None)
            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 50000 == 0:
                print(f"    Z-scores: {i+1:,}/{n:,} ...")

        df["WAZ"] = waz_vals
        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["WAZ_Status"] = df["WAZ"].apply(classify_waz)
        df["HAZ_Status"] = df["HAZ"].apply(classify_haz)
        df["BAZ_Status"] = df["BAZ"].apply(classify_baz)

        # Indicator flags (True/False)
        df["Ind_Kurang_Berat_Badan"] = df["WAZ"].apply(
            lambda z: z is not None and z < -2)
        df["Ind_Bantut"]             = df["HAZ"].apply(
            lambda z: z is not None and z < -2)
        df["Ind_Susut"]              = df["BAZ"].apply(
            lambda z: z is not None and z < -2)

        # ── Rule 7: Drop rows where ANY z-score is null ──────────────────────
        # This ensures the output is fully clean with no null z-scores
        before = len(df)
        null_zscore_mask = df["WAZ"].isna() | df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore_mask.sum())
        df = df[~null_zscore_mask].copy()
        print(f"    Rule 7: Dropped {stats['dropped_null_zscore']:,} rows with null z-scores")

        # ── Ind_Normal: TRUE when child is healthy (all indicators FALSE)
        df["Ind_Normal"] = ~(df["Ind_Kurang_Berat_Badan"] | df["Ind_Bantut"] | df["Ind_Susut"])

    else:
        for col in ["WAZ","HAZ","BAZ","WAZ_Status","HAZ_Status","BAZ_Status",
                    "Ind_Kurang_Berat_Badan","Ind_Bantut","Ind_Susut","Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats.setdefault("dropped_no_measurement", 0)
    stats.setdefault("dropped_bmi_outlier", 0)
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["waz_valid"]          = int(df["WAZ"].notna().sum())
        stats["haz_valid"]          = int(df["HAZ"].notna().sum())
        stats["baz_valid"]          = int(df["BAZ"].notna().sum())
        stats["ind_kurang_berat"]   = int(df["Ind_Kurang_Berat_Badan"].sum())
        stats["ind_bantut"]         = int(df["Ind_Bantut"].sum())
        stats["ind_susut"]          = int(df["Ind_Susut"].sum())
        stats["ind_normal"]         = int(df["Ind_Normal"].sum())
        stats["bawah_2_tahun"]      = int((df["Kumpulan_Umur"] == "Bawah 2 Tahun").sum())
        stats["bawah_5_tahun"]      = int(df["Kumpulan_Umur"].notna().sum())

    return df, stats


# BIV bounds unpacked for speed in the loop
_BIV_LO_WAZ, _BIV_HI_WAZ = BIV["WAZ"]
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr

def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(stats: dict, df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "MyVASS — DATA QUALITY REPORT"
    ws1["A1"].font  = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font  = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:D3")
    ws1["A3"].value = (
        "NOTE: Negative WAZ/HAZ/BAZ values are EXPECTED — z-scores are centred on 0 (WHO median). "
        "Negative = below median. WAZ/HAZ/BAZ < -2 = at-risk. "
        "Cleaned data contains NO null z-scores (rows with null z-scores are dropped via Rule 7). "
        "Bahagian is NULL for non-Sabah/Sarawak states — this is correct."
    )
    ws1["A3"].font      = Font(italic=True, size=9, color="7F7F7F")
    ws1["A3"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws1.row_dimensions[3].height = 36

    _hdr(ws1, 5, ["Metric", "Count", "% of Raw", "Notes"])
    total = stats["raw_count"]

    summary_rows = [
        ("Raw records (before cleaning)",   stats["raw_count"],  "100%",   "Source CSV"),
        ("Dropped — Jantina 'Others'",      stats["dropped_others_jantina"],
         f"{stats['dropped_others_jantina']/total*100:.2f}%",   "Rule 1"),
        ("Dropped — Assessment before DOB", stats["dropped_assessment_before_dob"],
         f"{stats['dropped_assessment_before_dob']/total*100:.2f}%", "Rule 4"),
        ("Dropped — Age > 5 years",         stats["dropped_age_over5"],
         f"{stats['dropped_age_over5']/total*100:.2f}%",        "Rule 3"),
        ("Dropped — Measurement outliers",  stats["dropped_measurement_outlier"],
         f"{stats['dropped_measurement_outlier']/total*100:.2f}%", "Rule 2"),
        ("Dropped — Implausible BMI > 40",  stats.get("dropped_bmi_outlier", 0),
         f"{stats.get('dropped_bmi_outlier',0)/total*100:.2f}%", "Rule 5"),
        ("Dropped — No measurement (both null)", stats.get("dropped_no_measurement", 0),
         f"{stats.get('dropped_no_measurement',0)/total*100:.2f}%", "Rule 6"),
        ("Dropped — Null z-score(s)", stats.get("dropped_null_zscore", 0),
         f"{stats.get('dropped_null_zscore',0)/total*100:.2f}%", "Rule 7"),
        ("Final cleaned records",           stats["final_count"],
         f"{stats['final_count']/total*100:.2f}%",              "After all rules"),
    ]
    for i, r in enumerate(summary_rows, 6):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(r, 1):
            _cell(ws1, i, c, v, bg=bg)

    row = len(summary_rows) + 7

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["WHO Indicator", "Positive Count", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        ind_rows = [
            ("Kurang Berat Badan (WAZ < -2SD)",
             stats.get("ind_kurang_berat", 0),
             f"{stats.get('ind_kurang_berat',0)/max(stats['final_count'],1)*100:.1f}%",
             "WAZ < -2"),
            ("Bantut (HAZ < -2SD)",
             stats.get("ind_bantut", 0),
             f"{stats.get('ind_bantut',0)/max(stats['final_count'],1)*100:.1f}%",
             "HAZ < -2"),
            ("Susut (BAZ < -2SD)",
             stats.get("ind_susut", 0),
             f"{stats.get('ind_susut',0)/max(stats['final_count'],1)*100:.1f}%",
             "BAZ < -2"),
            ("Normal (healthy child)",
             stats.get("ind_normal", 0),
             f"{stats.get('ind_normal',0)/max(stats['final_count'],1)*100:.1f}%",
             "All indicators FALSE"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)
        row += len(ind_rows) + 2

        _hdr(ws1, row, ["Age Group", "Count", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        _cell(ws1, row, 1, "Bawah 2 Tahun",   bg=ALT)
        _cell(ws1, row, 2, stats.get("bawah_2_tahun", 0), bg=ALT)
        _cell(ws1, row, 3, f"{stats.get('bawah_2_tahun',0)/max(stats['final_count'],1)*100:.1f}%", bg=ALT)
        _cell(ws1, row, 4, "0 – 23 months",   bg=ALT)
        row += 1
        _cell(ws1, row, 1, "Bawah 5 Tahun")
        _cell(ws1, row, 2, stats.get("bawah_5_tahun", 0))
        _cell(ws1, row, 3, f"{stats.get('bawah_5_tahun',0)/max(stats['final_count'],1)*100:.1f}%")
        _cell(ws1, row, 4, "0 – 59 months")

    _auto_width(ws1)

    # ── Sheet 2: Cleaning Rules ───────────────────────────────────────────
    ws2 = wb.create_sheet("Cleaning Rules")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "CLEANING RULES APPLIED"
    ws2["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _hdr(ws2, 3, ["Rule", "Column(s)", "Condition", "Action", "Rows Affected", "% of Raw"])
    rules = [
        ("Rule 1", "Jantina",
         "Jantina = 'Others'",
         "Drop row",
         stats["dropped_others_jantina"],
         f"{stats['dropped_others_jantina']/total*100:.2f}%"),
        ("Rule 2", "Berat Lahir (kg) / Panjang Lahir (kg)",
         f"Berat ≤ {BERAT_MIN} or > {BERAT_MAX} kg  |  Panjang ≤ {PANJANG_MIN} or > {PANJANG_MAX} cm",
         "Drop row",
         stats["dropped_measurement_outlier"],
         f"{stats['dropped_measurement_outlier']/total*100:.2f}%"),
        ("Rule 3", "Tarikh Lahir / Tarikh Antropometri",
         "Age at assessment ≥ 60 months (> 5 years)",
         "Drop row",
         stats["dropped_age_over5"],
         f"{stats['dropped_age_over5']/total*100:.2f}%"),
        ("Rule 4", "Tarikh Lahir / Tarikh Antropometri",
         "Tarikh Antropometri < Tarikh Lahir",
         "Drop row",
         stats["dropped_assessment_before_dob"],
         f"{stats['dropped_assessment_before_dob']/total*100:.2f}%"),        ("Rule 5", "Berat Lahir (kg) + Panjang Lahir (kg)",
         f"BMI > {BMI_MAX} (weight/height combination implausible e.g. birth-length with later weight)",
         "Drop row",
         stats.get("dropped_bmi_outlier", 0),
         f"{stats.get('dropped_bmi_outlier',0)/total*100:.2f}%"),
        ("Rule 6", "Berat Lahir (kg) + Panjang Lahir (kg)",
         "Both weight AND height are NULL (no measurement recorded)",
         "Drop row",
         stats.get("dropped_no_measurement", 0),
         f"{stats.get('dropped_no_measurement',0)/total*100:.2f}%"),    ]
    for i, r in enumerate(rules, 4):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        for c, v in enumerate(r, 1):
            _cell(ws2, i, c, v, bg=bg, align="center" if c >= 5 else "left")

    _auto_width(ws2)

    # ── Sheet 3: BMI / Z-score Summary ───────────────────────────────────
    if _ZSCORE_AVAILABLE and "WAZ" in df.columns:
        ws3 = wb.create_sheet("Z-Score Summary")
        ws3.sheet_view.showGridLines = False
        ws3.merge_cells("A1:E1")
        ws3["A1"].value = "WHO Z-SCORE DISTRIBUTION"
        ws3["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
        ws3["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 28

        def _dist_rows(col, label_col):
            vc = df[label_col].value_counts()
            total_valid = int(df[col].notna().sum())
            rows = []
            for status, cnt in vc.items():
                rows.append((str(status), int(cnt),
                             f"{int(cnt)/max(total_valid,1)*100:.1f}%"))
            rows.append(("(No valid z-score)", int(df[col].isna().sum()), ""))
            return rows

        r = 3
        for col, label_col, title in [
            ("WAZ", "WAZ_Status", "WAZ — Weight-for-Age"),
            ("HAZ", "HAZ_Status", "HAZ — Height-for-Age"),
            ("BAZ", "BAZ_Status", "BAZ — BMI-for-Age"),
        ]:
            if col not in df.columns:
                continue
            ws3.merge_cells(f"A{r}:E{r}")
            ws3.cell(row=r, column=1).value = title
            ws3.cell(row=r, column=1).font  = Font(bold=True, size=12)
            ws3.cell(row=r, column=1).fill  = PatternFill("solid", fgColor="2E75B6")
            ws3.cell(row=r, column=1).font  = Font(bold=True, color="FFFFFF", size=12)
            r += 1
            _hdr(ws3, r, ["Status", "Count", "% of Valid", "", ""], bg="4472C4")
            r += 1
            for j, row_data in enumerate(_dist_rows(col, label_col)):
                bg = ALT if j % 2 == 0 else "FFFFFF"
                for c, v in enumerate(row_data, 1):
                    _cell(ws3, r, c, v, bg=bg)
                r += 1
            r += 1
        _auto_width(ws3)

    # ── Sheet 4: By Negeri ────────────────────────────────────────────────
    ws4 = wb.create_sheet("By Negeri")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:F1")
    ws4["A1"].value = "DISTRIBUTION BY NEGERI"
    ws4["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    neg_cols = ["Negeri", "Total"]
    if _ZSCORE_AVAILABLE and "Ind_Kurang_Berat_Badan" in df.columns:
        neg_cols += ["Kurang Berat Badan", "Bantut", "Susut"]
    _hdr(ws4, 3, neg_cols)

    grp = df.groupby("Negeri")
    for i, (neg, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        row_vals = [neg, len(sub)]
        if _ZSCORE_AVAILABLE and "Ind_Kurang_Berat_Badan" in df.columns:
            row_vals += [
                int(sub["Ind_Kurang_Berat_Badan"].sum()),
                int(sub["Ind_Bantut"].sum()),
                int(sub["Ind_Susut"].sum()),
            ]
        for c, v in enumerate(row_vals, 1):
            _cell(ws4, i, c, v, bg=bg, align="right" if c > 1 else "left")
    _auto_width(ws4)

    # ── Sheet 5: Flagged / Dropped Log ────────────────────────────────────
    ws5 = wb.create_sheet("Drop Log")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:C1")
    ws5["A1"].value = "RECORDS DROPPED — Summary by Rule"
    ws5["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws5["A1"].fill  = PatternFill("solid", fgColor="C00000")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    _hdr(ws5, 3, ["Rule", "Records Dropped", "Cumulative Remaining"], bg="C00000")
    remaining = stats["raw_count"]
    drop_seq = [
        ("Rule 1 — Jantina 'Others'",       stats["dropped_others_jantina"]),
        ("Rule 4 — Assessment before DOB",   stats["dropped_assessment_before_dob"]),
        ("Rule 3 — Age > 5 years",           stats["dropped_age_over5"]),
        ("Rule 2 — Measurement outliers",    stats["dropped_measurement_outlier"]),        ("Rule 5 \u2014 Implausible BMI > 40",    stats.get("dropped_bmi_outlier", 0)),
        ("Rule 6 \u2014 No measurement (both null)", stats.get("dropped_no_measurement", 0)),    ]
    for i, (label, dropped) in enumerate(drop_seq, 4):
        remaining -= dropped
        bg = "FCE4D6" if i % 2 == 0 else "FFFFFF"
        _cell(ws5, i, 1, label, bg=bg)
        _cell(ws5, i, 2, dropped, bg=bg, align="right")
        _cell(ws5, i, 3, remaining, bg=bg, align="right")

    _auto_width(ws5)

    return wb


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def save_cleaned_csv(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()
    # Format dates as YYYY-MM-DD
    for col in ["Tarikh_Lahir", "Tarikh_Antropometri"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")
    # Booleans as readable
    for col in ["Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)
    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# File picker helpers
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=initial_dir,
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  MyVASS Data Cleaning Tool")
    print("=" * 60)

    # Step 1: Pick input CSV
    print("\nStep 1: Select the MyVASS CSV file")
    csv_path = _pick_file("Select MyVASS CSV", str(BASE_DIR))
    if not csv_path:
        print("  ✗ No file selected. Exiting.")
        return
    print(f"  ✓ File : {csv_path}")

    # Step 2: Pick output folder
    print("\nStep 2: Select the output folder")
    out_folder = _pick_folder("Select output folder", str(csv_path.parent))
    if not out_folder:
        out_folder = csv_path.parent / "Report"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output: {out_folder}")

    # Load
    print(f"\nLoading {csv_path.name} ...")
    try:
        raw = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    except UnicodeDecodeError:
        raw = pd.read_csv(csv_path, encoding="latin-1", low_memory=False)
    print(f"  Raw rows: {len(raw):,}")

    # Clean
    print("\nApplying cleaning rules ...")
    df, stats = clean_myvass(raw)
    print(f"  Rule 1 (Jantina Others)       dropped: {stats['dropped_others_jantina']:,}")
    print(f"  Rule 4 (Assessment < DOB)     dropped: {stats['dropped_assessment_before_dob']:,}")
    print(f"  Rule 3 (Age > 5 years)        dropped: {stats['dropped_age_over5']:,}")
    print(f"  Rule 2 (Measurement outliers) dropped: {stats['dropped_measurement_outlier']:,}")
    print(f"  Rule 5 (BMI > 40)             dropped: {stats.get('dropped_bmi_outlier', 0):,}")
    print(f"  Rule 6 (No measurement)       dropped: {stats.get('dropped_no_measurement', 0):,}")
    print(f"  Rule 7 (Null z-scores)        dropped: {stats.get('dropped_null_zscore', 0):,}")
    print(f"  ─────────────────────────────────────")
    print(f"  Total dropped : {stats['total_dropped']:,}  ({stats['total_dropped']/stats['raw_count']*100:.1f}%)")
    print(f"  Final records : {stats['final_count']:,}")

    if _ZSCORE_AVAILABLE:
        print(f"\n  WHO Indicators (in cleaned data):")
        print(f"    Kurang Berat Badan (WAZ < -2) : {stats.get('ind_kurang_berat', 0):,}")
        print(f"    Bantut             (HAZ < -2) : {stats.get('ind_bantut', 0):,}")
        print(f"    Susut              (BAZ < -2) : {stats.get('ind_susut', 0):,}")
        print(f"    Normal      (all FALSE)       : {stats.get('ind_normal', 0):,}")
        print(f"\n  Age groups:")
        print(f"    Bawah 2 Tahun (0–23 months)  : {stats.get('bawah_2_tahun', 0):,}")
        print(f"    Bawah 5 Tahun (0–59 months)  : {stats.get('bawah_5_tahun', 0):,}")

    # Save quality report
    print("\nBuilding quality report ...")
    out_report = out_folder / "MyVASS_QualityReport.xlsx"
    qr = build_quality_report(stats, df)
    qr.save(out_report)
    print(f"  Saved: {out_report}")

    # Save cleaned CSV
    print("\nSaving cleaned CSV ...")
    out_csv = out_folder / "MyVASS_Cleaned.csv"
    save_cleaned_csv(df, out_csv)

    print()
    print("=" * 60)
    print("  Done! Files saved to:")
    print(f"    {out_report}")
    print(f"    {out_csv}")
    print("=" * 60)

    root = tk.Tk(); root.withdraw()
    messagebox.showinfo(
        "Done",
        f"MyVASS cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n"
        f"• MyVASS_Cleaned.csv\n"
        f"• MyVASS_QualityReport.xlsx",
    )
    root.destroy()


if __name__ == "__main__":
    main()
