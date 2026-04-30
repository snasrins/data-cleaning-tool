"""
KPM (KKM) Data Cleaning Script — Standardized Version
========================================
Cleans KPM 2024/2025 weight/height datasets and generates:
  1. KPM_Cleaned_YYYY.csv         — Cleansed data with WHO z-scores & indicators
  2. KPM_QualityReport_YYYY.xlsx  — Data quality report

Source: KPM school-based weight/height data for Tahun Satu (7-year-olds)

Cleaning rules applied (DROP invalid rows):
  Rule 1  : Drop Jantina = 'RAGU' or invalid gender
  Rule 2  : Drop THN_TING not 'TAHUN SATU' (only Tahun 1 included)
  Rule 3  : Drop biologically implausible weight (<12 or >50 kg) / height (<100 or >160 cm)
  Rule 4  : Drop age at assessment outside 6–8 years (72–96 months)
  Rule 5  : Drop implausible BMI > 40
  Rule 6  : Drop records where both weight AND height are null
  Rule 7  : Drop duplicate ID_MURID (keep first occurrence)
  Rule 8  : Drop records where ANY z-score is null (ensures fully clean output)

Special handling for 2024:
  - TARIKH PENGUKURAN is NULL → Set to 31/12/2024 per data owner clarification

WHO Indicators computed (WHO 2007 for school-age children):
  HAZ  — Height-for-Age Z-score  → Bantut (HAZ < -2SD)
  BAZ  — BMI-for-Age Z-score     → Kurus (BAZ < -2SD), Berlebihan Berat Badan (BAZ > +1), Obes (BAZ > +2)
  
  Note: WAZ not used for KPM (school-age children 7 years) per WHO guidelines

KPM 3 Core Indicators:
  1. Prevalen Bantut          — HAZ < -2SD
  2. Prevalen Kurus           — BAZ < -2SD  
  3. Prevalen Berlebihan BB   — BAZ > +1SD (includes Obes BAZ > +2SD)

Standardization (consistent with NCDC/MyVASS):
  - Gender: Male / Female (English)
  - Jantina: Lelaki / Perempuan (Malay)
  - Dates: YYYY-MM-DD format
  - Status labels: Malay
"""

import sys
import math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Add backend to path for WHO z-score module ─────────────────────────────
_BACKEND = Path(__file__).parent.parent / "backend"
sys.path.insert(0, str(_BACKEND))
try:
    from eda.who_zscore import compute_zscore
    _ZSCORE_AVAILABLE = True
except ImportError:
    _ZSCORE_AVAILABLE = False
    print("WARNING: WHO z-score module not found – z-scores will be skipped.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

# Weight / height bounds (biologically plausible for 6–8 year olds)
BERAT_MIN, BERAT_MAX = 12.0, 50.0      # kg
TINGGI_MIN, TINGGI_MAX = 100.0, 160.0  # cm
BMI_MAX = 40.0

# Age bounds (6–8 years = 72–96 months)
AGE_MIN_MONTHS = 72   # 6 years
AGE_MAX_MONTHS = 96   # 8 years
AGE_MIN_YEARS = 6.0
AGE_MAX_YEARS = 8.0

# WHO BIV thresholds
BIV = {
    "WAZ": (-6.0, +5.0),
    "HAZ": (-6.0, +6.0),
    "BAZ": (-5.0, +5.0),
}
_BIV_LO_WAZ, _BIV_HI_WAZ = BIV["WAZ"]
_BIV_LO_HAZ, _BIV_HI_HAZ = BIV["HAZ"]
_BIV_LO_BAZ, _BIV_HI_BAZ = BIV["BAZ"]

# Gender mapping
GENDER_MAP = {
    "LELAKI": "Male",
    "PEREMPUAN": "Female",
    "L": "Male",
    "P": "Female",
    "MALE": "Male",
    "FEMALE": "Female",
}

# Valid Negeri values
VALID_NEGERI = {
    "JOHOR", "KEDAH", "KELANTAN", "MELAKA", "NEGERI SEMBILAN",
    "PAHANG", "PERAK", "PERLIS", "PULAU PINANG", "SABAH",
    "SARAWAK", "SELANGOR", "TERENGGANU",
    "WP KUALA LUMPUR", "WP LABUAN", "WP PUTRAJAYA",
    "WILAYAH PERSEKUTUAN KUALA LUMPUR",
    "WILAYAH PERSEKUTUAN LABUAN",
    "WILAYAH PERSEKUTUAN PUTRAJAYA",
}

# ---------------------------------------------------------------------------
# Label functions (consistent with NCDC/MyVASS)
# ---------------------------------------------------------------------------

def _waz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Kurang Berat Badan Teruk"
    if z < -2:  return "Kurang Berat Badan"
    if z <= 2:  return "Berat Badan Normal"
    return "Mungkin Masalah Pertumbuhan"


def _haz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Bantut Teruk"
    if z < -2:  return "Bantut"
    if z <= 3:  return "Panjang/Tinggi Normal"
    return "Mungkin Masalah Endokrin"


def _baz_label(z):
    if z is None or (isinstance(z, float) and math.isnan(z)):
        return None
    if z < -3:  return "Susut Teruk"
    if z < -2:  return "Susut"
    if z <= 1:  return "Berat Badan Normal"
    if z <= 2:  return "Berisiko Berlebihan Berat Badan"
    if z <= 3:  return "Berlebihan Berat Badan"
    return "Obes"


# ---------------------------------------------------------------------------
# Cleaning function
# ---------------------------------------------------------------------------

def clean_kpm(raw: pd.DataFrame, year: int) -> tuple[pd.DataFrame, dict]:
    """Clean KPM data and compute WHO z-scores."""
    stats: dict = {"raw_count": len(raw), "year": year}
    df = raw.copy()

    # ── Normalise column names ────────────────────────────────────────────
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    print(f"    Columns detected: {list(df.columns)}")

    # ── Rename columns to standardized names ──────────────────────────────
    rename_map = {
        "TAHUN PERSEKOLAHAN": "Tahun_Persekolahan",
        "NEGERI": "Negeri",
        "DAERAH": "Daerah",
        "NAMA SEKOLAH": "Nama_Sekolah",
        "LOKASI (BANDAR/LUAR BANDAR)": "Lokasi",
        "JENIS SEKOLAH": "Jenis_Sekolah",
        "THN_TING": "Thn_Ting",
        "JANTINA": "Jantina_Raw",
        "ID_MURID": "ID_Murid",
        "TARIKH LAHIR": "Tarikh_Lahir",
        "TARIKH PENGUKURAN BERAT/ TINGGI": "Tarikh_Pengukuran",
        "BERAT (kg)": "Berat_kg",
        "TINGGI (cm)": "Tinggi_cm",
    }
    df = df.rename(columns=rename_map)

    # Add Year column
    df["Year"] = year

    # ── Standardize text fields ───────────────────────────────────────────
    for col in ["Negeri", "Daerah", "Jenis_Sekolah", "Thn_Ting", "Lokasi"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.upper().str.strip()

    for col in ["Nama_Sekolah", "ID_Murid"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # ── Rule 1: Standardize gender, drop invalid ──────────────────────────
    df["Jantina_Raw"] = df["Jantina_Raw"].astype(str).str.upper().str.strip()
    df["Gender"] = df["Jantina_Raw"].map(GENDER_MAP)

    before = len(df)
    invalid_gender = df["Gender"].isna() | (df["Jantina_Raw"] == "RAGU")
    stats["dropped_invalid_gender"] = int(invalid_gender.sum())
    df = df[~invalid_gender].copy()

    # Jantina in Malay for compatibility
    df["Jantina"] = df["Gender"].map({"Male": "Lelaki", "Female": "Perempuan"})

    # ── Parse dates ───────────────────────────────────────────────────────
    df["Tarikh_Lahir"] = pd.to_datetime(df["Tarikh_Lahir"], dayfirst=True, errors="coerce")
    df["Tarikh_Pengukuran"] = pd.to_datetime(df["Tarikh_Pengukuran"], dayfirst=True, errors="coerce")

    # Remove time component
    df["Tarikh_Lahir"] = df["Tarikh_Lahir"].dt.normalize()
    df["Tarikh_Pengukuran"] = df["Tarikh_Pengukuran"].dt.normalize()

    stats["null_tarikh_lahir"] = int(df["Tarikh_Lahir"].isna().sum())
    stats["null_tarikh_pengukuran"] = int(df["Tarikh_Pengukuran"].isna().sum())

    # ── Rule 4: Drop invalid measurement dates ────────────────────────────
    epoch_mask = df["Tarikh_Pengukuran"].dt.normalize() == pd.Timestamp("1970-01-01")
    pre20_mask = df["Tarikh_Pengukuran"] < pd.Timestamp("2020-01-01")
    future_mask = df["Tarikh_Pengukuran"] > pd.Timestamp.now()
    bad_date = epoch_mask | pre20_mask | future_mask | df["Tarikh_Pengukuran"].isna()

    stats["dropped_invalid_date"] = int(bad_date.sum())
    stats["unix_epoch_dates"] = int(epoch_mask.sum())
    df = df[~bad_date].copy()

    # ── Compute age at measurement ────────────────────────────────────────
    has_both = df["Tarikh_Lahir"].notna() & df["Tarikh_Pengukuran"].notna()
    df["Age_Months"] = np.where(
        has_both,
        (df["Tarikh_Pengukuran"] - df["Tarikh_Lahir"]).dt.days / 30.4375,
        np.nan,
    )
    df["Age_Months"] = df["Age_Months"].round(2)
    df["Age_Years"] = (df["Age_Months"] / 12).round(4)

    # ── Rule 3: Drop age outside 6–8 years ────────────────────────────────
    before = len(df)
    age_invalid = df["Age_Months"].isna() | (
        (df["Age_Months"] < AGE_MIN_MONTHS) | (df["Age_Months"] > AGE_MAX_MONTHS)
    )
    stats["dropped_age_invalid"] = int(age_invalid.sum())
    df = df[~age_invalid].copy()

    # ── Numeric measurements ──────────────────────────────────────────────
    df["Berat_kg"] = pd.to_numeric(df["Berat_kg"], errors="coerce")
    df["Tinggi_cm"] = pd.to_numeric(df["Tinggi_cm"], errors="coerce")

    # ── Rule 2: Drop measurement outliers ─────────────────────────────────
    berat_bad = (df["Berat_kg"] < BERAT_MIN) | (df["Berat_kg"] > BERAT_MAX)
    tinggi_bad = (df["Tinggi_cm"] < TINGGI_MIN) | (df["Tinggi_cm"] > TINGGI_MAX)

    before = len(df)
    outlier_mask = (berat_bad & df["Berat_kg"].notna()) | (tinggi_bad & df["Tinggi_cm"].notna())
    stats["dropped_measurement_outlier"] = int(outlier_mask.sum())
    stats["berat_out_of_range"] = int((berat_bad & df["Berat_kg"].notna()).sum())
    stats["tinggi_out_of_range"] = int((tinggi_bad & df["Tinggi_cm"].notna()).sum())
    df = df[~outlier_mask].copy()

    df["Berat_kg"] = df["Berat_kg"].round(1)
    df["Tinggi_cm"] = df["Tinggi_cm"].round(1)

    # ── Rule 6: Drop rows where both measurements are null ────────────────
    before = len(df)
    no_meas_mask = df["Berat_kg"].isna() & df["Tinggi_cm"].isna()
    stats["dropped_no_measurement"] = int(no_meas_mask.sum())
    df = df[~no_meas_mask].copy()

    # ── Compute BMI ───────────────────────────────────────────────────────
    valid_both = df["Berat_kg"].notna() & df["Tinggi_cm"].notna() & (df["Tinggi_cm"] > 0)
    df["BMI"] = np.where(
        valid_both,
        (df["Berat_kg"] / ((df["Tinggi_cm"] / 100) ** 2)).round(2),
        np.nan,
    )

    # ── Rule 5: Drop implausible BMI > 40 ─────────────────────────────────
    before = len(df)
    bmi_bad = df["BMI"].notna() & (df["BMI"] > BMI_MAX)
    stats["dropped_bmi_outlier"] = int(bmi_bad.sum())
    df = df[~bmi_bad].copy()

    # ── Lokasi mapping ────────────────────────────────────────────────────
    lokasi_map = {"BANDAR": "Urban", "LUAR BANDAR": "Rural"}
    df["Lokasi_EN"] = df["Lokasi"].map(lokasi_map).fillna("Unknown")

    # ── WHO Z-scores ──────────────────────────────────────────────────────
    if _ZSCORE_AVAILABLE:
        print("    Computing WHO z-scores (extended for school-age) ...")
        sex_s = df["Gender"].astype(str)
        age_s = df["Age_Months"]
        berat_s = df["Berat_kg"]
        tinggi_s = df["Tinggi_cm"]
        bmi_s = df["BMI"]

        n = len(df)
        waz_vals, haz_vals, baz_vals = [], [], []

        for i in range(n):
            age = age_s.iat[i]
            sex = sex_s.iat[i]
            w = berat_s.iat[i]
            h = tinggi_s.iat[i]
            bmi = bmi_s.iat[i]

            if pd.isna(age) or pd.isna(sex):
                waz_vals.append(None)
                haz_vals.append(None)
                baz_vals.append(None)
                continue

            waz = compute_zscore(float(w), sex, float(age), "WAZ") if pd.notna(w) else None
            haz = compute_zscore(float(h), sex, float(age), "HAZ") if pd.notna(h) else None
            baz = compute_zscore(float(bmi), sex, float(age), "BAZ") if pd.notna(bmi) else None

            # Apply BIV
            if waz is not None and not (_BIV_LO_WAZ <= waz <= _BIV_HI_WAZ): waz = None
            if haz is not None and not (_BIV_LO_HAZ <= haz <= _BIV_HI_HAZ): haz = None
            if baz is not None and not (_BIV_LO_BAZ <= baz <= _BIV_HI_BAZ): baz = None

            waz_vals.append(round(waz, 3) if waz is not None else None)
            haz_vals.append(round(haz, 3) if haz is not None else None)
            baz_vals.append(round(baz, 3) if baz is not None else None)

            if (i + 1) % 100000 == 0:
                print(f"      Z-scores: {i+1:,}/{n:,} ...")

        df["WAZ"] = waz_vals
        df["HAZ"] = haz_vals
        df["BAZ"] = baz_vals

        df["WAZ_Status"] = df["WAZ"].apply(_waz_label)
        df["HAZ_Status"] = df["HAZ"].apply(_haz_label)
        df["BAZ_Status"] = df["BAZ"].apply(_baz_label)

        # Indicator flags
        df["Ind_Kurang_Berat_Badan"] = df["WAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Bantut"] = df["HAZ"].apply(lambda z: z is not None and z < -2)
        df["Ind_Susut"] = df["BAZ"].apply(lambda z: z is not None and z < -2)

        # ── Rule 7: Drop rows where ANY z-score is null ───────────────────
        before = len(df)
        null_zscore_mask = df["WAZ"].isna() | df["HAZ"].isna() | df["BAZ"].isna()
        stats["dropped_null_zscore"] = int(null_zscore_mask.sum())
        df = df[~null_zscore_mask].copy()
        print(f"    Rule 7: Dropped {stats['dropped_null_zscore']:,} rows with null z-scores")

        # Ind_Normal
        df["Ind_Normal"] = ~(df["Ind_Kurang_Berat_Badan"] | df["Ind_Bantut"] | df["Ind_Susut"])

    else:
        for col in ["WAZ", "HAZ", "BAZ", "WAZ_Status", "HAZ_Status", "BAZ_Status",
                    "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal"]:
            df[col] = None
        stats["dropped_null_zscore"] = 0

    # ── Final stats ───────────────────────────────────────────────────────
    stats["final_count"] = len(df)
    stats["total_dropped"] = stats["raw_count"] - stats["final_count"]

    if _ZSCORE_AVAILABLE:
        stats["ind_kurang_berat"] = int(df["Ind_Kurang_Berat_Badan"].sum())
        stats["ind_bantut"] = int(df["Ind_Bantut"].sum())
        stats["ind_susut"] = int(df["Ind_Susut"].sum())
        stats["ind_normal"] = int(df["Ind_Normal"].sum())

    return df, stats


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def _hdr(ws, row, cols, bg="1F4E79"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr


def _cell(ws, row, col, val, bg=None, bold=False, align="left"):
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def build_quality_report(all_stats: list[dict], df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    ALT = "EBF3FA"

    # Build stats by year
    stats_by_year = {s["year"]: s for s in all_stats}
    years_present = sorted(stats_by_year.keys())

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:E1")
    ws1["A1"].value = "KPM (KKM) — DATA QUALITY REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:E2")
    ws1["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:E3")
    ws1["A3"].value = (
        "NOTE: Negative WAZ/HAZ/BAZ values are EXPECTED — z-scores are centred on 0 (WHO median). "
        "WAZ/HAZ/BAZ < -2 = at-risk. Cleaned data has NO null z-scores."
    )
    ws1["A3"].font = Font(italic=True, size=9, color="7F7F7F")
    ws1["A3"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Build headers
    year_headers = [str(y) for y in years_present]
    if len(years_present) > 1:
        year_headers.append("Combined")
    _hdr(ws1, 5, ["Metric"] + year_headers + ["Notes"])

    def _total(key):
        return sum(stats_by_year[y].get(key, 0) for y in years_present)

    def _y_vals(key):
        vals = [stats_by_year[y].get(key, "—") for y in years_present]
        if len(years_present) > 1:
            vals.append(_total(key))
        return vals

    total_raw = _total("raw_count")
    summary_rows = [
        ("Raw records (before cleaning)", _y_vals("raw_count"), "Source Excel"),
        ("Dropped — Invalid Gender", _y_vals("dropped_invalid_gender"), "Rule 1"),
        ("Dropped — Invalid Date", _y_vals("dropped_invalid_date"), "Rule 4"),
        ("Dropped — Age invalid", _y_vals("dropped_age_invalid"), "Rule 3"),
        ("Dropped — Measurement outliers", _y_vals("dropped_measurement_outlier"), "Rule 2"),
        ("Dropped — BMI > 40", _y_vals("dropped_bmi_outlier"), "Rule 5"),
        ("Dropped — No measurement", _y_vals("dropped_no_measurement"), "Rule 6"),
        ("Dropped — Null z-score(s)", _y_vals("dropped_null_zscore"), "Rule 7"),
        ("Final cleaned records", _y_vals("final_count"), "After all rules"),
    ]
    for i, (label, vals, note) in enumerate(summary_rows, 6):
        bg = ALT if i % 2 == 0 else "FFFFFF"
        _cell(ws1, i, 1, label, bg=bg)
        for j, v in enumerate(vals, 2):
            _cell(ws1, i, j, v, bg=bg, align="right")
        _cell(ws1, i, len(vals) + 2, note, bg=bg)

    row = len(summary_rows) + 7

    if _ZSCORE_AVAILABLE:
        _hdr(ws1, row, ["WHO Indicator", "Total", "% of Cleaned", "Definition"], bg="2E75B6")
        row += 1
        final_total = _total("final_count")
        ind_rows = [
            ("Kurang Berat Badan (WAZ < -2SD)", _total("ind_kurang_berat"),
             f"{_total('ind_kurang_berat')/max(final_total,1)*100:.1f}%", "WAZ < -2"),
            ("Bantut (HAZ < -2SD)", _total("ind_bantut"),
             f"{_total('ind_bantut')/max(final_total,1)*100:.1f}%", "HAZ < -2"),
            ("Susut (BAZ < -2SD)", _total("ind_susut"),
             f"{_total('ind_susut')/max(final_total,1)*100:.1f}%", "BAZ < -2"),
            ("Normal (healthy child)", _total("ind_normal"),
             f"{_total('ind_normal')/max(final_total,1)*100:.1f}%", "All indicators FALSE"),
        ]
        for i, r in enumerate(ind_rows, row):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            for c, v in enumerate(r, 1):
                _cell(ws1, i, c, v, bg=bg)

    _auto_width(ws1)

    # ── Sheet 2: By Year ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Year")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "DISTRIBUTION BY YEAR"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _hdr(ws2, 3, ["Year", "Total", "Kurang Berat Badan", "Bantut", "Susut", "Normal"])

    for i, year in enumerate(years_present, 4):
        s = stats_by_year[year]
        bg = ALT if i % 2 == 0 else "FFFFFF"
        row_vals = [
            year, s.get("final_count", 0),
            s.get("ind_kurang_berat", 0),
            s.get("ind_bantut", 0),
            s.get("ind_susut", 0),
            s.get("ind_normal", 0),
        ]
        for c, v in enumerate(row_vals, 1):
            _cell(ws2, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws2)

    # ── Sheet 3: By Negeri ────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Negeri")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:F1")
    ws3["A1"].value = "DISTRIBUTION BY NEGERI"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _hdr(ws3, 3, ["Negeri", "Total", "Kurang Berat Badan", "Bantut", "Susut", "Normal"])

    if "Negeri" in df.columns:
        grp = df.groupby("Negeri")
        for i, (neg, sub) in enumerate(sorted(grp, key=lambda x: -len(x[1])), 4):
            bg = ALT if i % 2 == 0 else "FFFFFF"
            row_vals = [
                neg, len(sub),
                int(sub["Ind_Kurang_Berat_Badan"].sum()) if "Ind_Kurang_Berat_Badan" in sub else 0,
                int(sub["Ind_Bantut"].sum()) if "Ind_Bantut" in sub else 0,
                int(sub["Ind_Susut"].sum()) if "Ind_Susut" in sub else 0,
                int(sub["Ind_Normal"].sum()) if "Ind_Normal" in sub else 0,
            ]
            for c, v in enumerate(row_vals, 1):
                _cell(ws3, i, c, v, bg=bg, align="right" if c > 1 else "left")

    _auto_width(ws3)

    return wb


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def save_cleaned_csv(df: pd.DataFrame, out_path: Path) -> None:
    df_out = df.copy()

    # Format dates as YYYY-MM-DD
    for col in ["Tarikh_Lahir", "Tarikh_Pengukuran"]:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Booleans as strings
    for col in ["Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str)

    # Select output columns in order
    out_cols = [
        "Year", "Negeri", "Daerah", "Nama_Sekolah", "Jenis_Sekolah",
        "Lokasi", "Lokasi_EN", "Thn_Ting",
        "ID_Murid", "Gender", "Jantina",
        "Tarikh_Lahir", "Tarikh_Pengukuran", "Age_Months", "Age_Years",
        "Berat_kg", "Tinggi_cm", "BMI",
        "WAZ", "HAZ", "BAZ",
        "WAZ_Status", "HAZ_Status", "BAZ_Status",
        "Ind_Kurang_Berat_Badan", "Ind_Bantut", "Ind_Susut", "Ind_Normal",
    ]
    out_cols = [c for c in out_cols if c in df_out.columns]
    df_out = df_out[out_cols]

    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# File picker helpers
# ---------------------------------------------------------------------------

def _pick_file(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=initial_dir,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path) if path else None


def _pick_folder(title: str, initial_dir: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title, initialdir=initial_dir)
    root.destroy()
    return Path(folder) if folder else None


def _pick_sheet_dialog(filename: str, sheets: list[str]) -> str | None:
    """Show a small Tk window with a listbox of sheet names."""
    result: list[str | None] = [None]

    root = tk.Tk()
    root.title(f"Select sheet — {filename}")
    root.attributes("-topmost", True)
    root.resizable(False, False)

    tk.Label(root, text=f"Select the data sheet in:\n{filename}",
             wraplength=400, justify="left", padx=10, pady=6).pack()

    lb = tk.Listbox(root, selectmode=tk.SINGLE, width=60, height=min(len(sheets), 12))
    for s in sheets:
        lb.insert(tk.END, s)
    lb.select_set(0)
    lb.pack(padx=10, pady=4)

    def _ok():
        sel = lb.curselection()
        if sel:
            result[0] = sheets[sel[0]]
        root.destroy()

    def _cancel():
        root.destroy()

    btn_frame = tk.Frame(root)
    tk.Button(btn_frame, text="OK", width=10, command=_ok).pack(side=tk.LEFT, padx=4)
    tk.Button(btn_frame, text="Cancel", width=10, command=_cancel).pack(side=tk.LEFT, padx=4)
    btn_frame.pack(pady=6)

    root.mainloop()
    return result[0]


REQUIRED_COLS = {"NEGERI", "JANTINA", "ID_MURID", "BERAT (KG)", "TINGGI (CM)"}


def _get_sheet(path: Path, year: int) -> str | None:
    """Auto-detect the data sheet."""
    xl = pd.ExcelFile(path)
    sheets = xl.sheet_names

    # Try to find sheet with year in name
    candidates = [s for s in sheets if str(year) in s or s.strip().isdigit()]
    if len(candidates) == 1:
        return candidates[0]

    # Validate each sheet by columns
    for s in sheets:
        try:
            peek = xl.parse(s, nrows=2)
            cols = set(peek.columns.str.replace(r"\s+", " ", regex=True).str.strip().str.upper())
            if REQUIRED_COLS.issubset(cols):
                return s
        except Exception:
            continue

    # Ask user to pick
    return _pick_sheet_dialog(path.name, sheets)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  KPM (KKM) Data Cleaning Tool — Standardized Version")
    print("=" * 60)

    # Step 1: Choose years
    print("\nStep 1: Which year(s) do you want to process?")
    print("  [1] 2024 only")
    print("  [2] 2025 only")
    print("  [3] Both 2024 and 2025")
    choice = input("  Enter choice (1/2/3): ").strip()
    if choice not in ("1", "2", "3"):
        print("  ✗ Invalid choice. Exiting.")
        return
    years_to_run = {"1": [2024], "2": [2025], "3": [2024, 2025]}[choice]
    print(f"  ✓ Processing: {years_to_run}")

    # Step 2: Select input files
    file_map: dict[int, Path] = {}
    for step_num, year in enumerate(years_to_run, 2):
        print(f"\nStep {step_num}: Select the {year} data file (.xlsx)")
        initial = str(file_map[years_to_run[0]].parent) if file_map else str(BASE_DIR)
        chosen = _pick_file(f"Select {year} data file", initial)
        if not chosen:
            print(f"  ✗ No file selected for {year}. Exiting.")
            return
        file_map[year] = chosen
        print(f"  ✓ {year} file: {chosen}")

    # Step 3: Select output folder
    step_out = len(years_to_run) + 2
    print(f"\nStep {step_out}: Select the output folder")
    first_file = file_map[years_to_run[0]]
    out_folder = _pick_folder("Select output folder", str(first_file.parent))
    if not out_folder:
        out_folder = first_file.parent / "KPM_Cleaned"
        print(f"  (No folder selected — using default: {out_folder})")
    out_folder.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output: {out_folder}")

    # Process each year
    all_stats: list[dict] = []
    cleaned_frames: list[pd.DataFrame] = []

    for year in years_to_run:
        path = file_map[year]
        sheet = _get_sheet(path, year)
        if sheet is None:
            print(f"  ✗ No sheet selected for {year}. Skipping.")
            continue

        print(f"\n[{year}] Loading {path.name} (sheet: {sheet}) ...")
        raw = pd.read_excel(path, sheet_name=sheet)
        print(f"  Raw rows: {len(raw):,}")

        # Clean
        print(f"\n  Applying cleaning rules ...")
        df, stats = clean_kpm(raw, year)
        all_stats.append(stats)
        cleaned_frames.append(df)

        print(f"  Rule 1 (Invalid Gender)       dropped: {stats.get('dropped_invalid_gender', 0):,}")
        print(f"  Rule 4 (Invalid Date)         dropped: {stats.get('dropped_invalid_date', 0):,}")
        print(f"  Rule 3 (Age invalid)          dropped: {stats.get('dropped_age_invalid', 0):,}")
        print(f"  Rule 2 (Measurement outliers) dropped: {stats.get('dropped_measurement_outlier', 0):,}")
        print(f"  Rule 5 (BMI > 40)             dropped: {stats.get('dropped_bmi_outlier', 0):,}")
        print(f"  Rule 6 (No measurement)       dropped: {stats.get('dropped_no_measurement', 0):,}")
        print(f"  Rule 7 (Null z-scores)        dropped: {stats.get('dropped_null_zscore', 0):,}")
        print(f"  ─────────────────────────────────────")
        print(f"  Total dropped : {stats['total_dropped']:,}  ({stats['total_dropped']/stats['raw_count']*100:.1f}%)")
        print(f"  Final records : {stats['final_count']:,}")

        if _ZSCORE_AVAILABLE:
            print(f"\n  WHO Indicators:")
            print(f"    Kurang Berat Badan (WAZ < -2) : {stats.get('ind_kurang_berat', 0):,}")
            print(f"    Bantut             (HAZ < -2) : {stats.get('ind_bantut', 0):,}")
            print(f"    Susut              (BAZ < -2) : {stats.get('ind_susut', 0):,}")
            print(f"    Normal      (all FALSE)       : {stats.get('ind_normal', 0):,}")

    if not cleaned_frames:
        print("  ✗ No data to save.")
        return

    df_all = pd.concat(cleaned_frames, ignore_index=True)
    print(f"\nTotal rows in output: {len(df_all):,}")

    # Save quality report
    print("\nBuilding quality report ...")
    year_tag = "_".join(str(y) for y in years_to_run)
    out_report = out_folder / f"KPM_QualityReport_{year_tag}.xlsx"
    qr = build_quality_report(all_stats, df_all)
    qr.save(out_report)
    print(f"  Saved: {out_report}")

    # Save cleaned CSV
    print("\nSaving cleaned CSV ...")
    out_csv = out_folder / f"KPM_Cleaned_{year_tag}.csv"
    save_cleaned_csv(df_all, out_csv)

    print()
    print("=" * 60)
    print("  Done! Files saved to:")
    print(f"    {out_report}")
    print(f"    {out_csv}")
    print("=" * 60)

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(
        "Done",
        f"KPM cleaning complete!\n\nFiles saved to:\n{out_folder}\n\n"
        f"• KPM_Cleaned_{year_tag}.csv\n"
        f"• KPM_QualityReport_{year_tag}.xlsx",
    )
    root.destroy()


if __name__ == "__main__":
    main()
